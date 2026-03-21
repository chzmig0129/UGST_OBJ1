from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from flask_login import LoginManager, login_required, current_user
from flask_talisman import Talisman
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import os
import pandas as pd
import numpy as np
import re
import json
from werkzeug.utils import secure_filename
from flask import jsonify
import sqlite3
from datetime import datetime
from shapely.geometry import Polygon, Point
from geopy.distance import geodesic
import geopandas as gpd
import shapefile
import tempfile
import zipfile
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from shapefile_utils import plot_shapefile_to_png
from utils.pdf_generator import (generar_ficha_tecnica_desde_plantilla, verificar_instalacion_pymupdf, 
                                generar_ficha_tecnica_fallback, generar_ficha_tecnica_simple, 
                                garantizar_pymupdf, import_pymupdf)
from utils.shapefile_cache import shp_cache
import shutil
import math
import threading

app = Flask(__name__)

# Load configuration from the appropriate config class.
# Set FLASK_ENV or FLASK_CONFIG to 'production' / 'testing' to switch.
from config import get_config  # noqa: E402
app.config.from_object(get_config())

# SECRET_KEY must be set on app.secret_key as well for Flask session signing.
app.secret_key = app.config['SECRET_KEY']

# Configure logging (must be done after app creation and config loading)
from utils.logging_config import setup_logging  # noqa: E402
setup_logging(app)

# ---------------------------------------------------------------------------
# Security: Flask-Talisman (security headers) — production only
# ---------------------------------------------------------------------------
# Only enforce HTTPS and strict headers when not in debug mode.
# CSP allows 'self' plus the external origins needed by Leaflet maps.
if not app.debug:
    _csp = {
        'default-src': ["'self'"],
        'script-src': ["'self'", "'unsafe-inline'", 'cdn.jsdelivr.net', 'unpkg.com'],
        'style-src': ["'self'", "'unsafe-inline'", 'cdn.jsdelivr.net', 'unpkg.com'],
        'img-src': [
            "'self'",
            'data:',
            'tile.openstreetmap.org',
            '*.tile.openstreetmap.org',
            'cdn.jsdelivr.net',
            'server.arcgisonline.com',
        ],
        'font-src': ["'self'", 'cdn.jsdelivr.net', 'unpkg.com'],
        'connect-src': ["'self'", 'tile.openstreetmap.org', '*.tile.openstreetmap.org', 'server.arcgisonline.com'],
    }
    Talisman(
        app,
        force_https=os.environ.get('FORCE_HTTPS', 'false').lower() == 'true',
        strict_transport_security=True,
        content_security_policy=_csp,
        session_cookie_secure=os.environ.get('FORCE_HTTPS', 'false').lower() == 'true',
        session_cookie_http_only=True,
    )

# ---------------------------------------------------------------------------
# Security: Flask-Limiter (rate limiting)
# ---------------------------------------------------------------------------
# Uses in-memory storage (compatible with preload_app / single-worker setups).
# Default limits: 200 requests/day, 50 requests/hour.
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://",
)

# Añadir filtro personalizado para slice
@app.template_filter('slice')
def slice_filter(iterable, start, end=None):
    if iterable is None or len(iterable) == 0:
        return []
    if end is None:
        return iterable[start:]
    return iterable[start:end]



def obtener_estatus_validacion(row):
    """Obtiene el estatus de validación priorizando ESTA_CHA."""
    for field_name in ('ESTA_CHA', 'ESTATUS', 'Estatus'):
        estatus = row.get(field_name)
        if pd.notna(estatus):
            estatus_str = str(estatus).strip()
            if estatus_str:
                return estatus_str
    return ''


def normalizar_columna_estatus_validacion(gdf):
    """Garantiza compatibilidad de estatus 15K usando ESTA_CHA como nombre canonico."""
    if gdf is None:
        return gdf

    if 'ESTA_CHA' in gdf.columns:
        estatus_base = gdf['ESTA_CHA']
    elif 'ESTATUS' in gdf.columns:
        estatus_base = gdf['ESTATUS']
    elif 'Estatus' in gdf.columns:
        estatus_base = gdf['Estatus']
    else:
        return gdf

    gdf['ESTA_CHA'] = estatus_base
    gdf['ESTATUS'] = estatus_base

    return gdf

# Cache para el dashboard de estatus (se computa una vez al primer request)
_dashboard_cache = None
_indices_filtrados_cache = None
_clasif_nuevos_state = {'status': 'idle', 'progress': 0, 'processed': 0, 'total': 0, 'result': None, 'indices_por_clasif': None, 'error': None}
_nuevos_relacionados_cache = None


# Función para obtener municipio y estado desde coordenadas
def obtener_ubicacion(lat, lon):
    if shp_cache.municipios is None:
        return None
    try:
        punto = Point(lon, lat)  # Shapely usa (x=lon, y=lat)
        mask = shp_cache.municipios.contains(punto)
        resultados = shp_cache.municipios[mask]
        if not resultados.empty:
            # Corregir la codificación de caracteres
            municipio = resultados.iloc[0]["NOMGEO"]
            estado = resultados.iloc[0]["NOM_ENT"]
            
            # Intentar corregir la codificación si es necesario
            try:
                # Si los nombres están en Latin-1 pero interpretados como UTF-8
                if isinstance(municipio, str) and any(c in municipio for c in ['Ã', 'Â', 'Á', 'É', 'Í', 'Ó', 'Ú']):
                    municipio = municipio.encode('latin-1').decode('utf-8')
                if isinstance(estado, str) and any(c in estado for c in ['Ã', 'Â', 'Á', 'É', 'Í', 'Ó', 'Ú']):
                    estado = estado.encode('latin-1').decode('utf-8')
            except Exception as encoding_error:
                app.logger.error(f"Error al corregir codificación: {encoding_error}")
                
            return {
                "municipio": municipio,
                "estado": estado
            }
    except Exception as e:
        app.logger.error(f"Error al obtener ubicación: {e}")
    return None

# db and bcrypt are defined in extensions.py and initialised here via init_app.
# This is the standard Flask pattern that avoids circular imports: any module
# (auth.py, models/user.py, scripts/…) can safely import db/bcrypt from
# extensions without pulling in the full app object.
from extensions import db, bcrypt  # noqa: E402
db.init_app(app)
bcrypt.init_app(app)

# Flask-Login setup
login_manager = LoginManager(app)
login_manager.login_view = 'auth.login'
login_manager.login_message = 'Por favor inicia sesión para acceder a esta página.'
login_manager.login_message_category = 'warning'

# User model — import after extensions are wired up so db is ready.
from models.user import User  # noqa: E402

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Definición del modelo para la base de datos
class Poligono(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    # Columnas específicas mapeadas desde el Excel/Lista
    id_poligono = db.Column(db.Text, nullable=True)
    if_val = db.Column(db.Text, nullable=True) # 'if' es palabra reservada
    id_credito = db.Column(db.Text, nullable=True)
    id_persona = db.Column(db.Text, nullable=True)
    superficie = db.Column(db.Float, nullable=True) # Asumiendo numérico
    estado = db.Column(db.Text, nullable=True)
    municipio = db.Column(db.Text, nullable=True)
    coordenadas = db.Column(db.Text, nullable=True) # Coordenadas originales
    coordenadas_corregidas = db.Column(db.Text, nullable=True) # Coordenadas decimales corregidas
    area_digitalizada = db.Column(db.Float, nullable=True) # Área calculada/editada
    estatus = db.Column(db.Text, nullable=True) # Estatus (si existe)
    comentarios = db.Column(db.Text, nullable=True) # Comentarios editables
    descripcion = db.Column(db.Text, nullable=True) # Nueva columna para descripción
    orden = db.Column(db.Text, nullable=True) # Nueva columna para número de orden
    se_modifico = db.Column(db.Text, default='No') # Campo para indicar si se modificó el polígono en el mapa
    # Metadata
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    fecha_modificacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# Variable global para almacenar los datos del Excel (mantener por compatibilidad)
excel_data = {
    'data': [],
    'columns': [],
    'filename': '',
    'original_coords': []  # Nuevo: almacen coordenadas originales
}

# Asegurar que exista el directorio de uploads
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Register auth blueprint
from auth import auth as auth_blueprint  # noqa: E402
app.logger.debug(auth_blueprint)
app.register_blueprint(auth_blueprint)

# Apply per-route rate limits to auth blueprint views.
# After blueprint registration, view functions are keyed as 'blueprint_name.func_name'
# in app.view_functions (not in auth_blueprint.view_functions).
limiter.limit("5 per minute")(app.view_functions['auth.login'])
limiter.limit("3 per hour")(app.view_functions['auth.register'])

# Crear tablas que no existan todavía — nunca elimina datos existentes.
with app.app_context():
    try:
        db.create_all()
        app.logger.info("Base de datos inicializada (db.create_all completado).")
    except Exception as e:
        app.logger.error(f"Error al inicializar la base de datos: {e}")

# ==============================================
# Funciones para procesamiento de coordenadas
# ==============================================

def limpiar_coordenada(coord):
    coord = coord.replace('\t', '').replace('"', '').strip()
    coord = re.sub(' +', ' ', coord)
    return coord

def corregir_longitud(coord_decimales):
    if pd.isna(coord_decimales) or coord_decimales == '':
        return coord_decimales
        
    coord_list = coord_decimales.split(' | ')
    corrected_coords = []
    for coord in coord_list:
        if ',' not in coord:
            continue
        lat, lon = coord.split(',')
        try:
            lat = float(lat.strip())
            lon = float(lon.strip())

            if lon > 0:
                lon *= -1

            corrected_coords.append(f"{lat:.6f},{lon:.6f}")  # Más precisión
        except:
            continue

    return ' | '.join(corrected_coords)

def dms_a_decimal(coord):
    try:
        # Primero identificar la dirección (N, S, E, W)
        match_dir = re.search(r'([NSEW])$', coord.strip(), re.IGNORECASE)
        direccion = match_dir.group(1).upper() if match_dir else ''
        
        # Caso especial: formato compacto tipo 18°4811.1N (sin separadores entre minutos y segundos)
        special_match = re.match(r'(\d+)[°\s](\d{2})(\d{2}\.\d+)([NSEW])', coord)
        if special_match:
            grados = float(special_match.group(1))
            minutos = float(special_match.group(2))
            segundos = float(special_match.group(3))
            direccion = special_match.group(4).upper()
            
            decimal = grados + minutos/60 + segundos/3600
            if direccion in ['S', 'W']:
                decimal *= -1
            return round(decimal, 6)
            
        # Caso normal: formato separado por símbolos tradicionales
        coord_num = re.sub(r'[^\d\.\-]', ' ', coord)
        parts = coord_num.strip().split()
        
        if len(parts) == 3:
            grados, minutos, segundos = map(float, parts)
        elif len(parts) == 2:
            grados, minutos = map(float, parts)
            segundos = 0.0
        elif len(parts) == 1:
            # Intento detectar formato compacto dentro de un solo número (ej: 184811.1)
            part = parts[0]
            if len(part) >= 4:  # Al menos debe tener grados (2) y minutos (2)
                try:
                    # Intenta interpretar como GGMMSS.S
                    if '.' in part:
                        dot_pos = part.index('.')
                        # Si hay suficientes dígitos antes del punto para grados(2) + minutos(2)
                        if dot_pos >= 4:
                            grados = float(part[:dot_pos-4])
                            minutos = float(part[dot_pos-4:dot_pos-2])
                            segundos = float(part[dot_pos-2:])
                        else:
                            grados = float(part[:2])
                            minutos = float(part[2:4])
                            segundos = float('0.' + part.split('.')[1])
                    else:
                        # Sin punto decimal, interpretar como GGMMSS
                        grados = float(part[:2])
                        minutos = float(part[2:4])
                        if len(part) > 4:
                            segundos = float(part[4:])
                        else:
                            segundos = 0.0
                    
                    decimal = grados + minutos/60 + segundos/3600
                    if direccion in ['S', 'W']:
                        decimal *= -1
                    return round(decimal, 6)
                except:
                    grados = float(part)
                    minutos = segundos = 0.0
            else:
                grados = float(part)
                minutos = segundos = 0.0
        else:
            return np.nan
            
        decimal = grados + minutos/60 + segundos/3600
        if direccion in ['S', 'W']:
            decimal *= -1
        return round(decimal, 6)  # Más precisión
    except Exception as e:
        app.logger.error(f"Error al convertir DMS a decimal: {coord} - {str(e)}")
        return np.nan

def es_dms(coord):
    # Verificar si tiene símbolos de grados, minutos o segundos
    if re.search('[°\'"]', coord):
        return True
    # Verificar si tiene dirección N, S, E, W
    if re.search(r'[NSEW]$', coord, re.IGNORECASE):
        return True
    # Verificar formato de números separados
    coord_num = re.sub(r'[^\d\.]', ' ', coord)
    parts = coord_num.strip().split()
    return len(parts) > 1

def procesar_coordenadas_dms(fila):
    if 'COORDENADAS' not in fila or pd.isna(fila['COORDENADAS']):
        return ''
    
    coordenadas = str(fila['COORDENADAS'])
    coordenadas = coordenadas.replace('\n', ' ').replace('\r', ' ').strip()
    
    # Dividir por múltiples posibles separadores
    for sep in ['|', ';', ' y ', ',y,']:
        if sep in coordenadas:
            coord_list = coordenadas.split(sep)
            break
    else:
        # Si no se encontró ningún separador común, intentar dividir por espacios
        if ' ' in coordenadas and ',' not in coordenadas:
            # Asumir que cada par de coordenadas está separado por espacios
            parts = coordenadas.split()
            if len(parts) % 2 == 0:  # Debe haber un número par de partes
                coord_list = []
                for i in range(0, len(parts), 2):
                    if i+1 < len(parts):
                        coord_list.append(f"{parts[i]} {parts[i+1]}")
            else:
                coord_list = [coordenadas]  # Un solo par de coordenadas
        else:
            coord_list = [coordenadas]  # Un solo par de coordenadas
    
    coord_list = [c.strip() for c in coord_list]
    
    # Depuración para ver las coordenadas procesadas
    app.logger.debug(f"Coordenadas divididas: {coord_list}")
    
    coords_decimales = []
    
    for coord_pair in coord_list:
        coord_pair = coord_pair.strip()
        if not coord_pair:
            continue
        
        # Casos especiales: coordenadas tipo 18°4811.1N,103°5102.7W
        special_match = re.match(r'(\d+[°\s]\d{2}\d{2}\.\d+[NSEW])[,\s]+(\d+[°\s]\d{2}\d{2}\.\d+[NSEW])', coord_pair)
        if special_match:
            lat_str = special_match.group(1)
            lon_str = special_match.group(2)
            try:
                lat = dms_a_decimal(lat_str)
                lon = dms_a_decimal(lon_str)
                if not np.isnan(lat) and not np.isnan(lon):
                    coords_decimales.append(f"{lat:.6f},{lon:.6f}")
                    app.logger.debug(f"Par procesado especial: {lat_str},{lon_str} -> {lat:.6f},{lon:.6f}")
                continue
            except Exception as e:
                app.logger.error(f"Error procesando formato especial {coord_pair}: {e}")
            
        # Procesamiento normal
        if ' ' in coord_pair and ',' not in coord_pair:
            parts = coord_pair.split()
            
            patterns = [
                r'([0-9\.]+[°][0-9\.]+[\'"][0-9\.]*[\"]*[NS])\s+([0-9\.]+[°][0-9\.]+[\'"][0-9\.]*[\"]*[WE])',
                r'([0-9\.]+\s+[0-9\.]+\s+[0-9\.]+\s*[NS])\s+([0-9\.]+\s+[0-9\.]+\s+[0-9\.]+\s*[WE])',
                r'([0-9\.]+\s+[0-9\.]+\s*[NS])\s+([0-9\.]+\s+[0-9\.]+\s*[WE])',
                r'([0-9\.]+\s*[NS])\s+([0-9\.]+\s*[WE])',
                # Formatos para 18°4811.1N
                r'(\d+[°\s]\d{2}\d{2}\.\d+[NS])\s+(\d+[°\s]\d{2}\d{2}\.\d+[WE])'
            ]
            
            lat_str = None
            lon_str = None
            
            for pattern in patterns:
                match = re.search(pattern, coord_pair)
                if match:
                    lat_str, lon_str = match.groups()
                    break
                    
            if lat_str is None or lon_str is None:
                lat_parts = [p for p in parts if 'N' in p.upper() or 'S' in p.upper()]
                lon_parts = [p for p in parts if 'W' in p.upper() or 'E' in p.upper()]
                
                if len(lat_parts) == 1 and len(lon_parts) == 1:
                    lat_str = lat_parts[0]
                    lon_str = lon_parts[0]
                elif len(parts) >= 2:
                    mid = len(parts) // 2
                    lat_str = ' '.join(parts[:mid])
                    lon_str = ' '.join(parts[mid:])
                else:
                    continue
                    
        elif ',' in coord_pair:
            try:
                lat_str, lon_str = coord_pair.split(',', 1)
            except:
                continue
        else:
            # Intentar interpretar como un formato especial sin espacios ni comas
            match = re.match(r'(\d+[°\s]\d+\.\d+[NS])(\d+[°\s]\d+\.\d+[WE])', coord_pair)
            if match:
                lat_str, lon_str = match.groups()
            elif re.search(r'[NS]', coord_pair, re.IGNORECASE) and re.search(r'[WE]', coord_pair, re.IGNORECASE):
                # Intentar encontrar donde termina la latitud (marcada por N o S) y empieza longitud
                ns_pos = max(coord_pair.upper().rfind('N'), coord_pair.upper().rfind('S'))
                if ns_pos > 0:
                    lat_str = coord_pair[:ns_pos+1]
                    lon_str = coord_pair[ns_pos+1:]
                else:
                    continue
            else:
                if re.search(r'[0-9]', coord_pair):
                    try:
                        coords_clean = re.sub(r'[^\d\.\-]', ' ', coord_pair)
                        nums = [float(x) for x in coords_clean.split() if x.strip()]
                        if len(nums) >= 2:
                            lat, lon = nums[0], nums[1]
                            if lon > 0 and lon > 90:
                                lon *= -1
                            coords_decimales.append(f"{lat:.6f},{lon:.6f}")
                    except Exception as e:
                        app.logger.error(f"Error procesando parte numérica {coord_pair}: {e}")
                continue

        # Limpieza adicional
        lat_str = limpiar_coordenada(lat_str) if lat_str else ''
        lon_str = limpiar_coordenada(lon_str) if lon_str else ''
        
        # Intentar procesarlas como DMS
        app.logger.debug(f"Procesando: lat_str={lat_str}, lon_str={lon_str}")
        
        try:
            # Proceso de latitud
            if es_dms(lat_str):
                lat = dms_a_decimal(lat_str)
                app.logger.debug(f"Latitud DMS: {lat_str} -> {lat}")
            else:
                lat_str_numeric = re.sub(r'[^\d\.\-]', '', lat_str)
                lat = float(lat_str_numeric)
                if 'S' in lat_str.upper():
                    lat *= -1
                app.logger.debug(f"Latitud decimal: {lat_str} -> {lat}")
                
            if np.isnan(lat):
                app.logger.error(f"Latitud inválida: {lat_str}")
                continue
        except Exception as e:
            app.logger.error(f"Error procesando latitud {lat_str}: {e}")
            continue

        try:
            # Proceso de longitud
            if es_dms(lon_str):
                lon = dms_a_decimal(lon_str)
                app.logger.debug(f"Longitud DMS: {lon_str} -> {lon}")
            else:
                lon_str_numeric = re.sub(r'[^\d\.\-]', '', lon_str)
                lon = float(lon_str_numeric)
                if 'W' in lon_str.upper():
                    lon *= -1
                elif lon > 0:
                    lon *= -1  # Asumir oeste para América
                app.logger.debug(f"Longitud decimal: {lon_str} -> {lon}")
                
            if np.isnan(lon):
                app.logger.error(f"Longitud inválida: {lon_str}")
                continue
        except Exception as e:
            app.logger.error(f"Error procesando longitud {lon_str}: {e}")
            continue

        if not np.isnan(lat) and not np.isnan(lon):
            coords_decimales.append(f"{lat:.6f},{lon:.6f}")
            app.logger.debug(f"Par añadido: {lat:.6f},{lon:.6f}")

    # Eliminar duplicados
    coords_decimales = list(dict.fromkeys(coords_decimales))
    return ' | '.join(coords_decimales)

def calcular_area_poligono(coordenadas_str):
    """Calcula el área de un polígono en hectáreas usando cálculo geodésico"""
    if not coordenadas_str:
        return 0.0
    
    try:
        from shapely.geometry import Polygon
        from geopy.distance import geodesic
        import numpy as np
        
        # Parsear coordenadas - Soportar tanto | como espacios como separadores
        points = []
        # Determinar si se usa | o espacios como separador
        separador = '|' if '|' in coordenadas_str else ' '
        
        for pair in coordenadas_str.split(separador):
            if not pair.strip():
                continue
            parts = pair.strip().split(',')
            if len(parts) >= 2:
                try:
                    lat, lon = map(float, parts[:2])
                    points.append((lat, lon))
                except (ValueError, TypeError):
                    # Ignorar coordenadas inválidas
                    continue
        
        if len(points) < 3:
            return 0.0
        
        # Implementación del algoritmo geodésico para calcular área
        # Basado en el cálculo que usa Leaflet.GeometryUtil.geodesicArea
        area = 0.0
        coords = np.array(points)
        
        if len(coords) > 2:
            p1 = coords[0]
            for i in range(1, len(coords) - 1):
                p2 = coords[i]
                p3 = coords[i + 1]
                
                # Cálculo del área del triángulo geodésico usando la fórmula del semiperímetro
                a = geodesic(p1, p2).meters
                b = geodesic(p2, p3).meters
                c = geodesic(p3, p1).meters
                s = (a + b + c) / 2.0
                
                # Fórmula de Herón (evitar números negativos bajo la raíz)
                area_factor = s * (s - a) * (s - b) * (s - c)
                if area_factor > 0:
                    area_triangulo = np.sqrt(area_factor)
                    area += area_triangulo
                else:
                    # Si el factor es negativo, usar un enfoque alternativo o 0
                    app.logger.debug(f"Factor de área negativo: {area_factor}")
        
        # Convertir a hectáreas (1 ha = 10,000 m²)
        return area / 10000.0
    except Exception as e:
        app.logger.error(f"Error al calcular área geodésica: {e}")
        
        # Fallback: usar shapely para cálculo plano si el geodésico falla
        try:
            from shapely.geometry import Polygon
            coords = []
            
            # Determinar si se usa | o espacios como separador
            separador = '|' if '|' in coordenadas_str else ' '
            
            for pair in coordenadas_str.split(separador):
                if not pair.strip():
                    continue
                parts = pair.strip().split(',')
                if len(parts) >= 2:
                    try:
                        lat, lon = map(float, parts[:2])
                        coords.append((lon, lat))  # Shapely usa (x,y) = (lon,lat)
                    except (ValueError, TypeError):
                        # Ignorar coordenadas inválidas
                        continue
            
            if len(coords) < 3:
                return 0.0
                
            try:
                polygon = Polygon(coords)
                if polygon.is_valid:
                    return polygon.area / 10000  # Convertir m² a hectáreas
                else:
                    app.logger.error("Polígono inválido, regresando área 0")
                    return 0.0
            except:
                app.logger.error("No se pudo crear polígono válido, regresando área 0")
                return 0.0
        except Exception as inner_e:
            app.logger.error(f"Error en fallback de cálculo de área: {inner_e}")
            return 0.0

# ==============================================
# Rutas de la aplicación
# ==============================================

@app.route('/health')
def health_check():
    status = {
        'status': 'ok',
        'database': False,
        'shapefiles': False,
        'timestamp': datetime.utcnow().isoformat()
    }

    # Check database connectivity
    try:
        db.session.execute(db.text('SELECT 1'))
        status['database'] = True
    except Exception:
        status['status'] = 'degraded'

    # Check if shapefiles are loaded
    try:
        status['shapefiles'] = shp_cache.municipios is not None
    except Exception:
        status['status'] = 'degraded'

    http_code = 200 if status['status'] == 'ok' else 503
    return jsonify(status), http_code


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/validacion-rapida')
@login_required
def validacion_rapida():
    return "Página de validación rápida en desarrollo"

@app.route('/unir-archivos')
@login_required
def unir_archivos():
    # Verificar si hay resultados en la sesión
    resultado = session.pop('resultado_shp', None)
    # Pasar la fecha y hora actual para los logs
    from datetime import datetime
    now = datetime.now()
    return render_template('unir_archivos.html', resultado=resultado, now=now)

@app.route('/validacion-poligonos', defaults={'tab': 'cargar'})
@app.route('/validacion-poligonos/<tab>')
@login_required
def validacion_poligonos(tab):
    valid_tabs = ['cargar', 'lista', 'editar', 'generar']
    
    if tab not in valid_tabs:
        tab = 'cargar'
    
    if tab == 'lista':
        try:
            # Obtener datos de la base de datos
            app.logger.info("Consultando polígonos en la base de datos...")
            poligonos = Poligono.query.all()
            app.logger.info(f"Se encontraron {len(poligonos)} polígonos en la base de datos")
            
            # Convertir a formato compatible con la plantilla (LEYENDO DIRECTO DE COLUMNAS)
            data = []
            for p in poligonos:
                # Crear diccionario directamente desde los atributos del objeto Poligono
                datos = {
                    'ID_POLIGONO': p.id_poligono,
                    'IF': p.if_val,
                    'ID_CREDITO': p.id_credito,
                    'ID_PERSONA': p.id_persona,
                    'SUPERFICIE': p.superficie,
                    'ESTADO': corregir_codificacion(p.estado),
                    'MUNICIPIO': corregir_codificacion(p.municipio),
                    'COORDENADAS': p.coordenadas,
                    'COORDENADAS_DECIMALES_CORREGIDAS': p.coordenadas_corregidas,  # Cambiado para coincidir con el template
                    'AREA_DIGITALIZADA': p.area_digitalizada,
                    'ESTATUS': p.estatus,
                    'COMENTARIOS': p.comentarios,
                    'DESCRIPCION': p.descripcion,
                    'ORDEN': p.orden,
                    'db_id': p.id
                }
                # Ya no es necesario cargar JSON ni usar setdefault,
                # los atributos no presentes en BD serán None por defecto.
                data.append(datos)
            
            # --- Definir columnas fijas para la vista de lista ---
            columns_to_display = [
                'ID_POLIGONO', 'IF', 'ID_CREDITO', 'ID_PERSONA', 'SUPERFICIE',
                'ESTADO', 'MUNICIPIO', 'COORDENADAS', 'COORDENADAS_DECIMALES_CORREGIDAS',
                'AREA_DIGITALIZADA', 'ESTATUS', 'COMENTARIOS', 'DESCRIPCION', 'ORDEN', 'db_id'
            ]
            # --- FIN: Definir columnas fijas ---

            app.logger.info(f"Mostrando {len(columns_to_display)} columnas fijas: {columns_to_display}")

            return render_template('validacion_poligonos.html',
                               tab=tab,
                               data=data,
                               columns=columns_to_display, # Usar la lista fija
                               filename=excel_data['filename']) # Mantener filename por compatibilidad
        except Exception as e:
            app.logger.error(f"ERROR AL CARGAR LISTA: {str(e)}")
            import traceback
            traceback.print_exc()
            flash(f'Error al cargar datos: {str(e)}', 'error')
            return render_template('validacion_poligonos.html', 
                               tab=tab, 
                               data=[],
                               columns=[],
                               filename='')
    
    elif tab == 'editar':
        db_id = request.args.get('db_id')
        if db_id:
            try:
                # Buscar el polígono en la base de datos por su ID
                poligono = Poligono.query.get(int(db_id))
                
                if poligono is None:
                    flash('Polígono no encontrado', 'error')
                    return redirect(url_for('validacion_poligonos', tab='lista'))
                
                # Preparar coordenadas para el mapa
                coords_para_mapa = []
                if poligono.coordenadas_corregidas:
                    try:
                        # Parsear las coordenadas corregidas para el mapa
                        coord_pairs = poligono.coordenadas_corregidas.split(' | ')
                        for pair in coord_pairs:
                            if ',' in pair:
                                lat, lon = pair.split(',')
                                coords_para_mapa.append([float(lat.strip()), float(lon.strip())])
                    except Exception as e:
                        app.logger.error(f"Error al procesar coordenadas para el mapa: {e}")
                        coords_para_mapa = []
                
                # Detectar ubicación automáticamente si el estado y municipio están vacíos
                ubicacion_auto = False
                estado_detectado = poligono.estado
                municipio_detectado = poligono.municipio
                
                if (not estado_detectado or not municipio_detectado) and poligono.coordenadas_corregidas:
                    app.logger.debug("Detectando ubicación automáticamente...")
                    ubicacion = obtener_ubicacion_desde_poligono(poligono.coordenadas_corregidas)
                    if ubicacion:
                        if not estado_detectado:
                            estado_detectado = ubicacion['estado']
                            ubicacion_auto = True
                        if not municipio_detectado:
                            municipio_detectado = ubicacion['municipio']
                            ubicacion_auto = True
                        app.logger.debug(f"Ubicación detectada: {municipio_detectado}, {estado_detectado}")
                
                # Crear diccionario con datos del polígono para la plantilla
                poligono_data = {
                    'ID_POLIGONO': poligono.id_poligono,
                    'IF': poligono.if_val,
                    'ID_CREDITO': poligono.id_credito,
                    'ID_PERSONA': poligono.id_persona,
                    'SUPERFICIE': poligono.superficie,
                    'ESTADO': corregir_codificacion(estado_detectado) or '',
                    'MUNICIPIO': corregir_codificacion(municipio_detectado) or '',
                    'COORDENADAS': poligono.coordenadas,
                    'COORDENADAS_DECIMALES_CORREGIDAS': poligono.coordenadas_corregidas,  # Cambiado para coincidir con el template
                    'AREA_DIGITALIZADA': poligono.area_digitalizada,
                    'ESTATUS': poligono.estatus,
                    'COMENTARIOS': poligono.comentarios,
                    'DESCRIPCION': poligono.descripcion,
                    'ORDEN': poligono.orden,
                    'db_id': poligono.id,
                    'UBICACION_AUTO': ubicacion_auto  # Bandera para mostrar que se detectó automáticamente
                }
                
                return render_template('validacion_poligonos.html', 
                                      tab=tab, 
                                      db_id=db_id,
                                      poligono_data=poligono_data, 
                                      coords_para_mapa=coords_para_mapa)
            except ValueError:
                flash('ID de polígono inválido', 'error')
                return redirect(url_for('validacion_poligonos', tab='lista'))
            except Exception as e:
                app.logger.error(f"Error al cargar polígono para edición: {e}")
                flash('Error al cargar el polígono para edición', 'error')
                return redirect(url_for('validacion_poligonos', tab='lista'))
        else:
            # Si no hay db_id, redirigir a la lista
            flash('No se especificó qué polígono editar', 'warning')
            return redirect(url_for('validacion_poligonos', tab='lista'))
    
    elif tab == 'generar':
        try:
            # Obtener datos de la base de datos para generar reportes
            poligonos = Poligono.query.all()
            
            # Convertir a formato compatible con la plantilla (LEYENDO DIRECTO DE COLUMNAS)
            data = []
            for p in poligonos:
                # Crear diccionario directamente desde los atributos del objeto Poligono
                datos = {
                    'ID_POLIGONO': p.id_poligono,
                    'IF': p.if_val,
                    'ID_CREDITO': p.id_credito,
                    'ID_PERSONA': p.id_persona,
                    'SUPERFICIE': p.superficie,
                    'ESTADO': corregir_codificacion(p.estado),
                    'MUNICIPIO': corregir_codificacion(p.municipio),
                    'COORDENADAS': p.coordenadas,
                    'COORDENADAS_DECIMALES_CORREGIDAS': p.coordenadas_corregidas,  # Cambiado para coincidir con el template
                    'AREA_DIGITALIZADA': p.area_digitalizada,
                    'ESTATUS': p.estatus,
                    'COMENTARIOS': p.comentarios,
                    'DESCRIPCION': p.descripcion,
                    'ORDEN': p.orden,
                    'db_id': p.id
                }
                data.append(datos)
            
            # Si no hay datos en la base de datos, usar datos en memoria (mantener por si acaso)
            if not data and excel_data.get('data'):
                data = excel_data['data']
                flash('Generando reporte con datos en memoria. No hay datos guardados en la base de datos.', 'warning')
            
            # Asegurar que haya datos para prevenir división por cero
            if not data:
                flash('No hay datos disponibles para generar reportes. Por favor, cargue un archivo primero.', 'warning')
                return redirect(url_for('validacion_poligonos', tab='cargar'))
            
            # Determinar columnas disponibles de manera segura
            all_columns = set()
            for row in data:
                if isinstance(row, dict):  # Asegurar que row sea un diccionario
                    all_columns.update(row.keys())
            
            columns = sorted(list(all_columns)) if all_columns else []
            
            return render_template('validacion_poligonos.html', 
                               tab=tab,
                               data=data,
                               columns=columns)
        except Exception as e:
            app.logger.error(f"ERROR AL GENERAR REPORTE: {str(e)}")
            import traceback
            traceback.print_exc()
            flash(f'Error al generar reporte: {str(e)}', 'error')
            return redirect(url_for('validacion_poligonos', tab='lista'))
    
    else:  # tab == 'cargar'
        columnas_ejemplo = [
            'ID_POLIGONO', 'ESTADO', 'AREA_REPORTADA', 'AREA_DIGITALIZADA',
            'COORDENADAS', 'MUNICIPIO', 'ID_CREDITO_FIRA', 'ID_PERSONA',
            'NOMBRE_IF', 'OBSERVACIONES', 'COMENTARIOS', 'CURP_PRODUCTOR', 'RFC'
        ]
        return render_template('validacion_poligonos.html', 
                           tab=tab,
                           columnas=columnas_ejemplo,
                           uploaded_columns=excel_data['columns'],
                           filename=excel_data['filename'])

@app.route('/cargar-excel', methods=['POST'])
@login_required
@limiter.limit("10 per hour")
def cargar_excel():
    global excel_data
    
    if 'archivo' not in request.files:
        flash('No se seleccionó ningún archivo', 'error')
        return redirect(url_for('validacion_poligonos'))
    
    archivo = request.files['archivo']
    
    if archivo.filename == '':
        flash('No se seleccionó ningún archivo', 'error')
        return redirect(url_for('validacion_poligonos'))
    
    if archivo and allowed_file(archivo.filename):
        try:
            # Guardar el archivo
            filename = secure_filename(archivo.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            archivo.save(filepath)
            
            # Leer el archivo Excel
            app.logger.info(f"Leyendo archivo Excel: {filename}")
            df = pd.read_excel(filepath)
            app.logger.info(f"Columnas encontradas en el Excel: {df.columns.tolist()}")
            
            # Normalizar nombres de columnas (eliminar espacios, convertir a mayúsculas)
            df.columns = [col.strip().upper().replace(' ', '_') for col in df.columns]
            app.logger.info(f"Columnas normalizadas: {df.columns.tolist()}")
            
            # --- INICIO: Validar columnas requeridas ---
            required_columns = {'IF', 'ID_CREDITO', 'ID_PERSONA', 'ID_POLIGONO', 'SUPERFICIE', 'COORDENADAS'}
            actual_columns = set(df.columns)
            
            if actual_columns != required_columns:
                missing_cols = required_columns - actual_columns
                extra_cols = actual_columns - required_columns
                error_parts = []
                if missing_cols:
                    error_parts.append(f"Faltan columnas: {', '.join(sorted(list(missing_cols)))}")
                if extra_cols:
                    error_parts.append(f"Hay columnas extra: {', '.join(sorted(list(extra_cols)))}")

                error_msg = f"El excel no sigue el formato. Favor de verificar el nombre de las columnas. Columnas requeridas: {', '.join(sorted(list(required_columns)))}. Detalles: {'. '.join(error_parts)}"
                flash(error_msg, 'error')
                return redirect(url_for('validacion_poligonos', tab='cargar'))
            # --- FIN: Validar columnas requeridas ---

            # Asegurar que exista la columna COORDENADAS (Esta validación ya está cubierta arriba, se podría quitar pero la dejamos por si acaso)
            # if 'COORDENADAS' not in df.columns:
            #     # Buscar una columna que pueda contener coordenadas (buscar patrones como 26°47'54"N)
            #     for col in df.columns:
            #         if df[col].dtype == 'object' and df[col].astype(str).str.contains('°|\'|"|N|W', regex=True).any():
            #             app.logger.debug(f"Se encontró columna con posibles coordenadas: {col}")
            #             df['COORDENADAS'] = df[col]
            #             break
            #     
            #     if 'COORDENADAS' not in df.columns:
            #         flash('El archivo debe contener una columna con coordenadas', 'error')
            #         return redirect(url_for('validacion_poligonos'))
            
            # Procesar coordenadas
            df['COORDENADAS_DECIMALES'] = df.apply(procesar_coordenadas_dms, axis=1)
            df['COORDENADAS_DECIMALES_CORREGIDAS'] = df['COORDENADAS_DECIMALES'].apply(corregir_longitud)
            
            # No calculamos el área aquí, la dejamos como None inicialmente
            # df['AREA_DIGITALIZADA'] = areas

            # Limpiar variable global excel_data ya que usaremos la BD
            excel_data = {
                'data': [],
                'columns': [],
                'filename': filename, # Guardamos el nombre del último archivo cargado
                'original_coords': []
            }
            
            # GUARDAR EN LA BASE DE DATOS
            try:
                app.logger.info("Intentando guardar datos en la base de datos...")
                # Primero limpiamos la tabla para evitar duplicaciones al cargar un nuevo archivo
                db.session.query(Poligono).delete()
                db.session.commit()
                app.logger.info(f"Tabla 'poligono' limpiada. Insertando {len(df)} registros...")
                
                count = 0
                for index, row in df.iterrows():
                    # Crear objeto Poligono mapeando columnas del DF a atributos del modelo
                    # Usar .get() para manejar columnas opcionales en el Excel
                    try:
                        superficie_val = float(row.get('SUPERFICIE', None)) if pd.notna(row.get('SUPERFICIE')) else None
                    except (ValueError, TypeError):
                        superficie_val = None

                    poligono = Poligono(
                        id_poligono=str(row.get('ID_POLIGONO', '')),
                        if_val=str(row.get('IF', '')), # Mapeado a if_val
                        id_credito=str(row.get('ID_CREDITO', '')),
                        id_persona=str(row.get('ID_PERSONA', '')),
                        superficie=superficie_val,
                        estado=str(row.get('ESTADO', '')), # Añadir si existe en Excel
                        municipio=str(row.get('MUNICIPIO', '')), # Añadir si existe en Excel
                        coordenadas=str(row.get('COORDENADAS', '')),
                        coordenadas_corregidas=str(row.get('COORDENADAS_DECIMALES_CORREGIDAS', '')), # Usar las corregidas
                        area_digitalizada=None, # Se inicializa como None
                        estatus=str(row.get('ESTATUS', '')), # Añadir si existe en Excel
                        comentarios=None,        # Se inicializa como None
                        descripcion=str(row.get('DESCRIPCION', ''))  # Añadir descripción
                        # datos_json ya no existe
                    )
                    db.session.add(poligono)
                    count += 1

                    # Commit por lotes
                    if count % 100 == 0:
                        db.session.commit()
                        app.logger.info(f"Guardados {count} registros...")
                
                # Commit final
                db.session.commit()
                app.logger.info(f"¡Guardados {count} registros en total en la base de datos!")
                flash(f'Archivo \'{filename}\' cargado y {count} registros guardados en la base de datos', 'success')
                
            except Exception as db_error:
                app.logger.error(f"ERROR AL GUARDAR EN LA BASE DE DATOS: {str(db_error)}")
                import traceback
                traceback.print_exc()
                flash(f'Error al guardar en la base de datos: {str(db_error)}', 'error')
                try:
                    db.session.rollback()
                except: pass
                # Redirigir a cargar si falla la BD
                return redirect(url_for('validacion_poligonos', tab='cargar'))
            
            # Redirigir a la lista después de guardar exitosamente
            return redirect(url_for('validacion_poligonos', tab='lista'))
            
        except Exception as e:
            flash(f'Error al procesar el archivo: {str(e)}', 'error')
            app.logger.error(f"ERROR GENERAL: {str(e)}")
            import traceback
            traceback.print_exc()
            return redirect(url_for('validacion_poligonos'))
    
    flash('Formato de archivo no permitido. Solo se aceptan .xlsx o .xls', 'error')
    return redirect(url_for('validacion_poligonos'))

@app.route('/actualizar-fila', methods=['POST'])
@login_required
def actualizar_fila():
    global excel_data
    
    row_index = request.form.get('row_index', type=int)
    db_id = request.form.get('db_id', type=int)
    
    # Imprimir información de la solicitud para depuración
    app.logger.debug(f"Actualizando fila - db_id: {db_id}, row_index: {row_index}")
    app.logger.debug(f"Datos del formulario: {request.form}")
    
    try:
        # Si tenemos db_id, actualizamos en la base de datos
        if db_id is not None:
            poligono = Poligono.query.get(db_id)
            if poligono is None:
                flash('Registro no encontrado en la base de datos', 'error')
                return redirect(url_for('validacion_poligonos', tab='lista'))
            
            app.logger.debug(f"Actualizando polígono en la base de datos con ID: {db_id}")

            # Cargar datos JSON actuales -> YA NO SE USA JSON
            # try:
            #     datos_actuales = json.loads(poligono.datos_json)
            # except:
            #     datos_actuales = {}

            # Actualizar campos directamente en el objeto Poligono
            for campo_form, valor_form in request.form.items():
                # Evitar campos especiales
                if campo_form in ['row_index', 'db_id']:
                    continue

                # Mapear nombre de campo del formulario (UPPERCASE) a atributo del modelo (lowercase)
                atributo_modelo = None
                if campo_form == 'ID_POLIGONO': atributo_modelo = 'id_poligono'
                elif campo_form == 'IF': atributo_modelo = 'if_val'
                elif campo_form == 'ID_CREDITO': atributo_modelo = 'id_credito'
                elif campo_form == 'ID_PERSONA': atributo_modelo = 'id_persona'
                elif campo_form == 'SUPERFICIE': atributo_modelo = 'superficie'
                elif campo_form == 'ESTADO': atributo_modelo = 'estado'
                elif campo_form == 'MUNICIPIO': atributo_modelo = 'municipio'
                # COORDENADAS originales no se editan aquí
                elif campo_form == 'COORDENADAS_DECIMALES_CORREGIDAS': atributo_modelo = 'coordenadas_corregidas'
                elif campo_form == 'AREA_DIGITALIZADA': atributo_modelo = 'area_digitalizada'
                elif campo_form == 'ESTATUS': atributo_modelo = 'estatus'
                elif campo_form == 'COMENTARIOS': atributo_modelo = 'comentarios'
                elif campo_form == 'DESCRIPCION': atributo_modelo = 'descripcion'
                elif campo_form == 'ORDEN': atributo_modelo = 'orden'
                # Añadir más mapeos si se agregan más campos editables

                if atributo_modelo:
                    try:
                        # Intentar convertir a float si es un campo numérico
                        if atributo_modelo in ['superficie', 'area_digitalizada']:
                            valor_actualizado = float(valor_form) if valor_form.strip() else None
                        else:
                            # Para campos de texto, usar None en lugar de strings vacíos o 'None'
                            valor_actualizado = valor_form.strip() if valor_form.strip() and valor_form.strip().lower() != 'none' else None
                        setattr(poligono, atributo_modelo, valor_actualizado)
                        app.logger.debug(f"Actualizado {atributo_modelo} a: {valor_actualizado}")
                    except ValueError:
                         app.logger.error(f"Error al convertir {campo_form} ('{valor_form}') a número para {atributo_modelo}. Se guarda como None/String.")
                         # Si falla la conversión numérica, decidir si guardar como None o string (depende del campo)
                         if atributo_modelo in ['superficie', 'area_digitalizada']:
                             setattr(poligono, atributo_modelo, None)
                         else: # Para campos de texto, guardar el valor original
                             setattr(poligono, atributo_modelo, valor_form)
                    except Exception as set_err:
                         app.logger.error(f"Error al actualizar {atributo_modelo}: {set_err}")

            # Guardar explícitamente el área digitalizada del formulario (redundante con el bucle, pero asegura tipo)
            # if 'AREA_DIGITALIZADA' in request.form and request.form['AREA_DIGITALIZADA'].strip():
            #     try:
            #         area_manual = float(request.form['AREA_DIGITALIZADA'])
            #         poligono.area_digitalizada = area_manual
            #         # datos_actuales['AREA_DIGITALIZADA'] = area_manual # No más JSON
            #         app.logger.debug(f"Usando área ingresada manualmente: {area_manual} hectáreas")
            #     except ValueError:
            #         poligono.area_digitalizada = None # Poner None si no es válido

            # Actualizar coordenadas (redundante con el bucle)
            # if 'COORDENADAS_DECIMALES_CORREGIDAS' in request.form:
            #     nuevas_coords = request.form['COORDENADAS_DECIMALES_CORREGIDAS']
            #     poligono.coordenadas_corregidas = nuevas_coords
            #     # datos_actuales['COORDENADAS_DECIMALES_CORREGIDAS'] = nuevas_coords # No más JSON

            # Guardar los datos actualizados como JSON -> YA NO SE USA JSON
            # poligono.datos_json = json.dumps(datos_actuales, ensure_ascii=False)

            # Actualizar fecha de modificación
            poligono.fecha_modificacion = datetime.utcnow()
            
            # Guardar cambios en la base de datos
            try:
                db.session.commit()
                app.logger.info("Cambios guardados exitosamente en la base de datos")
                flash('Cambios guardados correctamente en la base de datos', 'success')
                
                # Verificar si el usuario quiere ir al siguiente registro
                if 'guardar_y_siguiente' in request.form:
                    # Buscar el siguiente registro en la base de datos
                    siguiente_poligono = Poligono.query.filter(Poligono.id > db_id).order_by(Poligono.id.asc()).first()
                    
                    if siguiente_poligono:
                        flash('Pasando al siguiente registro...', 'info')
                        return redirect(url_for('validacion_poligonos', tab='editar', db_id=siguiente_poligono.id))
                    else:
                        flash('No hay más registros. Este era el último.', 'warning')
                        return redirect(url_for('validacion_poligonos', tab='lista'))
                
            except Exception as db_error:
                app.logger.error(f"Error al guardar en la base de datos: {db_error}")
                db.session.rollback()
                flash(f'Error al guardar en la base de datos: {str(db_error)}', 'error')
            
            return redirect(url_for('validacion_poligonos', tab='lista'))
        
        # Compatibilidad con el código anterior (mediante index)
        elif row_index is not None and row_index < len(excel_data['data']):
            app.logger.debug(f"Actualizando polígono en memoria con índice: {row_index}")
            
            # Actualizar todos los campos editables
            for col in excel_data['columns']:
                if col in request.form:
                    excel_data['data'][row_index][col] = request.form[col]
            
            # Guardar explícitamente el área digitalizada del formulario
            if 'AREA_DIGITALIZADA' in request.form and request.form['AREA_DIGITALIZADA'].strip():
                try:
                    area_manual = float(request.form['AREA_DIGITALIZADA'])
                    excel_data['data'][row_index]['AREA_DIGITALIZADA'] = area_manual
                    app.logger.debug(f"Usando área ingresada manualmente: {area_manual} hectáreas")
                except ValueError:
                    app.logger.debug("Valor de área digitalizada no válido")
            
            # Actualizar coordenadas si se proporcionaron
            if 'COORDENADAS_DECIMALES_CORREGIDAS' in request.form:
                excel_data['data'][row_index]['COORDENADAS_DECIMALES_CORREGIDAS'] = request.form['COORDENADAS_DECIMALES_CORREGIDAS']
                # Ya no recalculamos el área basada en coordenadas
            
            flash('Cambios guardados correctamente (modo memoria)', 'success')
            
            # Verificar si el usuario quiere ir al siguiente registro
            if 'guardar_y_siguiente' in request.form:
                # Buscar el siguiente registro en memoria
                siguiente_indice = row_index + 1
                if siguiente_indice < len(excel_data['data']):
                    siguiente_row = excel_data['data'][siguiente_indice]
                    siguiente_db_id = siguiente_row.get('db_id')
                    
                    if siguiente_db_id:
                        flash('Pasando al siguiente registro...', 'info')
                        return redirect(url_for('validacion_poligonos', tab='editar', db_id=siguiente_db_id))
                    else:
                        flash('Pasando al siguiente registro...', 'info')
                        return redirect(url_for('validacion_poligonos', tab='editar', row_index=siguiente_indice))
                else:
                    flash('No hay más registros. Este era el último.', 'warning')
                    return redirect(url_for('validacion_poligonos', tab='lista'))
            
            return redirect(url_for('validacion_poligonos', tab='lista'))
        
        else:
            flash('Índice de fila inválido', 'error')
            return redirect(url_for('validacion_poligonos', tab='lista'))
    
    except Exception as e:
        app.logger.error(f"ERROR GENERAL AL ACTUALIZAR: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error al actualizar: {str(e)}', 'error')
        return redirect(url_for('validacion_poligonos', tab='lista'))

@app.route('/get-original-coords/<int:row_index>')
@login_required
def get_original_coords(row_index):
    """Endpoint para obtener coordenadas originales (AJAX)"""
    # Intentar obtener el ID de la base de datos si está presente
    db_id = request.args.get('db_id', type=int)
    
    if db_id is not None:
        # Obtener de la base de datos
        poligono = Poligono.query.get(db_id)
        if poligono is None:
            return jsonify({'error': 'Registro no encontrado en la base de datos'}), 404
        
        return jsonify({
            'coordenadas': poligono.coordenadas_corregidas
        })
    elif row_index >= 0 and row_index < len(excel_data.get('original_coords', [])):
        # Obtener del almacenamiento en memoria (compatibilidad)
        return jsonify({
            'coordenadas': excel_data['original_coords'][row_index]
        })
    else:
        return jsonify({'error': 'Índice inválido'}), 404

@app.route('/marcar-como-modificado', methods=['POST'])
@login_required
def marcar_como_modificado():
    """Endpoint para marcar un polígono como modificado cuando se edita en el mapa"""
    try:
        data = request.get_json()
        db_id = data.get('db_id')
        
        if db_id is None:
            return jsonify({'error': 'ID de polígono no proporcionado'}), 400
        
        # Buscar el polígono en la base de datos
        poligono = Poligono.query.get(db_id)
        if poligono is None:
            return jsonify({'error': 'Polígono no encontrado'}), 404
        
        # Marcar como modificado
        poligono.se_modifico = 'Sí'
        poligono.fecha_modificacion = datetime.utcnow()
        
        # Guardar cambios
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Polígono marcado como modificado'})
        
    except Exception as e:
        app.logger.error(f"Error al marcar como modificado: {str(e)}")
        db.session.rollback()
        return jsonify({'error': f'Error al marcar como modificado: {str(e)}'}), 500

@app.route('/diagnostico-poligono/<int:db_id>')
@login_required
def diagnostico_poligono(db_id):
    """Endpoint para mostrar información de diagnóstico de un polígono"""
    poligono = Poligono.query.get(db_id)
    if poligono is None:
        return jsonify({'error': 'Registro no encontrado en la base de datos'}), 404
    
    # Devolver todos los datos del polígono para diagnosticar
    datos = {
        'id': poligono.id,
        'id_poligono': poligono.id_poligono,
        'if_val': poligono.if_val,
        'id_credito': poligono.id_credito,
        'id_persona': poligono.id_persona,
        'superficie': poligono.superficie,
        'estado': poligono.estado,
        'municipio': poligono.municipio,
        'coordenadas': poligono.coordenadas,
        'coordenadas_corregidas': poligono.coordenadas_corregidas,
        'area_digitalizada': poligono.area_digitalizada,
        'estatus': poligono.estatus,
        'comentarios': poligono.comentarios,
        'descripcion': poligono.descripcion,
        'orden': poligono.orden,
        'fecha_creacion': str(poligono.fecha_creacion),
        'fecha_modificacion': str(poligono.fecha_modificacion)
    }
    
    return jsonify(datos)

@app.route('/get-historico-poligonos')
@login_required
def get_historico_poligonos():
    """Endpoint para cargar y devolver los polígonos históricos como GeoJSON"""
    try:
        # Ruta al archivo shapefile histórico
        historico_shapefile = "data/MEGA_CAPA_V1_OL.shp"
        
        # Leer el shapefile con geopandas
        historico_gdf = gpd.read_file(historico_shapefile)
        
        # Verificar/convertir CRS a WGS84 (EPSG:4326) si es necesario
        if historico_gdf.crs != "EPSG:4326":
            historico_gdf = historico_gdf.to_crs(epsg=4326)
        
        # Convertir a GeoJSON
        geojson_data = json.loads(historico_gdf.to_json())
        
        # Asegurar que tenemos el campo ID_POLIGON (si existe)
        id_field = None
        orden_field = None
        for field in historico_gdf.columns:
            if field.upper() == 'ID_POLIGON':
                id_field = field
            elif field.upper() in ['ORDEN', 'ORDER']:
                orden_field = field
            
        # Agregar información sobre el campo de ID y orden para facilitar el etiquetado en el frontend
        respuesta = {
            'geojson': geojson_data,
            'id_field': id_field,
            'orden_field': orden_field
        }
        
        return jsonify(respuesta)
    except Exception as e:
        app.logger.error(f"Error al cargar el shapefile histórico: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/get-historico-poligonos-radio/<int:polygon_id>')
@login_required
def get_historico_poligonos_radio(polygon_id):
    """Endpoint para cargar y devolver los polígonos históricos dentro de un radio de 5km"""
    try:
        # Buscar el polígono en la base de datos
        poligono = Poligono.query.get(polygon_id)
        if poligono is None:
            return jsonify({'error': 'Polígono no encontrado'}), 404
            
        # Obtener coordenadas del polígono
        coordenadas_corregidas = poligono.coordenadas_corregidas
        if not coordenadas_corregidas:
            return jsonify({'error': 'El polígono no tiene coordenadas válidas'}), 400
            
        # Extraer el primer punto del polígono como punto de referencia
        coords_list = coordenadas_corregidas.split(' | ')
        if not coords_list:
            return jsonify({'error': 'Formato de coordenadas inválido'}), 400
            
        first_point = coords_list[0].split(',')
        if len(first_point) < 2:
            return jsonify({'error': 'Formato de coordenadas inválido'}), 400
            
        lat_ref = float(first_point[0])
        lon_ref = float(first_point[1])
        
        # Ruta al archivo shapefile histórico
        historico_shapefile = "data/MEGA_CAPA_V1_OL.shp"
        
        # Leer el shapefile con geopandas
        historico_gdf = gpd.read_file(historico_shapefile)
        
        # Verificar/convertir CRS a WGS84 (EPSG:4326) si es necesario
        if historico_gdf.crs != "EPSG:4326":
            historico_gdf = historico_gdf.to_crs(epsg=4326)
        
        # Filtrar polígonos en el radio de 5km
        from shapely.geometry import Point
        import math
        
        # Radio de la tierra en km
        R = 6371.0
        
        # Función para calcular distancia haversine
        def haversine(lat1, lon1, lat2, lon2):
            # Convertir de grados a radianes
            lat1 = math.radians(lat1)
            lon1 = math.radians(lon1)
            lat2 = math.radians(lat2)
            lon2 = math.radians(lon2)
            
            # Fórmula haversine
            dlon = lon2 - lon1
            dlat = lat2 - lat1
            a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
            c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
            distance = R * c
            
            return distance
        
        # Crear una función para aplicar a cada geometría
        def en_radio(geometry):
            # Para polígonos, usamos el centroide
            centroid = geometry.centroid
            lat = centroid.y
            lon = centroid.x
            
            # Calcular la distancia
            distancia = haversine(lat_ref, lon_ref, lat, lon)
            
            # Retornar True si está dentro del radio (5km)
            return distancia <= 5.0
        
        # Aplicar el filtro a todas las geometrías
        mask = historico_gdf.geometry.apply(en_radio)
        historico_filtrado = historico_gdf[mask]
        
        # Convertir a GeoJSON
        geojson_data = json.loads(historico_filtrado.to_json())
        
        # Asegurar que tenemos el campo ID_POLIGON y ORDEN (si existen)
        id_field = None
        orden_field = None
        for field in historico_filtrado.columns:
            if field.upper() == 'ID_POLIGON':
                id_field = field
            elif field.upper() in ['ORDEN', 'ORDER']:
                orden_field = field
            
        # Agregar información sobre el campo de ID, orden y contador
        respuesta = {
            'geojson': geojson_data,
            'id_field': id_field,
            'orden_field': orden_field,
            'total': len(historico_filtrado),
            'radio_km': 5.0
        }
        
        return jsonify(respuesta)
    except Exception as e:
        app.logger.error(f"Error al cargar el shapefile histórico filtrado: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/obtener_ubicacion', methods=['POST'])
@login_required
def get_ubicacion():
    """Endpoint para obtener municipio y estado desde coordenadas"""
    try:
        data = request.get_json()
        lat = float(data.get('lat'))
        lon = float(data.get('lon'))
        
        # Usar la función para obtener el municipio y estado
        ubicacion = obtener_ubicacion(lat, lon)
        
        if ubicacion:
            return jsonify(ubicacion)
        return jsonify({"error": "Ubicación no encontrada"}), 404
    except Exception as e:
        return jsonify({"error": f"Datos inválidos: {str(e)}"}), 400

# Función para obtener ubicación desde las coordenadas de un polígono
def obtener_ubicacion_desde_poligono(coordenadas_str):
    """Obtiene el municipio y estado desde las coordenadas de un polígono"""
    if not coordenadas_str:
        return None
    
    try:
        # Usar el primer punto del polígono para determinar ubicación
        coords_list = coordenadas_str.split(' | ')
        if not coords_list:
            return None
            
        first_point = coords_list[0].split(',')
        if len(first_point) < 2:
            return None
            
        lat = float(first_point[0])
        lon = float(first_point[1])
        
        ubicacion = obtener_ubicacion(lat, lon)
        if ubicacion:
            # Asegurar que los nombres tengan codificación correcta
            ubicacion['municipio'] = corregir_codificacion(ubicacion['municipio'])
            ubicacion['estado'] = corregir_codificacion(ubicacion['estado'])
        return ubicacion
    except Exception as e:
        app.logger.error(f"Error al obtener ubicación desde polígono: {e}")
        return None

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

@app.route('/generar_shapefiles', methods=['POST'])
@login_required
def generar_shapefiles():
    """Ruta para generar archivos shapefile de polígonos seleccionados"""
    # Obtener los índices de polígonos seleccionados
    selected_rows = request.json.get('selected_rows', [])
    
    if not selected_rows:
        return jsonify({'error': 'No se seleccionaron polígonos'}), 400
    
    try:
        # Preparar un archivo ZIP en memoria para contener todos los shapefiles
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w') as zf:
            # Para cada polígono seleccionado
            for row_id in selected_rows:
                # Buscar el polígono en la base de datos por su ID
                try:
                    row_id = int(row_id)
                    # Primero intentar buscar por ID exacto
                    poligono = Poligono.query.get(row_id)
                    
                    if poligono is None:
                        # Si no se encuentra, imprimir para depuración
                        app.logger.error(f"No se encontró polígono con ID {row_id}, buscando en posición")
                        
                        # Intentar buscar por posición como fallback
                        poligonos = Poligono.query.all()
                        if 0 <= row_id < len(poligonos):
                            poligono = poligonos[row_id]
                        else:
                            app.logger.debug(f"Índice {row_id} fuera de rango, hay {len(poligonos)} polígonos")
                            continue
                    
                    app.logger.debug(f"Generando shapefile para polígono ID={poligono.id}, ID_POLIGONO={poligono.id_poligono}")
                except Exception as e:
                    app.logger.error(f"Error al recuperar polígono {row_id}: {e}")
                    # Si no es un índice válido, continuar con el siguiente
                    continue
                
                # Generar shapefile para este polígono
                shapefile_buffer = generar_shapefile_individual(poligono, f'polygon-{row_id}')
                
                # Añadir el shapefile al archivo ZIP
                if shapefile_buffer:
                    zf.writestr(f'polygon-{row_id}.zip', shapefile_buffer.getvalue())
        
        # Regresar al inicio del archivo en memoria
        memory_file.seek(0)
        
        # Enviar el archivo ZIP como respuesta
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name='poligonos_shapefiles.zip'
        )
    
    except Exception as e:
        app.logger.error(f"Error al generar shapefiles: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/generar_paquete_completo', methods=['POST'])
@login_required
def generar_paquete_completo():
    """Ruta para generar un paquete completo con fichas PDF y shapefiles"""
    # Obtener los índices de polígonos seleccionados
    selected_rows = request.json.get('selected_rows', [])
    
    if not selected_rows:
        return jsonify({'error': 'No se seleccionaron polígonos'}), 400
    
    try:
        # Preparar un archivo ZIP en memoria para contener todos los archivos
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w') as zf:
            # Crear carpeta para fichas técnicas
            zf.writestr('fichas_tecnicas/', '')
            # Crear carpeta para shapefiles
            zf.writestr('shapefiles/', '')
            # Crear carpeta para mapas
            zf.writestr('mapas/', '')
            
            # Para cada polígono seleccionado
            for row_id in selected_rows:
                try:
                    row_id = int(row_id)
                    # Primero intentar buscar por ID exacto
                    poligono = Poligono.query.get(row_id)
                    
                    if poligono is None:
                        # Si no se encuentra, imprimir para depuración
                        app.logger.error(f"No se encontró polígono con ID {row_id}, buscando en posición")
                        
                        # Intentar buscar por posición como fallback
                        poligonos = Poligono.query.all()
                        if 0 <= row_id < len(poligonos):
                            poligono = poligonos[row_id]
                        else:
                            app.logger.debug(f"Índice {row_id} fuera de rango, hay {len(poligonos)} polígonos")
                            continue
                            
                    app.logger.debug(f"Generando fichas para polígono ID={poligono.id}, ID_POLIGONO={poligono.id_poligono}")
                except Exception as e:
                    app.logger.error(f"Error al recuperar polígono {row_id}: {e}")
                    # Si no es un ID válido, continuar con el siguiente
                    continue
                
                # Generar shapefile para este polígono
                shapefile_buffer = generar_shapefile_individual(poligono, f'polygon-{row_id}')
                png_filepath = None
                
                if shapefile_buffer:
                    # Usar ID_POLIGONO para nombrar el archivo si está disponible
                    archivo_nombre = poligono.id_poligono if poligono.id_poligono else f'polygon-{row_id}'
                    zf.writestr(f'shapefiles/{archivo_nombre}.zip', shapefile_buffer.getvalue())
                    
                    # Generar mapas PNG a partir del shapefile
                    try:
                        # Crear un directorio temporal para guardar los PNG
                        with tempfile.TemporaryDirectory() as temp_png_dir:
                            # Generar PNG a partir del shapefile
                            png_dir = plot_shapefile_to_png(shapefile_buffer, temp_png_dir)
                            
                            # Añadir todos los archivos PNG al ZIP y guardar la ruta del primer PNG para la ficha técnica
                            if png_dir:
                                for png_filename in os.listdir(png_dir):
                                    if png_filename.endswith('.png'):
                                        png_path = os.path.join(png_dir, png_filename)
                                        # Guardar la ruta del primer PNG para usarla en la ficha
                                        if png_filepath is None:
                                            png_filepath = png_path
                                        
                                        # Guardar la imagen en un archivo temporal más permanente que podamos usar para el PDF
                                        temp_img_path = tempfile.NamedTemporaryFile(delete=False, suffix='.png').name
                                        shutil.copy2(png_path, temp_img_path)
                                        png_filepath = temp_img_path
                                        
                                        with open(png_path, 'rb') as png_file:
                                            # Guardar con un nombre predecible basado en ID_POLIGONO
                                            png_name = f"{poligono.id_poligono or f'polygon-{row_id}'}.png"
                                            zf.writestr(f'mapas/{png_name}', png_file.read())
                    except Exception as e:
                        app.logger.error(f"Error al generar mapa PNG para polígono {row_id}: {e}")
                        error_msg = f"Error al generar mapa PNG para polígono {row_id}: {e}"
                        errores_detalles.append(error_msg)
                        errores += 1
                        import traceback
                        traceback.print_exc()
                
                # Generar ficha técnica PDF con la nueva plantilla
                if png_filepath:
                    pdf_buffer = generar_ficha_tecnica_desde_plantilla(poligono, png_filepath)
                    if pdf_buffer:
                        # Usar ID_POLIGONO para nombrar el archivo si está disponible
                        archivo_nombre = poligono.id_poligono if poligono.id_poligono else f'polygon-{row_id}'
                        zf.writestr(f'fichas_tecnicas/ficha_{archivo_nombre}.pdf', pdf_buffer.getvalue())
                    else:
                        # Si falla la generación con la plantilla, intentar el método original como respaldo
                        pdf_buffer = generar_ficha_tecnica(poligono, f'polygon-{row_id}')
                        if pdf_buffer:
                            # Usar ID_POLIGONO para nombrar el archivo si está disponible
                            archivo_nombre = poligono.id_poligono if poligono.id_poligono else f'polygon-{row_id}'
                            zf.writestr(f'fichas_tecnicas/ficha_{archivo_nombre}.pdf', pdf_buffer.getvalue())
                else:
                    # Si no hay imagen, usar el método tradicional
                    pdf_buffer = generar_ficha_tecnica(poligono, f'polygon-{row_id}')
                    if pdf_buffer:
                        # Usar ID_POLIGONO para nombrar el archivo si está disponible
                        archivo_nombre = poligono.id_poligono if poligono.id_poligono else f'polygon-{row_id}'
                        zf.writestr(f'fichas_tecnicas/ficha_{archivo_nombre}.pdf', pdf_buffer.getvalue())
        
        # Regresar al inicio del archivo en memoria
        memory_file.seek(0)
        
        # Enviar el archivo ZIP como respuesta
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name='paquete_completo.zip'
        )
    
    except Exception as e:
        app.logger.error(f"Error al generar paquete completo: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/generar_shapefile_unico', methods=['POST'])
@login_required
def generar_shapefile_unico():
    """Ruta para generar un único shapefile con todos los polígonos seleccionados"""
    try:
        data = request.get_json()
        selected_rows = data.get('selected_rows', [])
        
        if not selected_rows:
            return jsonify({'error': 'No se seleccionaron filas'}), 400
        
        # Obtener polígonos de la base de datos
        poligonos = []
        for row_id in selected_rows:
            try:
                row_id = int(row_id)
                poligono = Poligono.query.get(row_id)
                
                if poligono:
                    poligonos.append(poligono)
                else:
                    app.logger.debug(f"Polígono con ID {row_id} no encontrado")
                    
            except Exception as e:
                app.logger.error(f"Error al recuperar polígono {row_id}: {e}")
                continue
        
        if not poligonos:
            return jsonify({'error': 'No se encontraron polígonos válidos'}), 400
        
        # Generar shapefile único con todos los polígonos
        shapefile_bytes = generar_shapefile_unificado(poligonos, 'poligonos_unificados')
        
        if not shapefile_bytes:
            return jsonify({'error': 'Error al generar el shapefile'}), 500
        
        return send_file(
            shapefile_bytes,
            mimetype='application/zip',
            as_attachment=True,
            download_name='poligonos_unificados.zip'
        )
    
    except Exception as e:
        app.logger.error(f"Error general: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def generar_shapefile_individual(poligono, nombre_archivo):
    """Genera un archivo shapefile para un polígono individual"""
    try:
        # Crear un objeto de memoria para el archivo ZIP
        zip_buffer = io.BytesIO()
        
        # Crear un directorio temporal para los archivos del shapefile
        with tempfile.TemporaryDirectory() as tempdir:
            # Crear el writer de shapefile
            w = shapefile.Writer(os.path.join(tempdir, 'poligono'))
            
            # Definir campos de atributos
            w.field('ID_POLIG', 'C', 40)
            w.field('IF', 'C', 40)
            w.field('ID_CRED', 'C', 40)
            w.field('ID_PERS', 'C', 40)
            w.field('SUPERF', 'N', 10, 4)
            w.field('ESTADO', 'C', 40)
            w.field('MUNICIP', 'C', 40)
            w.field('AREA_HA', 'N', 10, 4)
            w.field('ESTATUS', 'C', 10)
            w.field('COMENT', 'C', 254)
            w.field('DESCRIP', 'C', 254)  # Campo para descripción
            w.field('ORDEN', 'C', 100)    # Campo para número de orden
            
            # Función auxiliar para limpiar campos de texto que pueden tener saltos de línea
            def limpiar_campo_texto(valor):
                """Limpia un campo de texto reemplazando saltos de línea con comas y eliminando caracteres problemáticos"""
                if valor is None:
                    return ''
                
                # Convertir a string si no lo es
                valor_str = str(valor)
                
                # Reemplazar saltos de línea con comas y espacios
                valor_limpio = valor_str.replace('\n', ', ').replace('\r\n', ', ').replace('\r', ', ')
                
                # Reemplazar múltiples comas seguidas con una sola coma
                import re
                valor_limpio = re.sub(r',\s*,+', ', ', valor_limpio)
                
                # Eliminar comas al inicio y al final
                valor_limpio = valor_limpio.strip(', ')
                
                # Reemplazar caracteres problemáticos que pueden causar errores en DBF
                valor_limpio = valor_limpio.replace('\t', ' ')  # Reemplazar tabs con espacios
                valor_limpio = re.sub(r'\s+', ' ', valor_limpio)  # Reemplazar múltiples espacios con uno solo
                
                return valor_limpio
            
            # Obtener coordenadas del polígono
            coords = []
            if poligono.coordenadas_corregidas:
                # Verificar qué separador usa: ' | ' o '|'
                if ' | ' in poligono.coordenadas_corregidas:
                    pares = poligono.coordenadas_corregidas.split(' | ')
                else:
                    pares = poligono.coordenadas_corregidas.split('|')
                
                for par in pares:
                    par = par.strip()
                    if par and ',' in par:
                        try:
                            partes = par.split(',')
                            lat = float(partes[0].strip())
                            lon = float(partes[1].strip())
                            coords.append([lon, lat])  # Shapefile usa [lon, lat]
                        except (ValueError, IndexError) as e:
                            app.logger.error(f"Error al procesar coordenada {par}: {e}")
                            continue
                
                app.logger.debug(f"Coordenadas procesadas para shapefile: {coords}")
            
            # Limpiar todos los campos de texto antes de escribirlos al DBF
            id_poligono_limpio = limpiar_campo_texto(poligono.id_poligono)[:40]
            if_val_limpio = limpiar_campo_texto(poligono.if_val)[:40]
            id_credito_limpio = limpiar_campo_texto(poligono.id_credito)[:40]
            id_persona_limpio = limpiar_campo_texto(poligono.id_persona)[:40]
            estado_limpio = limpiar_campo_texto(corregir_codificacion(poligono.estado))[:40]
            municipio_limpio = limpiar_campo_texto(corregir_codificacion(poligono.municipio))[:40]
            estatus_limpio = limpiar_campo_texto(poligono.estatus)[:10]
            comentarios_limpio = limpiar_campo_texto(poligono.comentarios)[:254]
            descripcion_limpio = limpiar_campo_texto(poligono.descripcion)[:254]
            orden_limpio = limpiar_campo_texto(poligono.orden)[:100]
            
            # Si no hay suficientes coordenadas, usar un punto
            if len(coords) < 3:
                if len(coords) == 1:
                    # Crear un punto
                    w.point(coords[0][0], coords[0][1])
                    w.record(
                        id_poligono_limpio,
                        if_val_limpio,
                        id_credito_limpio,
                        id_persona_limpio,
                        poligono.superficie or 0,
                        estado_limpio,
                        municipio_limpio,
                        poligono.area_digitalizada or 0,
                        estatus_limpio,
                        comentarios_limpio,
                        descripcion_limpio,
                        orden_limpio
                    )
                else:
                    # No hay coordenadas válidas
                    return None
            else:
                # Crear un polígono
                w.poly([coords])
                w.record(
                    id_poligono_limpio,
                    if_val_limpio,
                    id_credito_limpio,
                    id_persona_limpio,
                    poligono.superficie or 0,
                    estado_limpio,
                    municipio_limpio,
                    poligono.area_digitalizada or 0,
                    estatus_limpio,
                    comentarios_limpio,
                    descripcion_limpio,
                    orden_limpio
                )
            
            # Guardar el shapefile
            w.close()
            
            # Crear archivo .prj para la proyección (WGS84)
            with open(os.path.join(tempdir, 'poligono.prj'), 'w') as prj:
                prj.write('GEOGCS["GCS_WGS_1984",DATUM["D_WGS_1984",SPHEROID["WGS_1984",6378137.0,298.257223563]],PRIMEM["Greenwich",0.0],UNIT["Degree",0.0174532925199433]]')
            
            # Comprimir todos los archivos en un ZIP
            with zipfile.ZipFile(zip_buffer, 'w') as zf:
                for filename in os.listdir(tempdir):
                    filepath = os.path.join(tempdir, filename)
                    zf.write(filepath, filename)
        
        # Regresar al inicio del buffer
        zip_buffer.seek(0)
        return zip_buffer
    
    except Exception as e:
        app.logger.error(f"Error al generar shapefile individual: {e}")
        import traceback
        traceback.print_exc()
        return None

def generar_ficha_tecnica(poligono, nombre_archivo):
    """Genera una ficha técnica en formato PDF para un polígono"""
    try:
        # Crear un buffer de memoria para el PDF
        buffer = io.BytesIO()
        
        # Crear el canvas
        c = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # Agregar logos
        try:
            # Logo FIRA (izquierda)
            logo_fira_path = "static/images/logo_fira.png"
            if os.path.exists(logo_fira_path):
                c.drawImage(logo_fira_path, 1*inch, 9.5*inch, width=2*inch, height=0.75*inch, preserveAspectRatio=True)
            
            # Logo secundario (derecha)
            logo_sec_path = "static/images/logo_sec.png"
            if os.path.exists(logo_sec_path):
                c.drawImage(logo_sec_path, 6.5*inch, 9.5*inch, width=1*inch, height=1*inch, preserveAspectRatio=True)
        except Exception as e:
            app.logger.error(f"Error al cargar logos: {e}")
        
        # Título
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(width/2, 9.25*inch, "FICHA TÉCNICA")
        
        # Línea separadora
        c.line(1*inch, 9.15*inch, width-1*inch, 9.15*inch)
        
        # Detalles del polígono en formato tabular
        c.setFont("Helvetica-Bold", 10)
        y_start = 8.9*inch
        
        # Primera columna (etiquetas)
        c.drawString(1*inch, y_start, "Nombre del IF:")
        c.drawString(1*inch, y_start - 0.3*inch, "ID Polígono:")
        c.drawString(1*inch, y_start - 0.6*inch, "ID Crédito FIRA:")
        c.drawString(1*inch, y_start - 0.9*inch, "ID Persona:")
        c.drawString(1*inch, y_start - 1.2*inch, "Superficie (reportada):")
        c.drawString(1*inch, y_start - 1.5*inch, "Superficie (digitalizada):")
        
        # Segunda columna (valores) - Desplazado para alinear mejor
        c.setFont("Helvetica", 10)
        c.drawString(2.5*inch, y_start, f"{poligono.if_val or 'N/A'}")
        c.drawString(2.5*inch, y_start - 0.3*inch, f"{poligono.id_poligono or 'N/A'}")
        c.drawString(2.5*inch, y_start - 0.6*inch, f"{poligono.id_credito or 'N/A'}")
        c.drawString(2.5*inch, y_start - 0.9*inch, f"{poligono.id_persona or 'N/A'}")
        c.drawString(2.5*inch, y_start - 1.2*inch, f"{poligono.superficie or 0} ha")
        c.drawString(2.5*inch, y_start - 1.5*inch, f"{poligono.area_digitalizada or 0} ha")
        
        # Tercera columna (etiquetas) - Mayor separación horizontal
        c.setFont("Helvetica-Bold", 10)
        c.drawString(5*inch, y_start, "Estado:")
        c.drawString(5*inch, y_start - 0.3*inch, "Municipio:")
        
        # Cuarta columna (valores) - Desplazado para alinear mejor
        c.setFont("Helvetica", 10)
        c.drawString(5.8*inch, y_start, f"{corregir_codificacion(poligono.estado) or 'N/A'}")
        c.drawString(5.8*inch, y_start - 0.3*inch, f"{corregir_codificacion(poligono.municipio) or 'N/A'}")
        
        # Ajustar posición del mapa
        mapa_y_pos = 4.3*inch
        
        # Añadir borde para el mapa
        c.rect(1*inch, mapa_y_pos, 6.5*inch, 3*inch, stroke=1, fill=0)
        
        # Generar el mapa para este polígono
        mapa_image_path = None
        try:
            # Generar shapefile para este polígono
            shapefile_buffer = generar_shapefile_individual(poligono, f'temp-{nombre_archivo}')
            
            if shapefile_buffer:
                # Crear un directorio temporal para guardar el PNG
                with tempfile.TemporaryDirectory() as temp_png_dir:
                    # Generar PNG a partir del shapefile
                    png_dir = plot_shapefile_to_png(shapefile_buffer, temp_png_dir)
                    
                    # Buscar el archivo PNG generado
                    if png_dir:
                        for png_filename in os.listdir(png_dir):
                            if png_filename.endswith('.png'):
                                mapa_image_path = os.path.join(png_dir, png_filename)
                                break
                        
                        # Insertar el mapa si se encontró
                        if mapa_image_path and os.path.exists(mapa_image_path):
                            # Ajustar dimensiones para mantener el aspecto pero ajustarse al espacio disponible
                            map_width = 6.3*inch
                            map_height = 2.8*inch
                            # Centrar el mapa en el recuadro
                            c.drawImage(mapa_image_path, 1.1*inch, mapa_y_pos + 0.1*inch, 
                                       width=map_width, height=map_height, preserveAspectRatio=True)
        except Exception as map_error:
            app.logger.error(f"Error al generar o insertar el mapa: {map_error}")
            import traceback
            traceback.print_exc()
        
        # Información del metadata (parte inferior)
        y_metadata = 3.9*inch
        c.setFont("Helvetica-Bold", 10)
        c.drawString(1*inch, y_metadata, "Información del metadato:")
        
        # Crear recuadros para los metadatos
        # Primero dibujamos los recuadros - Ajustar altura para evitar superposición
        c.setFillColorRGB(0.9, 0.9, 0.9)  # Gris claro
        c.rect(1*inch, y_metadata - 2.4*inch, 3.5*inch, 2.2*inch, fill=1, stroke=1)  # Recuadro izquierdo
        c.rect(5*inch, y_metadata - 2.4*inch, 2.5*inch, 2.2*inch, fill=1, stroke=1)  # Recuadro derecho
        
        # Texto metadatos (izquierda)
        c.setFillColorRGB(0, 0, 0)  # Negro
        c.setFont("Helvetica-Bold", 9)
        # Aumentar espacio entre etiquetas
        metadata_y = y_metadata - 0.2*inch
        
        # Calcular espaciados más uniformes
        meta_spacing = 0.27*inch
        
        # Etiquetas de metadatos izquierda
        c.drawString(1.1*inch, metadata_y, "1.- Polígono")
        c.drawString(1.1*inch, metadata_y - meta_spacing, "2.- Fecha de referencia del conjunto de datos")
        c.drawString(1.1*inch, metadata_y - (meta_spacing*1.7), "    espaciales o producto:")
        c.drawString(1.1*inch, metadata_y - (meta_spacing*2.7), "3.- Unidad del estado responsable del conjunto")
        c.drawString(1.1*inch, metadata_y - (meta_spacing*3.4), "    de datos espaciales o producto:")
        c.drawString(1.1*inch, metadata_y - (meta_spacing*4.4), "4.- Calidad de la información, alcance o ámbito;")
        c.drawString(1.1*inch, metadata_y - (meta_spacing*5.1), "    nivel: Atributo:")
        
        # Valores metadatos (izquierda) - Alineados horizontalmente con las etiquetas
        c.setFont("Helvetica", 9)
        # ID de polígono alineado
        c.drawString(2.5*inch, metadata_y, f"{poligono.id_poligono or 'N/A'}")
        
        # Fecha actual
        from datetime import datetime
        fecha_actual = datetime.now().strftime("%d de %B de %Y")
        c.drawString(2.5*inch, metadata_y - (meta_spacing*1.7), fecha_actual)
        
        # Texto de "Instituto vinculados..." alineado
        c.drawString(1.5*inch, metadata_y - (meta_spacing*3.4), "Institutos vinculados en Relación con la")
        c.drawString(1.5*inch, metadata_y - (meta_spacing*4.0), "Agricultura (FIRA).")
        
        # Información aplicada al valor...
        c.drawString(1.5*inch, metadata_y - (meta_spacing*5.1), "Información aplicada al valor de atributo")
        
        # Información adicional - Observaciones (derecha)
        c.setFillColorRGB(0, 0, 0)  # Negro
        c.setFont("Helvetica-Bold", 9)
        c.drawString(5.1*inch, metadata_y, "Observaciones:")
        
        # Comentarios con mejor espaciado
        c.setFont("Helvetica", 9)
        comentarios = poligono.comentarios or "NO CUMPLE CON LA SUPERFICIE."
        # Ajustar comentarios al espacio disponible
        import textwrap
        comentario_lines = textwrap.wrap(comentarios, width=30)
        for i, line in enumerate(comentario_lines[:3]):  # Limitar a 3 líneas para dejar espacio a la descripción
            c.drawString(5.1*inch, metadata_y - 0.3*inch - (i * 0.2*inch), line)
        
        # Añadir descripción
        descripcion_y = metadata_y - 0.3*inch - (len(comentario_lines[:3]) * 0.2*inch) - 0.3*inch
        c.setFont("Helvetica-Bold", 9)
        c.drawString(5.1*inch, descripcion_y, "Descripción:")
        
        c.setFont("Helvetica", 9)
        descripcion = poligono.descripcion or ""
        descripcion_lines = textwrap.wrap(descripcion, width=30)
        for i, line in enumerate(descripcion_lines[:2]):  # Limitar a 2 líneas
            c.drawString(5.1*inch, descripcion_y - 0.2*inch - (i * 0.2*inch), line)
        
        # Información SRC mejor espaciada
        c.setFont("Helvetica-Bold", 9)
        
        # Ajustar la posición vertical del sistema de coordenadas
        src_y = metadata_y - (meta_spacing*3.5)
        c.drawString(5.1*inch, src_y, "Sistema de coordenadas")
        c.drawString(5.1*inch, src_y - 0.2*inch, "geográficas:")
        c.drawString(5.1*inch, src_y - 0.6*inch, "Dato:")
        c.drawString(5.1*inch, src_y - 1*inch, "Unidad:")
        
        # Valores SRC alineados con etiquetas
        c.setFont("Helvetica", 9)
        c.drawString(6.3*inch, src_y - 0.1*inch, "GCS WGS 1984")
        c.drawString(5.6*inch, src_y - 0.6*inch, "D WGS 1984")
        c.drawString(5.6*inch, src_y - 1*inch, "Grados")
        
        # Metadata adicional
        c.setFont("Helvetica-Bold", 9)
        c.drawString(1*inch, y_metadata - 2.6*inch, "5.- Información del contexto para los metadatos: FIRA -")
        c.drawString(1*inch, y_metadata - 2.9*inch, "    Subdirector Técnico y de Redes de Valor")
        
        # Línea divisoria
        c.line(1*inch, 1.2*inch, width-1*inch, 1.2*inch)
        
        # Firmas
        firma_y = 0.9*inch
        c.setFont("Helvetica-Bold", 10)
        nombre1 = "José Renato Navarrete Pérez"
        nombre2 = "Oswaldo Rahmses Castro Martínez"
        
        # Firma 1 (izquierda)
        c.drawCentredString(width/4, firma_y, nombre1)
        c.line(width/8, firma_y - 0.1*inch, 3*width/8, firma_y - 0.1*inch)
        c.setFont("Helvetica", 9)
        c.drawCentredString(width/4, firma_y - 0.3*inch, "Subdirector en Innovación Tecnológica")
        
        # Firma 2 (derecha)
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(3*width/4, firma_y, nombre2)
        c.line(5*width/8, firma_y - 0.1*inch, 7*width/8, firma_y - 0.1*inch)
        c.setFont("Helvetica", 9)
        c.drawCentredString(3*width/4, firma_y - 0.3*inch, "Responsable Operativo del Proyecto")
        
        # Fecha
        c.setFont("Helvetica", 9)
        today = datetime.now().strftime("%d de %B de %Y")
        c.drawString(width/8, 0.3*inch, f"FECHA: {today}")
        
        # Guardar el PDF
        c.save()
        
        # Regresar al inicio del buffer
        buffer.seek(0)
        return buffer
    
    except Exception as e:
        app.logger.error(f"Error al generar ficha técnica: {e}")
        import traceback
        traceback.print_exc()
        return None

def generar_shapefile_unificado(poligonos, nombre_archivo):
    """Genera un único archivo shapefile con todos los polígonos especificados"""
    try:
        # Crear un objeto de memoria para el archivo ZIP
        zip_buffer = io.BytesIO()
        
        # Crear un directorio temporal para los archivos del shapefile
        with tempfile.TemporaryDirectory() as tempdir:
            # Crear el writer de shapefile
            w = shapefile.Writer(os.path.join(tempdir, nombre_archivo))
            
            # Definir campos de atributos
            w.field('ID_POLIG', 'C', 40)
            w.field('IF', 'C', 40)
            w.field('ID_CRED', 'C', 40)
            w.field('ID_PERS', 'C', 40)
            w.field('SUPERF', 'N', 10, 4)
            w.field('ESTADO', 'C', 40)
            w.field('MUNICIP', 'C', 40)
            w.field('AREA_HA', 'N', 10, 4)
            w.field('ESTATUS', 'C', 10)
            w.field('COMENT', 'C', 254)
            w.field('DESCRIP', 'C', 254)  # Campo para descripción
            w.field('ORDEN', 'C', 100)    # Campo para número de orden
            
            # Función auxiliar para limpiar campos de texto que pueden tener saltos de línea
            def limpiar_campo_texto(valor):
                """Limpia un campo de texto reemplazando saltos de línea con comas y eliminando caracteres problemáticos"""
                if valor is None:
                    return ''
                
                # Convertir a string si no lo es
                valor_str = str(valor)
                
                # Reemplazar saltos de línea con comas y espacios
                valor_limpio = valor_str.replace('\n', ', ').replace('\r\n', ', ').replace('\r', ', ')
                
                # Reemplazar múltiples comas seguidas con una sola coma
                import re
                valor_limpio = re.sub(r',\s*,+', ', ', valor_limpio)
                
                # Eliminar comas al inicio y al final
                valor_limpio = valor_limpio.strip(', ')
                
                # Reemplazar caracteres problemáticos que pueden causar errores en DBF
                valor_limpio = valor_limpio.replace('\t', ' ')  # Reemplazar tabs con espacios
                valor_limpio = re.sub(r'\s+', ' ', valor_limpio)  # Reemplazar múltiples espacios con uno solo
                
                return valor_limpio
            
            # Procesar cada polígono
            for poligono in poligonos:
                # Obtener coordenadas del polígono
                coords = []
                if poligono.coordenadas_corregidas:
                    # Verificar qué separador usa: ' | ' o '|'
                    if ' | ' in poligono.coordenadas_corregidas:
                        pares = poligono.coordenadas_corregidas.split(' | ')
                    else:
                        pares = poligono.coordenadas_corregidas.split('|')
                    
                    for par in pares:
                        par = par.strip()
                        if par and ',' in par:
                            try:
                                partes = par.split(',')
                                lat = float(partes[0].strip())
                                lon = float(partes[1].strip())
                                coords.append([lon, lat])  # Shapefile usa [lon, lat]
                            except (ValueError, IndexError) as e:
                                app.logger.error(f"Error al procesar coordenada {par}: {e}")
                                continue
                    
                    app.logger.debug(f"Coordenadas procesadas para polígono {poligono.id}: {coords}")
                
                # Limpiar todos los campos de texto antes de escribirlos al DBF
                id_poligono_limpio = limpiar_campo_texto(poligono.id_poligono)[:40]
                if_val_limpio = limpiar_campo_texto(poligono.if_val)[:40]
                id_credito_limpio = limpiar_campo_texto(poligono.id_credito)[:40]
                id_persona_limpio = limpiar_campo_texto(poligono.id_persona)[:40]
                estado_limpio = limpiar_campo_texto(corregir_codificacion(poligono.estado))[:40]
                municipio_limpio = limpiar_campo_texto(corregir_codificacion(poligono.municipio))[:40]
                estatus_limpio = limpiar_campo_texto(poligono.estatus)[:10]
                comentarios_limpio = limpiar_campo_texto(poligono.comentarios)[:254]
                descripcion_limpio = limpiar_campo_texto(poligono.descripcion)[:254]
                orden_limpio = limpiar_campo_texto(poligono.orden)[:100]
                
                # Si no hay suficientes coordenadas, usar un punto o saltar
                if len(coords) < 3:
                    if len(coords) == 1:
                        # Crear un punto
                        w.point(coords[0][0], coords[0][1])
                        w.record(
                            id_poligono_limpio,
                            if_val_limpio,
                            id_credito_limpio,
                            id_persona_limpio,
                            poligono.superficie or 0,
                            estado_limpio,
                            municipio_limpio,
                            poligono.area_digitalizada or 0,
                            estatus_limpio,
                            comentarios_limpio,
                            descripcion_limpio,
                            orden_limpio
                        )
                    else:
                        # No hay coordenadas válidas, saltar este polígono
                        app.logger.debug(f"Saltando polígono {poligono.id} - coordenadas insuficientes")
                        continue
                else:
                    # Crear un polígono
                    w.poly([coords])
                    w.record(
                        id_poligono_limpio,
                        if_val_limpio,
                        id_credito_limpio,
                        id_persona_limpio,
                        poligono.superficie or 0,
                        estado_limpio,
                        municipio_limpio,
                        poligono.area_digitalizada or 0,
                        estatus_limpio,
                        comentarios_limpio,
                        descripcion_limpio,
                        orden_limpio
                    )
            
            # Guardar el shapefile
            w.close()
            
            # Crear archivo .prj para la proyección (WGS84)
            with open(os.path.join(tempdir, f'{nombre_archivo}.prj'), 'w') as prj:
                prj.write('GEOGCS["GCS_WGS_1984",DATUM["D_WGS_1984",SPHEROID["WGS_1984",6378137.0,298.257223563]],PRIMEM["Greenwich",0.0],UNIT["Degree",0.0174532925199433]]')
            
            # Comprimir todos los archivos en un ZIP
            with zipfile.ZipFile(zip_buffer, 'w') as zf:
                for filename in os.listdir(tempdir):
                    filepath = os.path.join(tempdir, filename)
                    zf.write(filepath, filename)
        
        # Regresar al inicio del buffer
        zip_buffer.seek(0)
        return zip_buffer
    
    except Exception as e:
        app.logger.error(f"Error al generar shapefile unificado: {e}")
        import traceback
        traceback.print_exc()
        return None

# Función para corregir la codificación de un texto
def corregir_codificacion(texto):
    if not texto:
        return texto
        
    try:
        # Si los nombres están en Latin-1 pero interpretados como UTF-8
        if isinstance(texto, str) and any(c in texto for c in ['Ã', 'Â']):
            return texto.encode('latin-1').decode('utf-8')
        return texto
    except Exception as e:
        app.logger.error(f"Error al corregir codificación: {e}")
        return texto

@app.route('/generar_shapefiles_y_mapas', methods=['POST'])
@login_required
def generar_shapefiles_y_mapas():
    """Ruta para generar archivos shapefile y mapas PNG de polígonos seleccionados"""
    # Obtener los índices de polígonos seleccionados
    selected_rows = request.json.get('selected_rows', [])
    
    if not selected_rows:
        return jsonify({'error': 'No se seleccionaron polígonos'}), 400
    
    try:
        # Preparar un archivo ZIP en memoria para contener todos los shapefiles y mapas
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w') as zf:
            # Crear carpetas dentro del ZIP
            zf.writestr('shapefiles/', '')
            zf.writestr('mapas/', '')
            
            # Para cada polígono seleccionado
            for row_id in selected_rows:
                try:
                    row_id = int(row_id)
                    # Primero intentar buscar por ID exacto
                    poligono = Poligono.query.get(row_id)
                    
                    if poligono is None:
                        # Si no se encuentra, imprimir para depuración
                        app.logger.error(f"No se encontró polígono con ID {row_id}, buscando en posición")
                        
                        # Intentar buscar por posición como fallback
                        poligonos = Poligono.query.all()
                        if 0 <= row_id < len(poligonos):
                            poligono = poligonos[row_id]
                        else:
                            app.logger.debug(f"Índice {row_id} fuera de rango, hay {len(poligonos)} polígonos")
                            continue
                    
                    app.logger.debug(f"Generando shapefile para polígono ID={poligono.id}, ID_POLIGONO={poligono.id_poligono}")
                except Exception as e:
                    app.logger.error(f"Error al recuperar polígono {row_id}: {e}")
                    # Si no es un índice válido, continuar con el siguiente
                    continue
                
                # Generar shapefile para este polígono
                shapefile_buffer = generar_shapefile_individual(poligono, f'polygon-{row_id}')
                
                if shapefile_buffer:
                    # Usar ID_POLIGONO para nombrar el archivo si está disponible
                    archivo_nombre = poligono.id_poligono if poligono.id_poligono else f'polygon-{row_id}'
                    zf.writestr(f'shapefiles/{archivo_nombre}.zip', shapefile_buffer.getvalue())
                    
                    # Generar y añadir el mapa PNG
                    try:
                        # Crear un directorio temporal para guardar los PNG
                        with tempfile.TemporaryDirectory() as temp_png_dir:
                            # Generar PNG a partir del shapefile
                            png_dir = plot_shapefile_to_png(shapefile_buffer, temp_png_dir)
                            
                            # Añadir todos los archivos PNG al ZIP
                            if png_dir:
                                for png_filename in os.listdir(png_dir):
                                    if png_filename.endswith('.png'):
                                        png_path = os.path.join(png_dir, png_filename)
                                        with open(png_path, 'rb') as png_file:
                                            zf.writestr(f'mapas/{png_filename}', png_file.read())
                    except Exception as e:
                        app.logger.error(f"Error al generar mapa PNG para polígono {row_id}: {e}")
                        import traceback
                        traceback.print_exc()
        
        # Regresar al inicio del archivo en memoria
        memory_file.seek(0)
        
        # Enviar el archivo ZIP como respuesta
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name='poligonos_shapefiles_y_mapas.zip'
        )
    
    except Exception as e:
        app.logger.error(f"Error al generar shapefiles y mapas: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/procesar-shp', methods=['POST'])
@login_required
@limiter.limit("10 per hour")
def procesar_shp():
    try:
        app.logger.debug("Ruta /procesar-shp llamada")
        
        # Verificar que el directorio de uploads existe
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            os.makedirs(app.config['UPLOAD_FOLDER'])
            app.logger.info(f"Directorio de uploads creado: {app.config['UPLOAD_FOLDER']}")
        
        # Verificar que el directorio de uploads tiene permisos de escritura
        if not os.access(app.config['UPLOAD_FOLDER'], os.W_OK):
            error_msg = f"Error: No hay permisos de escritura en el directorio {app.config['UPLOAD_FOLDER']}"
            app.logger.error(error_msg)
            return jsonify({'error': error_msg}), 500
        
        if 'zipfile' not in request.files:
            app.logger.error("Error: No hay archivo en la solicitud")
            # Verificar si es una solicitud AJAX o un formulario directo
            if request.is_xhr or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'error': 'No se ha enviado ningún archivo'}), 400
            else:
                flash('No se ha enviado ningún archivo', 'error')
                return redirect(url_for('unir_archivos'))
        
        archivo = request.files['zipfile']
        app.logger.debug(f"Archivo recibido: {archivo.filename}")
        
        if archivo.filename == '':
            app.logger.error("Error: Nombre de archivo vacío")
            # Verificar si es una solicitud AJAX o un formulario directo
            if request.is_xhr or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'error': 'No se ha seleccionado ningún archivo'}), 400
            else:
                flash('No se ha seleccionado ningún archivo', 'error')
                return redirect(url_for('unir_archivos'))
        
        # Verificar tamaño del archivo
        MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB
        archivo.seek(0, os.SEEK_END)
        file_size = archivo.tell()
        archivo.seek(0)  # Resetear el puntero al inicio
        
        if file_size > MAX_FILE_SIZE:
            error_msg = f"El archivo es demasiado grande. Tamaño máximo permitido: 50 MB"
            app.logger.error(error_msg)
            return jsonify({'error': error_msg}), 413  # Request Entity Too Large
        
        if archivo and archivo.filename.endswith('.zip'):
            try:
                app.logger.debug(f"Procesando archivo ZIP: {archivo.filename}")
                
                # Crear directorio temporal para extracción
                try:
                    temp_dir = tempfile.mkdtemp()
                    app.logger.debug(f"Directorio temporal creado: {temp_dir}")
                except Exception as e:
                    error_msg = f"Error al crear directorio temporal: {str(e)}"
                    app.logger.error(error_msg)
                    return jsonify({'error': error_msg}), 500
                
                # Guardar archivo ZIP
                try:
                    zip_path = os.path.join(temp_dir, 'input.zip')
                    archivo.save(zip_path)
                    app.logger.debug(f"Archivo guardado en: {zip_path}")
                except Exception as e:
                    error_msg = f"Error al guardar archivo: {str(e)}"
                    app.logger.error(error_msg)
                    return jsonify({'error': error_msg}), 500
                
                # Verificar si es un ZIP válido
                try:
                    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                        # Verificar si el ZIP no está dañado
                        if zip_ref.testzip() is not None:
                            error_msg = "El archivo ZIP está dañado"
                            app.logger.error(error_msg)
                            return jsonify({'error': error_msg}), 400
                        
                        # Limitar el número de archivos dentro del ZIP
                        MAX_FILES = 500
                        if len(zip_ref.namelist()) > MAX_FILES:
                            error_msg = f"El archivo ZIP contiene demasiados archivos (máximo {MAX_FILES})"
                            app.logger.error(error_msg)
                            return jsonify({'error': error_msg}), 413
                        
                        # Verificar que el tamaño descomprimido no sea excesivo
                        MAX_UNCOMPRESSED_SIZE = 200 * 1024 * 1024  # 200 MB
                        total_size = sum(info.file_size for info in zip_ref.infolist())
                        if total_size > MAX_UNCOMPRESSED_SIZE:
                            error_msg = f"El tamaño descomprimido del ZIP es demasiado grande (máximo 200 MB)"
                            app.logger.error(error_msg)
                            return jsonify({'error': error_msg}), 413
                        
                        # Extraer el ZIP
                        zip_ref.extractall(temp_dir)
                    app.logger.debug(f"Archivo ZIP extraído en: {temp_dir}")
                except zipfile.BadZipFile:
                    error_msg = "El archivo no es un ZIP válido"
                    app.logger.error(error_msg)
                    return jsonify({'error': error_msg}), 400
                except Exception as e:
                    error_msg = f"Error al extraer archivo ZIP: {str(e)}"
                    app.logger.error(error_msg)
                    return jsonify({'error': error_msg}), 500
                
                # Buscar archivos SHP o ZIPs anidados
                try:
                    shp_files = []
                    internal_zips = []
                    
                    # Buscar archivos SHP y ZIPs anidados en el primer nivel
                    for root, dirs, files in os.walk(temp_dir):
                        for file in files:
                            if file.endswith('.shp'):
                                shp_files.append(os.path.join(root, file))
                            elif file.endswith('.zip'):
                                internal_zips.append(os.path.join(root, file))
                    
                    app.logger.debug(f"Archivos SHP encontrados (primer nivel): {len(shp_files)}")
                    app.logger.debug(f"Archivos ZIP internos encontrados: {len(internal_zips)}")
                    
                    # Extraer y procesar ZIPs anidados si no se encontraron archivos SHP
                    if not shp_files and internal_zips:
                        app.logger.debug("Extrayendo archivos ZIP internos...")
                        # Limitar el número de ZIPs anidados a procesar
                        MAX_NESTED_ZIPS = 10
                        if len(internal_zips) > MAX_NESTED_ZIPS:
                            app.logger.debug(f"Limitando a {MAX_NESTED_ZIPS} ZIPs anidados")
                            internal_zips = internal_zips[:MAX_NESTED_ZIPS]
                        
                        for zip_file in internal_zips:
                            zip_name = os.path.basename(zip_file)
                            extract_subdir = os.path.join(temp_dir, f"extracted_{zip_name.replace('.zip', '')}")
                            os.makedirs(extract_subdir, exist_ok=True)
                            
                            try:
                                app.logger.debug(f"Extrayendo ZIP interno: {zip_name} en {extract_subdir}")
                                # Verificar el ZIP interno antes de extraerlo
                                with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                                    # Verificar ZIP no dañado
                                    if zip_ref.testzip() is not None:
                                        app.logger.error(f"ZIP interno {zip_name} está dañado, omitiendo")
                                        continue
                                    
                                    # Verificar número de archivos
                                    if len(zip_ref.namelist()) > MAX_FILES:
                                        app.logger.debug(f"ZIP interno {zip_name} tiene demasiados archivos, omitiendo")
                                        continue
                                    
                                    # Verificar tamaño descomprimido
                                    nested_total_size = sum(info.file_size for info in zip_ref.infolist())
                                    if nested_total_size > MAX_UNCOMPRESSED_SIZE:
                                        app.logger.debug(f"ZIP interno {zip_name} es demasiado grande, omitiendo")
                                        continue
                                    
                                    # Extraer archivos
                                    zip_ref.extractall(extract_subdir)
                                
                                # Buscar archivos SHP en el ZIP extraído
                                for root, dirs, files in os.walk(extract_subdir):
                                    for file in files:
                                        if file.endswith('.shp'):
                                            shp_path = os.path.join(root, file)
                                            shp_files.append(shp_path)
                                            app.logger.debug(f"  - SHP encontrado en ZIP interno: {shp_path}")
                            except zipfile.BadZipFile:
                                app.logger.debug(f"ZIP interno {zip_name} no es válido, omitiendo")
                                continue
                            except Exception as e:
                                app.logger.error(f"Error al extraer ZIP interno {zip_name}: {str(e)}")
                                # Continúa con el siguiente ZIP
                    
                    app.logger.debug(f"Total de archivos SHP encontrados: {len(shp_files)}")
                    for shp in shp_files:
                        app.logger.debug(f"  - {shp}")
                    
                    if not shp_files:
                        error_msg = "No se encontraron archivos SHP en el archivo ZIP"
                        app.logger.error(error_msg)
                        return jsonify({'error': error_msg}), 400
                except Exception as e:
                    error_msg = f"Error al buscar archivos SHP: {str(e)}"
                    app.logger.error(error_msg)
                    return jsonify({'error': error_msg}), 500
                
                # Limitar el número de archivos SHP a procesar
                MAX_SHP_FILES = 20
                if len(shp_files) > MAX_SHP_FILES:
                    app.logger.debug(f"Limitando a {MAX_SHP_FILES} archivos SHP")
                    shp_files = shp_files[:MAX_SHP_FILES]
                
                # Unir archivos SHP con geopandas
                try:
                    merged_gdf = None
                    for shp_file in shp_files:
                        app.logger.debug(f"Procesando archivo: {shp_file}")
                        try:
                            # Verificar tamaño del archivo SHP
                            if os.path.getsize(shp_file) > 20 * 1024 * 1024:  # 20 MB
                                app.logger.debug(f"  - SHP demasiado grande, omitiendo: {shp_file}")
                                continue
                            
                            gdf = gpd.read_file(shp_file)
                            
                            # Limitar el número de geometrías
                            MAX_FEATURES = 5000
                            if len(gdf) > MAX_FEATURES:
                                app.logger.debug(f"  - Demasiadas geometrías ({len(gdf)}), limitando a {MAX_FEATURES}")
                                gdf = gdf.head(MAX_FEATURES)
                            
                            app.logger.debug(f"  - Geometrías: {len(gdf)}, CRS: {gdf.crs}")
                            
                            if merged_gdf is None:
                                merged_gdf = gdf
                            else:
                                # Asegurarse de que tienen el mismo CRS
                                if gdf.crs != merged_gdf.crs and gdf.crs is not None:
                                    app.logger.debug(f"  - Convirtiendo CRS de {gdf.crs} a {merged_gdf.crs}")
                                    gdf = gdf.to_crs(merged_gdf.crs)
                                
                                # Concatenar con seguridad
                                try:
                                    merged_gdf = pd.concat([merged_gdf, gdf])
                                except Exception as concat_error:
                                    app.logger.error(f"  - Error al concatenar: {str(concat_error)}")
                                    # Si falla la concatenación, intentar solo con geometrías
                                    try:
                                        app.logger.debug("  - Intentando concatenar solo geometrías...")
                                        # Crear un nuevo GeoDataFrame con solo geometrías
                                        simple_gdf = gpd.GeoDataFrame(geometry=gdf.geometry)
                                        merged_gdf = pd.concat([merged_gdf, simple_gdf])
                                    except Exception as simple_concat_error:
                                        app.logger.error(f"  - Error en concatenación simple: {str(simple_concat_error)}")
                                        # Continuar con el siguiente archivo
                                        continue
                        except Exception as e:
                            error_msg = f"Error al procesar archivo {os.path.basename(shp_file)}: {str(e)}"
                            app.logger.error(error_msg)
                            # Continuamos con el siguiente archivo en lugar de fallar completamente
                            continue
                    
                    if merged_gdf is None or len(merged_gdf) == 0:
                        error_msg = "No se pudieron procesar los archivos SHP"
                        app.logger.error(error_msg)
                        return jsonify({'error': error_msg}), 500
                    
                    # Limitar el tamaño final del GeoDataFrame
                    MAX_FINAL_FEATURES = 10000
                    if len(merged_gdf) > MAX_FINAL_FEATURES:
                        app.logger.debug(f"GeoDataFrame final demasiado grande ({len(merged_gdf)}), limitando a {MAX_FINAL_FEATURES}")
                        merged_gdf = merged_gdf.head(MAX_FINAL_FEATURES)
                    
                    app.logger.info(f"Unión completada: {len(merged_gdf)} geometrías")
                except Exception as e:
                    error_msg = f"Error al unir archivos SHP: {str(e)}"
                    app.logger.error(error_msg)
                    return jsonify({'error': error_msg}), 500
                
                # Guardar el archivo unificado
                try:
                    output_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'shp_unified')
                    if not os.path.exists(output_dir):
                        os.makedirs(output_dir)
                    app.logger.debug(f"Directorio de salida creado: {output_dir}")
                    
                    output_shp = os.path.join(temp_dir, 'unified.shp')
                    app.logger.debug(f"Guardando archivo unificado en: {output_shp}")
                    
                    # Simplificar el GeoDataFrame para la escritura
                    try:
                        # Intentar guardar con todas las columnas
                        merged_gdf.to_file(output_shp)
                    except Exception as save_error:
                        app.logger.error(f"Error al guardar GeoDataFrame completo: {str(save_error)}")
                        app.logger.debug("Intentando guardar con columnas reducidas...")
                        
                        # Crear un GeoDataFrame simplificado con solo la geometría
                        simple_gdf = gpd.GeoDataFrame(geometry=merged_gdf.geometry)
                        simple_gdf.to_file(output_shp)
                    
                    app.logger.debug(f"Archivo guardado correctamente")
                except Exception as e:
                    error_msg = f"Error al guardar archivo unificado: {str(e)}"
                    app.logger.error(error_msg)
                    return jsonify({'error': error_msg}), 500
                
                # Crear archivo ZIP con los archivos resultantes
                try:
                    output_zip = os.path.join(output_dir, 'unified_shp.zip')
                    app.logger.debug(f"Creando archivo ZIP de salida: {output_zip}")
                    
                    # Incluir archivos auxiliares (.dbf, .shx, .prj)
                    base_name = os.path.splitext(output_shp)[0]
                    with zipfile.ZipFile(output_zip, 'w') as zipf:
                        for ext in ['.shp', '.dbf', '.shx', '.prj']:
                            file_path = base_name + ext
                            if os.path.exists(file_path):
                                app.logger.debug(f"  - Añadiendo archivo: {os.path.basename(file_path)}")
                                zipf.write(file_path, os.path.basename(file_path))
                except Exception as e:
                    error_msg = f"Error al crear archivo ZIP de salida: {str(e)}"
                    app.logger.error(error_msg)
                    return jsonify({'error': error_msg}), 500
                
                # Preparar datos para respuesta
                try:
                    # Crear una versión extremadamente simplificada del GeoJSON para la respuesta
                    # En lugar de enviar todas las geometrías, enviar solo un resumen o un subconjunto muy pequeño
                    simplified_gdf = None
                    try:
                        # Intentar crear una versión muy simplificada con solo los primeros polígonos
                        if len(merged_gdf) > 0:
                            # Tomar solo los primeros 5 polígonos como muestra
                            sample_gdf = merged_gdf.head(5).copy()
                            
                            # Aplicar una simplificación agresiva a las geometrías
                            try:
                                sample_gdf.geometry = sample_gdf.geometry.simplify(tolerance=0.01)
                            except Exception as simplify_error:
                                app.logger.error(f"Error al simplificar geometrías de muestra: {str(simplify_error)}")
                            
                            # Eliminar todas las columnas excepto la geometría
                            simplified_gdf = gpd.GeoDataFrame(geometry=sample_gdf.geometry)
                            app.logger.debug(f"GeoJSON simplificado creado con {len(simplified_gdf)} geometrías de muestra")
                    except Exception as sample_error:
                        app.logger.error(f"Error al crear muestra de GeoJSON: {str(sample_error)}")
                        # Continuar sin GeoJSON si hay error
                    
                    # Si no se pudo crear una versión simplificada, usar un GeoJSON vacío
                    if simplified_gdf is None or len(simplified_gdf) == 0:
                        geojson_data = '{"type":"FeatureCollection","features":[]}'
                        app.logger.debug("Usando GeoJSON vacío para la respuesta")
                    else:
                        # Convertir a GeoJSON con manejo de errores
                        try:
                            geojson_data = simplified_gdf.to_json()
                            # Verificar tamaño del JSON
                            if len(geojson_data) > 1000000:  # Más de 1MB
                                app.logger.debug(f"GeoJSON demasiado grande ({len(geojson_data)} bytes), usando vacío")
                                geojson_data = '{"type":"FeatureCollection","features":[]}'
                        except Exception as json_error:
                            app.logger.error(f"Error al convertir a GeoJSON: {str(json_error)}")
                            geojson_data = '{"type":"FeatureCollection","features":[]}'
                    
                    # Obtener conteo de polígonos
                    num_poligonos = len(merged_gdf)
                    
                    # Calcular área con manejo de errores
                    try:
                        area_total = merged_gdf.geometry.area.sum() / 10000  # Convertir a hectáreas
                    except Exception as area_error:
                        app.logger.error(f"Error al calcular área: {str(area_error)}")
                        area_total = 0
                    
                    app.logger.debug(f"Datos preparados: {num_poligonos} polígonos, {area_total:.2f} ha")
                except Exception as e:
                    error_msg = f"Error al preparar datos para respuesta: {str(e)}"
                    app.logger.error(error_msg)
                    # No fallar aquí, continuar con valores predeterminados
                    geojson_data = '{"type":"FeatureCollection","features":[]}'
                    num_poligonos = 0
                    area_total = 0
                
                # Crear un diccionario de respuesta mínimo
                response_data = {
                    'success': True,
                    'message': 'Archivos SHP unidos correctamente',
                    'archivo_salida': '/uploads/shp_unified/unified_shp.zip',
                    'num_archivos': len(shp_files),
                    'num_poligonos': num_poligonos,
                    'area_total': round(area_total, 2)
                }
                
                # Añadir geojson solo si no está vacío y es pequeño
                if geojson_data != '{"type":"FeatureCollection","features":[]}':
                    response_data['geojson'] = geojson_data
                else:
                    # Indicar que el GeoJSON está disponible pero no se incluye en la respuesta
                    response_data['geojson_status'] = 'no_incluido_por_tamano'
                
                # Limpiar directorio temporal
                try:
                    import shutil
                    shutil.rmtree(temp_dir)
                    app.logger.debug(f"Directorio temporal eliminado: {temp_dir}")
                except Exception as e:
                    app.logger.error(f"Advertencia: No se pudo eliminar el directorio temporal: {str(e)}")
                
                # Verificar si es una solicitud AJAX o un formulario directo
                is_ajax = request.is_xhr or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                if is_ajax:
                    app.logger.debug("Enviando respuesta JSON")
                    try:
                        return jsonify(response_data)
                    except Exception as json_error:
                        app.logger.error(f"Error al serializar respuesta JSON: {str(json_error)}")
                        # Intentar con una respuesta más sencilla sin GeoJSON
                        del response_data['geojson']
                        response_data['geojson_status'] = 'error_serializacion'
                        return jsonify(response_data)
                else:
                    # Si es un formulario directo, guardar datos en sesión y redirigir
                    app.logger.debug("Redireccionando con datos en sesión")
                    flash('Archivos SHP unidos correctamente. Puede descargar el resultado.', 'success')
                    session['resultado_shp'] = {
                        'num_archivos': len(shp_files),
                        'num_poligonos': num_poligonos,
                        'area_total': round(area_total, 2)
                    }
                    return redirect(url_for('unir_archivos'))
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                error_msg = f"Error al procesar archivos: {str(e)}"
                app.logger.error(f"Error al procesar: {error_msg}")
                
                # Verificar si es una solicitud AJAX o un formulario directo
                if request.is_xhr or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'error': error_msg}), 500
                else:
                    flash(error_msg, 'error')
                    return redirect(url_for('unir_archivos'))
        else:
            error_msg = "Formato de archivo no válido. Debe ser un archivo ZIP"
            app.logger.error(error_msg)
            # Verificar si es una solicitud AJAX o un formulario directo
            if request.is_xhr or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'error': error_msg}), 400
            else:
                flash(error_msg, 'error')
                return redirect(url_for('unir_archivos'))
    except Exception as e:
        # Capturar cualquier excepción no manejada para evitar respuestas HTML de error 500
        import traceback
        traceback.print_exc()
        error_msg = f"Error interno del servidor: {str(e)}"
        app.logger.error(f"ERROR NO MANEJADO: {error_msg}")
        
        # Siempre devolver una respuesta JSON válida
        if request.is_xhr or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': error_msg}), 500
        else:
            flash(error_msg, 'error')
            return redirect(url_for('unir_archivos'))

@app.route('/descargar-shp-unificado')
@login_required
def descargar_shp_unificado():
    zip_path = os.path.join(app.config['UPLOAD_FOLDER'], 'shp_unified', 'unified_shp.zip')
    if os.path.exists(zip_path):
        return send_file(zip_path, as_attachment=True, download_name='poligonos_unificados.zip')
    else:
        flash('No se encontró el archivo unificado. Procese los archivos primero.', 'error')
        return redirect(url_for('unir_archivos'))

@app.route('/generar_ficha_tecnica_template/<int:db_id>')
@login_required
def generar_ficha_tecnica_template_route(db_id):
    """Genera una ficha técnica con la nueva plantilla para un polígono específico"""
    try:
        # Obtener las fechas del formulario
        fecha_referencia = request.args.get('fecha_referencia')
        fecha_final = request.args.get('fecha_final')
        
        # Buscar el polígono en la base de datos
        poligono = Poligono.query.get(db_id)
        if not poligono:
            return jsonify({"error": "Polígono no encontrado"}), 404
        
        # Verificar que PyMuPDF esté correctamente configurado
        pymupdf_funcional = garantizar_pymupdf()
        if not pymupdf_funcional:
            app.logger.error("ADVERTENCIA: PyMuPDF no está correctamente configurado, se usará el método alternativo")
        
        # Generar mapa del polígono
        shapefile_buffer = generar_shapefile_individual(poligono, f'polygon-{db_id}')
        if not shapefile_buffer:
            return jsonify({"error": "No se pudo generar el shapefile"}), 500
        
        png_filepath = None
        # Crear un directorio temporal para guardar los PNG
        with tempfile.TemporaryDirectory() as temp_png_dir:
            # Generar PNG a partir del shapefile
            png_dir = plot_shapefile_to_png(shapefile_buffer, temp_png_dir)
            
            # Buscar el archivo PNG generado
            if png_dir:
                for png_filename in os.listdir(png_dir):
                    if png_filename.endswith('.png'):
                        png_filepath = os.path.join(png_dir, png_filename)
                        break
            
            if not png_filepath:
                return jsonify({"error": "No se pudo generar la imagen del mapa"}), 500
            
            # Generar ficha técnica con la plantilla
            pdf_buffer = generar_ficha_tecnica_desde_plantilla(poligono, png_filepath, fecha_referencia, fecha_final)
            
            if not pdf_buffer:
                return jsonify({"error": "No se pudo generar la ficha técnica"}), 500
            
            # Enviar el PDF como respuesta
            return send_file(
                pdf_buffer,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f'ficha_tecnica_{poligono.id_poligono or db_id}.pdf'
            )
    
    except Exception as e:
        app.logger.error(f"Error al generar ficha técnica con plantilla: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/generar_paquete_completo_con_plantilla', methods=['POST'])
@login_required
def generar_paquete_completo_con_plantilla():
    """Ruta para generar un paquete completo con fichas PDF basadas en plantilla y shapefiles"""
    # Obtener los índices de polígonos seleccionados
    selected_rows = request.json.get('selected_rows', [])
    fecha_referencia = request.json.get('fecha_referencia')
    fecha_final = request.json.get('fecha_final')
    
    # Debugging: Log the received dates
    app.logger.debug(f"DEBUG - Fecha de Referencia recibida: {fecha_referencia}")
    app.logger.debug(f"DEBUG - Fecha Final recibida: {fecha_final}")
    
    if not selected_rows:
        return jsonify({'error': 'No se seleccionaron polígonos'}), 400
    
    try:
        # Verificamos PyMuPDF usando la función mejorada
        from utils.pdf_generator import garantizar_pymupdf
        pymupdf_funcional = garantizar_pymupdf()
        if not pymupdf_funcional:
            app.logger.error("ADVERTENCIA: PyMuPDF no está correctamente configurado, se usará el método alternativo")
            
        # Verificar que la plantilla existe - usar el directorio static/plantilla
        plantilla_path = os.path.join('static', 'plantilla', 'Plantilla_2.pdf')
        plantilla_encontrada = False
        
        if os.path.exists(plantilla_path):
            plantilla_encontrada = True
            app.logger.debug(f"Plantilla encontrada en: {os.path.abspath(plantilla_path)}")
        else:
            # Verificar rutas alternativas
            rutas_alternativas = [
                'Plantilla_2.pdf',
                './plantilla/Plantilla_2.pdf',
                '../plantilla/Plantilla_2.pdf',
                './static/plantilla/Plantilla_2.pdf'
            ]
            
            for ruta in rutas_alternativas:
                if os.path.exists(ruta):
                    plantilla_path = ruta
                    plantilla_encontrada = True
                    app.logger.debug(f"Plantilla encontrada en ruta alternativa: {os.path.abspath(ruta)}")
                    break
            
        # Añadir mensaje de advertencia si no se encontró la plantilla
        if not plantilla_encontrada:
            app.logger.error("ADVERTENCIA: No se encontró la plantilla, se usará el método alternativo")
            
        # Continuamos con el proceso independientemente de la plantilla

        # Verificar permisos de lectura del archivo
        try:
            if plantilla_encontrada:
                with open(plantilla_path, 'rb') as f:
                    _ = f.read(1)  # Leer un byte para verificar acceso
                    app.logger.debug(f"Archivo de plantilla accesible: {plantilla_path}")
        except Exception as e:
            app.logger.error(f"Error al acceder a la plantilla: {e}")
            app.logger.debug("Se utilizará el método alternativo para generar PDFs")
            
        # Preparar un archivo ZIP en memoria para contener todos los archivos
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w') as zf:
            # Crear carpeta para fichas técnicas
            zf.writestr('fichas_tecnicas/', '')
            # Crear carpeta para shapefiles
            zf.writestr('shapefiles/', '')
            # Crear carpeta para mapas
            zf.writestr('mapas/', '')
            # Crear una carpeta para logs
            zf.writestr('logs/', '')
            
            # Contadores para estadísticas
            total_poligonos = len(selected_rows)
            poligonos_procesados = 0
            errores = 0
            errores_detalles = []
            
            # Para cada polígono seleccionado
            for row_id in selected_rows:
                try:
                    row_id = int(row_id)
                    # Primero intentar buscar por ID exacto
                    poligono = Poligono.query.get(row_id)
                    
                    if poligono is None:
                        # Si no se encuentra, imprimir para depuración
                        app.logger.error(f"No se encontró polígono con ID {row_id}, buscando en posición")
                        
                        # Intentar buscar por posición como fallback
                        poligonos = Poligono.query.all()
                        if 0 <= row_id < len(poligonos):
                            poligono = poligonos[row_id]
                        else:
                            app.logger.debug(f"Índice {row_id} fuera de rango, hay {len(poligonos)} polígonos")
                            error_msg = f"Índice {row_id} fuera de rango, hay {len(poligonos)} polígonos"
                            errores_detalles.append(error_msg)
                            errores += 1
                            continue
                            
                    app.logger.debug(f"Generando fichas para polígono ID={poligono.id}, ID_POLIGONO={poligono.id_poligono}")
                except Exception as e:
                    app.logger.error(f"Error al recuperar polígono {row_id}: {e}")
                    error_msg = f"Error al recuperar polígono {row_id}: {e}"
                    errores_detalles.append(error_msg)
                    errores += 1
                    # Si no es un ID válido, continuar con el siguiente
                    continue
                
                # Generar shapefile para este polígono
                shapefile_buffer = generar_shapefile_individual(poligono, f'polygon-{row_id}')
                png_filepath = None
                
                if shapefile_buffer:
                    # Usar ID_POLIGONO para nombrar el archivo si está disponible
                    archivo_nombre = poligono.id_poligono if poligono.id_poligono else f'polygon-{row_id}'
                    zf.writestr(f'shapefiles/{archivo_nombre}.zip', shapefile_buffer.getvalue())
                    
                    # Generar mapas PNG a partir del shapefile
                    try:
                        # Crear un directorio temporal para guardar los PNG
                        with tempfile.TemporaryDirectory() as temp_png_dir:
                            # Generar PNG a partir del shapefile
                            png_dir = plot_shapefile_to_png(shapefile_buffer, temp_png_dir)
                            
                            # Añadir todos los archivos PNG al ZIP y guardar la ruta del primer PNG para la ficha técnica
                            if png_dir:
                                for png_filename in os.listdir(png_dir):
                                    if png_filename.endswith('.png'):
                                        png_path = os.path.join(png_dir, png_filename)
                                        # Guardar la ruta del primer PNG para usarla en la ficha
                                        if png_filepath is None:
                                            png_filepath = png_path
                                        
                                        # Guardar la imagen en un archivo temporal más permanente que podamos usar para el PDF
                                        temp_img_path = tempfile.NamedTemporaryFile(delete=False, suffix='.png').name
                                        shutil.copy2(png_path, temp_img_path)
                                        png_filepath = temp_img_path
                                        
                                        with open(png_path, 'rb') as png_file:
                                            # Guardar con un nombre predecible basado en ID_POLIGONO
                                            png_name = f"{poligono.id_poligono or f'polygon-{row_id}'}.png"
                                            zf.writestr(f'mapas/{png_name}', png_file.read())
                    except Exception as e:
                        app.logger.error(f"Error al generar mapa PNG para polígono {row_id}: {e}")
                        error_msg = f"Error al generar mapa PNG para polígono {row_id}: {e}"
                        errores_detalles.append(error_msg)
                        errores += 1
                        import traceback
                        traceback.print_exc()
                
                # Generar ficha técnica PDF con la nueva plantilla
                if png_filepath:
                    try:
                        # Configurar un log específico para este polígono
                        log_buffer = io.StringIO()
                        log_buffer.write(f"=== Log de generación de PDF para polígono {row_id} ===\n")
                        log_buffer.write(f"ID_POLIGONO: {poligono.id_poligono}\n")
                        log_buffer.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                        
                        # Verificar que la imagen existe
                        if os.path.exists(png_filepath):
                            log_buffer.write(f"Imagen encontrada en: {os.path.abspath(png_filepath)}\n")
                        else:
                            log_buffer.write(f"ADVERTENCIA: No se encontró la imagen en: {os.path.abspath(png_filepath)}\n")
                        
                        # Intentar generar el PDF con la plantilla
                        log_buffer.write("Intentando generar PDF con la plantilla...\n")
                        pdf_buffer = generar_ficha_tecnica_desde_plantilla(poligono, png_filepath, fecha_referencia, fecha_final)
                        
                        # Guardar una copia del PDF (si se generó) para depuración
                        if pdf_buffer:
                            log_buffer.write(f"PDF generado correctamente con plantilla ({len(pdf_buffer.getvalue())} bytes)\n")
                            # Guardar una copia del PDF generado con plantilla para diagnóstico
                            try:
                                debug_pdf_path = os.path.join('debug', f'template_pdf_{poligono.id_poligono or row_id}.pdf')
                                os.makedirs('debug', exist_ok=True)
                                with open(debug_pdf_path, 'wb') as f:
                                    f.write(pdf_buffer.getvalue())
                                log_buffer.write(f"Copia de diagnóstico guardada en: {debug_pdf_path}\n")
                                pdf_buffer.seek(0)  # Resetear el buffer
                            except Exception as debug_error:
                                log_buffer.write(f"No se pudo guardar copia de diagnóstico: {debug_error}\n")
                        
                        # Si falla la generación con plantilla, usar el método simple como respaldo
                        if pdf_buffer is None:
                            log_buffer.write("Error al generar PDF con plantilla, intentando método simple como respaldo...\n")
                            pdf_buffer = generar_ficha_tecnica_simple(poligono, png_filepath)
                            
                            # Guardar una copia del PDF simple para comparación
                            if pdf_buffer:
                                try:
                                    debug_simple_path = os.path.join('debug', f'simple_pdf_{poligono.id_poligono or row_id}.pdf')
                                    os.makedirs('debug', exist_ok=True)
                                    with open(debug_simple_path, 'wb') as f:
                                        f.write(pdf_buffer.getvalue())
                                    log_buffer.write(f"Copia de PDF simple guardada en: {debug_simple_path}\n")
                                    pdf_buffer.seek(0)  # Resetear el buffer
                                except Exception as debug_error:
                                    log_buffer.write(f"No se pudo guardar copia de PDF simple: {debug_error}\n")
                        
                        if pdf_buffer:
                            log_buffer.write(f"PDF generado correctamente ({len(pdf_buffer.getvalue())} bytes)\n")
                            nombre_archivo = f"{poligono.id_poligono or f'polygon-{row_id}'}.pdf"
                            zf.writestr(f'fichas_tecnicas/{nombre_archivo}', pdf_buffer.getvalue())
                            poligonos_procesados += 1
                        else:
                            log_buffer.write("ERROR: No se pudo generar el PDF (buffer vacío)\n")
                            errores += 1
                            error_msg = f"No se pudo generar el PDF para el polígono {row_id} - ID_POLIGONO={poligono.id_poligono}"
                            errores_detalles.append(error_msg)
                    except Exception as e:
                        app.logger.error(f"Error al generar PDF para polígono {row_id}: {e}")
                        log_buffer.write(f"EXCEPCIÓN: {str(e)}\n")
                        import traceback
                        error_traceback = traceback.format_exc()
                        log_buffer.write(f"Traceback:\n{error_traceback}\n")
                        
                        error_msg = f"Error al generar PDF para polígono {row_id}: {e}"
                        errores_detalles.append(error_msg)
                        errores += 1
                    finally:
                        # Guardar el log en el archivo ZIP
                        archivo_nombre = poligono.id_poligono if poligono.id_poligono else f'polygon-{row_id}'
                        zf.writestr(f'logs/log_{archivo_nombre}.txt', log_buffer.getvalue())
                else:
                    error_msg = f"No se pudo generar mapa para polígono {row_id}"
                    errores_detalles.append(error_msg)
                    errores += 1
                    app.logger.error(f"Error: No se pudo generar mapa para polígono {row_id}")
            
            # Añadir un resumen en el ZIP
            resumen = f"""
            Resumen de generación de fichas técnicas:
            --------------------------------------
            Total de polígonos seleccionados: {total_poligonos}
            Polígonos procesados exitosamente: {poligonos_procesados}
            Errores: {errores}
            
            Detalles de errores:
            {chr(10).join(errores_detalles)}
            
            Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
            """
            zf.writestr('resumen.txt', resumen)
        
        # Regresar al inicio del archivo en memoria
        memory_file.seek(0)
        
        # Enviar el archivo ZIP como respuesta
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name='paquete_completo_con_plantilla.zip'
        )
    
    except Exception as e:
        app.logger.error(f"Error al generar paquete completo con plantilla: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/generar_excel', methods=['GET'])
@login_required
def generar_excel():
    """Ruta para generar un archivo Excel con todos los registros de la base de datos - réplica completa"""
    try:
        # Obtener todos los polígonos de la base de datos
        poligonos = Poligono.query.all()
        
        if not poligonos:
            flash('No hay datos disponibles para generar el archivo Excel', 'warning')
            return redirect(url_for('validacion_poligonos', tab='generar'))
        
        # Crear un DataFrame con TODAS las columnas de la base de datos
        data = []
        for p in poligonos:
            datos = {
                'ID_POLIGONO': p.id_poligono,
                'IF': p.if_val,
                'ID_CREDITO': p.id_credito,
                'ID_PERSONA': p.id_persona,
                'SUPERFICIE': p.superficie,  # Superficie original
                'ESTADO': corregir_codificacion(p.estado) if p.estado else '',
                'MUNICIPIO': corregir_codificacion(p.municipio) if p.municipio else '',
                'COORDENADAS': p.coordenadas if p.coordenadas else '',
                'COORDENADAS_DECIMALES_CORREGIDAS': p.coordenadas_corregidas if p.coordenadas_corregidas else '',
                'AREA_DIGITALIZADA': p.area_digitalizada if p.area_digitalizada else 0.0,
                'ESTATUS': p.estatus if p.estatus else '',
                'COMENTARIOS': p.comentarios if p.comentarios else '',
                'DESCRIPCION': p.descripcion if p.descripcion else '',
                'ORDEN': p.orden if p.orden else '',
                'SE_MODIFICO': p.se_modifico if p.se_modifico else 'No',
                'FECHA_CREACION': p.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if p.fecha_creacion else '',
                'FECHA_MODIFICACION': p.fecha_modificacion.strftime('%Y-%m-%d %H:%M:%S') if p.fecha_modificacion else '',
                'DB_ID': p.id  # ID interno de la base de datos
            }
            data.append(datos)
        
        # Crear un DataFrame de pandas con los datos
        df = pd.DataFrame(data)
        
        # Reordenar las columnas para que las más importantes estén primero
        columnas_ordenadas = [
            'ID_POLIGONO', 'IF', 'ID_CREDITO', 'ID_PERSONA', 'SUPERFICIE', 
            'ESTADO', 'MUNICIPIO', 'COORDENADAS', 'COORDENADAS_DECIMALES_CORREGIDAS',
            'AREA_DIGITALIZADA', 'ESTATUS', 'COMENTARIOS', 'DESCRIPCION', 'ORDEN',
            'SE_MODIFICO', 'FECHA_CREACION', 'FECHA_MODIFICACION', 'DB_ID'
        ]
        
        # Verificar que todas las columnas existen antes de reordenar
        columnas_existentes = [col for col in columnas_ordenadas if col in df.columns]
        df = df[columnas_existentes]
        
        # Crear un objeto BytesIO para guardar el Excel en memoria
        excel_file = io.BytesIO()
        
        # Guardar el DataFrame como un archivo Excel
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Polígonos_BD')
            
            # Obtener el workbook y worksheet para aplicar formato
            workbook = writer.book
            worksheet = writer.sheets['Polígonos_BD']
            
            # Aplicar formato a las cabeceras
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Aplicar estilo a la primera fila (cabeceras)
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Ajustar el ancho de las columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar el archivo para la descarga
        excel_file.seek(0)
        
        # Crear una respuesta con el archivo Excel
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f'base_datos_poligonos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        app.logger.error(f"ERROR AL GENERAR EXCEL: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error al generar archivo Excel: {str(e)}', 'error')
        return redirect(url_for('validacion_poligonos', tab='generar'))

# Import necessary modules for SHP handling
import zipfile
import tempfile
import os
import geopandas as gpd
import pandas as pd
from shapely.geometry import shape, Point, Polygon
import json
from werkzeug.utils import secure_filename

# ------- Validación Rápida SHP Routes -------

def safe_process_coordinates(geometry):
    """Safely process geometry coordinates for display in the map."""
    try:
        # For standard shapefile polygons
        if hasattr(geometry, 'exterior') and hasattr(geometry.exterior, 'coords'):
            # Handle polygon geometry
            exterior_coords = list(geometry.exterior.coords)
            # Only take the first two coordinates (x,y) and ignore z or other dimensions if present
            coords = [[y, x] for x, y in [(p[0], p[1]) for p in exterior_coords]]
            return coords
        
        # For MultiPolygon geometries
        elif hasattr(geometry, 'geoms'):
            # Use the first polygon in the multipolygon
            first_geom = geometry.geoms[0]
            if hasattr(first_geom, 'exterior'):
                exterior_coords = list(first_geom.exterior.coords)
                # Only take the first two coordinates (x,y) and ignore z or other dimensions
                coords = [[y, x] for x, y in [(p[0], p[1]) for p in exterior_coords]]
                return coords
        
        # For Point geometries
        elif geometry.geom_type == 'Point':
            return [[geometry.y, geometry.x]]  # [lat, lng] for Leaflet
        
        # For LineString geometries
        elif geometry.geom_type == 'LineString':
            line_coords = list(geometry.coords)
            # Only take the first two coordinates (x,y) and ignore z or other dimensions
            coords = [[y, x] for x, y in [(p[0], p[1]) for p in line_coords]]
            return coords
        
        # Return empty list if geometry type is not handled
        app.logger.debug(f"Geometry type not handled: {geometry.geom_type}")
        return []
        
    except Exception as e:
        app.logger.error(f"Error safely processing coordinates: {e}")
        import traceback
        traceback.print_exc()
        return []

@app.route('/validacion_rapida_shp')
@app.route('/validacion_rapida_shp/<tab>')
@login_required
def validacion_rapida_shp(tab=None):
    """
    Main route for the SHP validation functionality.
    Handles different tabs based on the parameter.
    """
    if not tab:
        tab = 'cargar'
    
    # Get available SHP files and data
    shp_data = []
    shp_columns = []
    shp_archivos = []
    
    # Query the database for SHP records
    if tab in ['lista', 'generar']:
        # Get filter parameter if available
        shp_filter = request.args.get('shp_filter')
        
        # Connect to database
        conn = get_db_connection()
        
        if shp_filter:
            # Filter by SHP file name
            shp_data = conn.execute('SELECT * FROM shp_records WHERE shp_origen = ?', 
                                   (shp_filter,)).fetchall()
        else:
            # Get all records
            shp_data = conn.execute('SELECT * FROM shp_records').fetchall()
        
        # Get list of unique SHP files
        shp_archivos = conn.execute(
            'SELECT DISTINCT shp_origen FROM shp_records ORDER BY shp_origen'
        ).fetchall()
        shp_archivos = [row['shp_origen'] for row in shp_archivos]
        
        # Get columns for display (excluding geometry data)
        if shp_data:
            shp_columns = [column for column in shp_data[0].keys() 
                          if column not in ['shp_id', 'geometry_wkt', 'atributos']]
        
        conn.close()
    
    # Handle Edit tab
    elif tab == 'editar':
        db_id = request.args.get('db_id')
        if db_id:
            try:
                # Buscar el polígono en la base de datos por su ID
                poligono = Poligono.query.get(int(db_id))
                
                if poligono is None:
                    flash('Polígono no encontrado', 'error')
                    return redirect(url_for('validacion_poligonos', tab='lista'))
                
                # Preparar coordenadas para el mapa
                coords_para_mapa = []
                if poligono.coordenadas_corregidas:
                    try:
                        # Parsear las coordenadas corregidas para el mapa
                        coord_pairs = poligono.coordenadas_corregidas.split(' | ')
                        for pair in coord_pairs:
                            if ',' in pair:
                                lat, lon = pair.split(',')
                                coords_para_mapa.append([float(lat.strip()), float(lon.strip())])
                    except Exception as e:
                        app.logger.error(f"Error al procesar coordenadas para el mapa: {e}")
                        coords_para_mapa = []
                
                # Detectar ubicación automáticamente si el estado y municipio están vacíos
                ubicacion_auto = False
                estado_detectado = poligono.estado
                municipio_detectado = poligono.municipio
                
                if (not estado_detectado or not municipio_detectado) and poligono.coordenadas_corregidas:
                    app.logger.debug("Detectando ubicación automáticamente...")
                    ubicacion = obtener_ubicacion_desde_poligono(poligono.coordenadas_corregidas)
                    if ubicacion:
                        if not estado_detectado:
                            estado_detectado = ubicacion['estado']
                            ubicacion_auto = True
                        if not municipio_detectado:
                            municipio_detectado = ubicacion['municipio']
                            ubicacion_auto = True
                        app.logger.debug(f"Ubicación detectada: {municipio_detectado}, {estado_detectado}")
                
                # Crear diccionario con datos del polígono para la plantilla
                poligono_data = {
                    'ID_POLIGONO': poligono.id_poligono,
                    'IF': poligono.if_val,
                    'ID_CREDITO': poligono.id_credito,
                    'ID_PERSONA': poligono.id_persona,
                    'SUPERFICIE': poligono.superficie,
                    'ESTADO': corregir_codificacion(estado_detectado) or '',
                    'MUNICIPIO': corregir_codificacion(municipio_detectado) or '',
                    'COORDENADAS': poligono.coordenadas,
                    'COORDENADAS_DECIMALES_CORREGIDAS': poligono.coordenadas_corregidas,  # Cambiado para coincidir con el template
                    'AREA_DIGITALIZADA': poligono.area_digitalizada,
                    'ESTATUS': poligono.estatus,
                    'COMENTARIOS': poligono.comentarios,
                    'DESCRIPCION': poligono.descripcion,
                    'db_id': poligono.id,
                    'UBICACION_AUTO': ubicacion_auto  # Bandera para mostrar que se detectó automáticamente
                }
                
                return render_template('validacion_poligonos.html', 
                                      tab=tab, 
                                      db_id=db_id,
                                      poligono_data=poligono_data, 
                                      coords_para_mapa=coords_para_mapa)
            except ValueError:
                flash('ID de polígono inválido', 'error')
                return redirect(url_for('validacion_poligonos', tab='lista'))
            except Exception as e:
                app.logger.error(f"Error al cargar polígono para edición: {e}")
                flash('Error al cargar el polígono para edición', 'error')
                return redirect(url_for('validacion_poligonos', tab='lista'))
        else:
            # Si no hay db_id, redirigir a la lista
            flash('No se especificó qué polígono editar', 'warning')
            return redirect(url_for('validacion_poligonos', tab='lista'))
    
    return render_template('validacion_rapida_shp.html', 
                          tab=tab, 
                          shp_data=shp_data,
                          shp_columns=shp_columns,
                          shp_archivos=shp_archivos)

@app.route('/cargar_shp_zip', methods=['POST'])
@login_required
def cargar_shp_zip():
    """
    Handle the upload of a ZIP file containing SHP files.
    Extract the ZIP, process each SHP, and store in database.
    """
    if 'archivo' not in request.files:
        flash('No se seleccionó ningún archivo')
        return redirect(url_for('validacion_rapida_shp', tab='cargar'))
    
    archivo = request.files['archivo']
    
    if archivo.filename == '':
        flash('No se seleccionó ningún archivo')
        return redirect(url_for('validacion_rapida_shp', tab='cargar'))
    
    if not archivo.filename.endswith('.zip'):
        flash('El archivo debe ser un archivo ZIP')
        return redirect(url_for('validacion_rapida_shp', tab='cargar'))
    
    # Create a temporary directory to extract the ZIP
    with tempfile.TemporaryDirectory() as temp_dir:
        zip_path = os.path.join(temp_dir, secure_filename(archivo.filename))
        archivo.save(zip_path)
        
        # Extract the ZIP file
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # Función recursiva para extraer ZIPs anidados
        def extract_nested_zips(directory):
            # Buscar todos los archivos ZIP en el directorio actual
            zip_files = []
            for root, dirs, files in os.walk(directory):
                for file in files:
                    if file.endswith('.zip'):
                        zip_files.append(os.path.join(root, file))
            
            # Si no hay archivos ZIP, terminar la recursión
            if not zip_files:
                return
            
            # Mantener un registro de los ZIPs ya procesados para evitar duplicados
            processed_zips = set()
            
            # Extraer cada archivo ZIP encontrado
            for zip_file in zip_files:
                try:
                    # Evitar procesar el mismo archivo ZIP más de una vez
                    if zip_file in processed_zips:
                        continue
                    
                    processed_zips.add(zip_file)
                    
                    # Crear un subdirectorio para la extracción basado en el nombre del ZIP
                    extract_dir = os.path.join(os.path.dirname(zip_file), 
                                              os.path.basename(zip_file).replace('.zip', ''))
                    os.makedirs(extract_dir, exist_ok=True)
                    
                    # Extraer el ZIP
                    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                        zip_ref.extractall(extract_dir)
                    
                    # Eliminar el ZIP original después de extraerlo para evitar duplicados
                    try:
                        os.remove(zip_file)
                    except Exception as remove_error:
                        app.logger.error(f"Error removing zip after extraction: {remove_error}")
                except Exception as e:
                    app.logger.error(f"Error extracting nested ZIP {zip_file}: {e}")
            
            # Buscar nuevos archivos ZIP después de la extracción
            new_zip_files = []
            for root, dirs, files in os.walk(directory):
                for file in files:
                    if file.endswith('.zip'):
                        new_zip_files.append(os.path.join(root, file))
            
            # Si hay nuevos archivos ZIP, extraerlos recursivamente
            if new_zip_files:
                extract_nested_zips(directory)
        
        # Extraer recursivamente todos los ZIPs anidados
        extract_nested_zips(temp_dir)
        
        # Find all SHP files in the extracted directory
        shp_files = []
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                if file.endswith('.shp'):
                    shp_files.append(os.path.join(root, file))
        
        if not shp_files:
            flash('No se encontraron archivos SHP en el ZIP')
            return redirect(url_for('validacion_rapida_shp', tab='cargar'))
        
        # Conjunto para rastrear geometrías ya procesadas y evitar duplicados
        processed_geometries = set()
        
        # Process each SHP file
        registros_procesados = 0
        
        conn = get_db_connection()
        
        for shp_file in shp_files:
            try:
                # Read the shapefile using GeoPandas
                gdf = gpd.read_file(shp_file)
                
                # Get the shapefile name without path and extension
                shp_origen = os.path.basename(shp_file).replace('.shp', '')
                
                # Process each record in the shapefile
                for idx, row in gdf.iterrows():
                    # Convert geometry to WKT (Well-Known Text) for storage
                    geometry_wkt = row.geometry.wkt
                    
                    # Calculate area in hectares for polygons
                    area = None
                    if row.geometry.geom_type in ['Polygon', 'MultiPolygon']:
                        # Convert to GeoSeries with correct CRS for area calculation
                        gs = gpd.GeoSeries([row.geometry], crs=gdf.crs)
                        # Convert to UTM for accurate area calculation
                        gs_utm = gs.to_crs('+proj=utm +zone=14 +datum=WGS84 +units=m +no_defs')
                        # Calculate area in hectares
                        area = gs_utm.area.values[0] / 10000  # m² to hectares
                    
                    # Store all other attributes as JSON
                    atributos = {}
                    for col in gdf.columns:
                        if col != 'geometry':
                            # Convert non-JSON serializable types
                            if isinstance(row[col], (int, float, str, bool)) or row[col] is None:
                                atributos[col] = row[col]
                            else:
                                atributos[col] = str(row[col])
                    
                    # Find ID field if available
                    id_campo = None
                    for key in ['ID', 'FID', 'OBJECTID', 'id', 'fid', 'objectid']:
                        if key in atributos:
                            id_campo = atributos[key]
                            break
                    
                    # Find municipality and state if available
                    municipio = None
                    estado = None
                    for key in ['MUNICIPIO', 'municipio', 'MUN', 'mun']:
                        if key in atributos:
                            municipio = atributos[key]
                            break
                    
                    for key in ['ESTADO', 'estado', 'EDO', 'edo']:
                        if key in atributos:
                            estado = atributos[key]
                            break
                    
                    # Crear una firma única para detectar duplicados (WKT + área + estado)
                    geometry_signature = f"{geometry_wkt}|{area}|{estado}"
                    
                    # Verificar si esta geometría ya fue procesada
                    if geometry_signature in processed_geometries:
                        app.logger.debug(f"Geometría duplicada detectada, saltando: {geometry_signature[:50]}...")
                        continue
                    
                    # Marcar esta geometría como procesada
                    processed_geometries.add(geometry_signature)
                    
                    # Insert into database
                    conn.execute('''
                        INSERT INTO shp_records 
                        (shp_origen, geometry_wkt, area, id_campo, atributos, municipio, estado, comentario, estatus)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        shp_origen,
                        geometry_wkt,
                        area,
                        id_campo,
                        json.dumps(atributos),
                        municipio,
                        estado,
                        '',  # Default empty comment
                        'no_aprobado'  # Default status is no_aprobado (previously 6)
                    ))
                    
                    registros_procesados += 1
            
            except Exception as e:
                flash(f'Error al procesar el archivo {os.path.basename(shp_file)}: {str(e)}')
                app.logger.error(f"Error processing SHP: {e}")
                conn.rollback()
                continue
        
        conn.commit()
        conn.close()
        
        if registros_procesados > 0:
            flash(f'Archivo procesado correctamente. {registros_procesados} registros importados.')
            return redirect(url_for('validacion_rapida_shp', tab='lista'))
        else:
            flash('No se pudo procesar ningún registro desde los archivos SHP.')
            return redirect(url_for('validacion_rapida_shp', tab='cargar'))

@app.route('/actualizar_shp_record', methods=['POST'])
@login_required
def actualizar_shp_record():
    """
    Update an SHP record with new information (comments, status, and geometry if provided).
    """
    shp_id = request.form.get('shp_id')
    comentario = request.form.get('comentario', '')
    estatus = request.form.get('estatus', '6')
    
    # Optional fields that might be editable
    municipio = request.form.get('municipio')
    estado = request.form.get('estado')
    
    # Get the area and wkt_geometry from the form
    area = request.form.get('area')
    wkt_geometry = request.form.get('wkt_geometry')
    
    # Get new coordinates if provided
    nuevas_coordenadas = request.form.get('nuevas_coordenadas')
    
    if not shp_id:
        flash('ID de registro no válido')
        return redirect(url_for('validacion_rapida_shp', tab='lista'))
    
    conn = get_db_connection()
    
    # First, get the current record to retain any values we're not updating
    current_record = conn.execute('SELECT * FROM shp_records WHERE shp_id = ?', (shp_id,)).fetchone()
    
    if not current_record:
        conn.close()
        flash('Registro no encontrado')
        return redirect(url_for('validacion_rapida_shp', tab='lista'))
    
    # Keep original values if not provided
    if municipio is None:
        municipio = current_record['municipio']
    if estado is None:
        estado = current_record['estado']
    
    # Update geometry and area if provided
    if wkt_geometry:
        try:
            # If we have a new geometry, update the record with it and the new area
            conn.execute('''
                UPDATE shp_records
                SET comentario = ?, estatus = ?, municipio = ?, estado = ?, 
                    geometry_wkt = ?, area = ?, nuevas_coordenadas = ?
                WHERE shp_id = ?
            ''', (comentario, estatus, municipio, estado, wkt_geometry, area, nuevas_coordenadas, shp_id))
        except Exception as e:
            conn.close()
            flash(f'Error al actualizar la geometría: {str(e)}')
            return redirect(url_for('validacion_rapida_shp', tab='editar', shp_id=shp_id))
    else:
        # Update just the attributes without changing geometry
        conn.execute('''
            UPDATE shp_records
            SET comentario = ?, estatus = ?, municipio = ?, estado = ?, nuevas_coordenadas = ?
            WHERE shp_id = ?
        ''', (comentario, estatus, municipio, estado, nuevas_coordenadas, shp_id))
    
    conn.commit()
    conn.close()
    
    flash('Registro actualizado correctamente')
    return redirect(url_for('validacion_rapida_shp', tab='lista'))

@app.route('/eliminar_shp_record', methods=['POST'])
@login_required
def eliminar_shp_record():
    """
    Delete an SHP record from the database.
    """
    shp_id = request.form.get('shp_id')
    
    if not shp_id:
        flash('ID de registro no válido')
        return redirect(url_for('validacion_rapida_shp', tab='lista'))
    
    conn = get_db_connection()
    conn.execute('DELETE FROM shp_records WHERE shp_id = ?', (shp_id,))
    conn.commit()
    conn.close()
    
    flash('Registro eliminado correctamente')
    return redirect(url_for('validacion_rapida_shp', tab='lista'))

@app.route('/exportar_shp_lista')
@login_required
def exportar_shp_lista():
    """
    Export the SHP records list to Excel.
    """
    # Get filter parameter if available
    shp_filter = request.args.get('shp_filter')
    
    conn = get_db_connection()
    
    if shp_filter:
        # Filter by SHP file name
        records = conn.execute('SELECT * FROM shp_records WHERE shp_origen = ?', 
                              (shp_filter,)).fetchall()
    else:
        # Get all records
        records = conn.execute('SELECT * FROM shp_records').fetchall()
    
    conn.close()
    
    if not records:
        flash('No hay registros para exportar')
        return redirect(url_for('validacion_rapida_shp', tab='lista'))
    
    # Create a DataFrame with the records
    data = []
    for record in records:
        # Skip geometry WKT to keep the excel clean
        row_data = {k: v for k, v in dict(record).items() if k != 'geometry_wkt'}
        
        # Parse the JSON attributes
        if 'atributos' in row_data and row_data['atributos']:
            try:
                atributos = json.loads(row_data['atributos'])
                for key, value in atributos.items():
                    row_data[f'attr_{key}'] = value
            except:
                pass
        
        data.append(row_data)
    
    df = pd.DataFrame(data)
    
    # Create a temporary file for the Excel
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        excel_path = temp_file.name
        df.to_excel(excel_path, index=False)
    
    # Return the file as attachment
    return send_file(
        excel_path,
        as_attachment=True,
        download_name='registros_shp.xlsx'
    )

@app.route('/generar_shp_archivos')
@login_required
def generar_shp_archivos():
    """
    Generate Excel file with all SHP records data
    """
    incluir_comentarios = request.args.get('incluir_comentarios', 'true') == 'true'
    filtro_estatus = request.args.get('filtro_estatus', 'todos')
    ids = request.args.get('ids', '')
    
    if ids:
        id_list = ids.split(',')
    else:
        flash('No se seleccionaron registros')
        return redirect(url_for('validacion_rapida_shp', tab='generar'))
    
    # Query the database for selected records
    conn = get_db_connection()
    
    placeholders = ','.join(['?'] * len(id_list))
    
    if filtro_estatus != 'todos':
        # Handle both old numeric and new string status values
        status_values = []
        if filtro_estatus == 'aprobado':
            status_values = ['7', 7, 'aprobado']  # Include both old and new values
        elif filtro_estatus == 'no_aprobado':
            status_values = ['6', 6, 'no_aprobado']  # Include both old and new values
        else:
            status_values = [filtro_estatus]  # Use the value as is
            
        # Create placeholders for status values
        status_placeholders = ','.join(['?'] * len(status_values))
        
        query = f'''
            SELECT * FROM shp_records 
            WHERE shp_id IN ({placeholders}) AND estatus IN ({status_placeholders})
        '''
        # Combine ID list and status values for the query parameters
        params = id_list + status_values
        records = conn.execute(query, params).fetchall()
    else:
        query = f'''
            SELECT * FROM shp_records 
            WHERE shp_id IN ({placeholders})
        '''
        records = conn.execute(query, id_list).fetchall()
    
    conn.close()
    
    if not records:
        flash('No se encontraron registros con los criterios seleccionados')
        return redirect(url_for('validacion_rapida_shp', tab='generar'))
    
    # Create Excel with all records
    data = []
    
    for record in records:
        # Skip geometry_wkt field to keep the Excel clean
        row_data = {k: v for k, v in dict(record).items() if k != 'geometry_wkt'}
        
        # Parse the JSON attributes and add them as individual columns
        if 'atributos' in row_data and row_data['atributos']:
            try:
                atributos = json.loads(row_data['atributos'])
                for key, value in atributos.items():
                    # Add attributes without prefix
                    row_data[key] = value
                
                # Remove the original JSON string to avoid duplication
                del row_data['atributos']
            except Exception as e:
                app.logger.error(f"Error parsing JSON attributes: {e}")
                # Keep the original column if parsing fails
        
        data.append(row_data)
    
    df = pd.DataFrame(data)
    
    # Create a temporary file for the Excel
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        excel_path = temp_file.name
        df.to_excel(excel_path, index=False)
    
    # Return the file as attachment
    return send_file(
        excel_path,
        as_attachment=True,
        download_name='registros_shp.xlsx'
    )

@app.route('/generar_shp_zip_completo')
@login_required
def generar_shp_zip_completo():
    """
    Generate a complete Excel file with all SHP records, expanding attributes into columns.
    """
    # Query all records from database
    conn = get_db_connection()
    records = conn.execute('SELECT * FROM shp_records').fetchall()
    conn.close()
    
    if not records:
        flash('No hay registros para procesar')
        return redirect(url_for('validacion_rapida_shp', tab='generar'))
    
    # Create Excel with all records
    data = []
    
    # Track all possible JSON attributes to ensure all records have all columns
    all_attributes = set()
    
    # First pass - extract all possible attribute names from all records
    for record in records:
        if record['atributos']:
            try:
                atributos = json.loads(record['atributos'])
                for key in atributos.keys():
                    all_attributes.add(key)
            except Exception as e:
                app.logger.error(f"Error parsing JSON attributes: {e}")
    
    # Second pass - create complete data rows with all attributes
    for record in records:
        # Skip geometry_wkt field to keep the Excel clean
        row_data = {k: v for k, v in dict(record).items() if k != 'geometry_wkt'}
        
        # Parse the JSON attributes and add them as individual columns
        attr_values = {}
        if 'atributos' in row_data and row_data['atributos']:
            try:
                atributos = json.loads(row_data['atributos'])
                # Initialize all possible attributes as None
                for attr in all_attributes:
                    attr_values[attr] = None
                
                # Set values for attributes present in this record
                for key, value in atributos.items():
                    attr_values[key] = value
                
                # Remove the original JSON string to avoid duplication
                del row_data['atributos']
            except Exception as e:
                app.logger.error(f"Error parsing JSON attributes for record {record['shp_id']}: {e}")
        
        # Combine base record data with attribute data
        row_data.update(attr_values)
        data.append(row_data)
    
    df = pd.DataFrame(data)
    
    # Create a temporary file for the Excel
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        excel_path = temp_file.name
        df.to_excel(excel_path, index=False)
    
    # Return the file as attachment
    return send_file(
        excel_path,
        as_attachment=True,
        download_name='todos_registros_shp.xlsx'
    )

# Add a database initialization function to create the SHP records table if it doesn't exist
def get_db_connection():
    """Create a connection to the SQLite database for SHP records."""
    import sqlite3
    conn = sqlite3.connect('shp_records.db')
    conn.row_factory = sqlite3.Row
    return conn

def init_shp_db():
    """
    Initialize the database table for SHP records if it doesn't exist.
    """
    conn = get_db_connection()
    
    # Create the table if it doesn't exist
    conn.execute('''
    CREATE TABLE IF NOT EXISTS shp_records (
        shp_id INTEGER PRIMARY KEY AUTOINCREMENT,
        shp_origen TEXT NOT NULL,
        geometry_wkt TEXT NOT NULL,
        area REAL,
        id_campo TEXT,
        atributos TEXT,
        municipio TEXT,
        estado TEXT,
        comentario TEXT,
        estatus TEXT,
        nuevas_coordenadas TEXT
    )
    ''')
    
    conn.commit()
    conn.close()

# Call the initialization function when the app starts
init_shp_db()

def init_validacion_15k_db():
    """
    Initialize the database table for 15K validation results if it doesn't exist.
    Reuses the same shp_records.db database as init_shp_db().
    """
    conn = get_db_connection()

    conn.execute('''
    CREATE TABLE IF NOT EXISTS validacion_15k (
        val_id INTEGER PRIMARY KEY AUTOINCREMENT,
        idx INTEGER NOT NULL UNIQUE,
        id_poligon_validacion TEXT,
        id_credito_validacion TEXT,
        nombre_zip TEXT,
        estatus TEXT DEFAULT 'pendiente',
        estatus_chapingo TEXT,
        id_poligono_unico TEXT,
        superficie_chapingo REAL,
        comentario_chapingo TEXT,
        id_poligon_historico TEXT,
        mega_idx INTEGER,
        overlap_pct REAL,
        fecha_validacion TIMESTAMP,
        validado_por TEXT DEFAULT 'usuario',
        superficie_calculada REAL
    )
    ''')

    existing_columns = {
        row['name']
        for row in conn.execute("PRAGMA table_info(validacion_15k)").fetchall()
    }
    required_columns = {
        'estatus_chapingo': 'TEXT',
        'id_poligono_unico': 'TEXT',
        'superficie_chapingo': 'REAL',
        'comentario_chapingo': 'TEXT',
        'superficie_calculada': 'REAL',
    }
    for column_name, column_type in required_columns.items():
        if column_name not in existing_columns:
            conn.execute(
                f'ALTER TABLE validacion_15k ADD COLUMN {column_name} {column_type}'
            )

    conn.commit()
    conn.close()

# Call the initialization function when the app starts
init_validacion_15k_db()

# Filtro personalizado para convertir strings JSON a diccionarios
@app.template_filter('ensure_dict')
def ensure_dict(value):
    """
    Asegura que el valor es un diccionario. Si es una cadena JSON, la parsea.
    Si ya es un diccionario, lo devuelve tal como está.
    Si es otra cosa, devuelve un diccionario vacío.
    """
    if isinstance(value, dict):
        return value
    elif isinstance(value, str):
        try:
            return json.loads(value)
        except (json.JSONDecodeError, ValueError):
            return {}
    else:
        return {}

@app.template_filter('clean_none')
def clean_none(value):
    """
    Limpiar valores None y convertirlos a strings vacíos
    """
    if value is None or str(value).lower() == 'none':
        return ''
    return str(value)

@app.route('/get-poligonos-actuales-traslapes/<int:polygon_id>')
@login_required
def get_poligonos_actuales_traslapes(polygon_id):
    """Endpoint para detectar y devolver polígonos actuales que traslapan con el polígono dado"""
    try:
        from shapely.geometry import Polygon
        import json
        
        # Buscar el polígono actual en la base de datos
        poligono_actual = Poligono.query.get(polygon_id)
        if poligono_actual is None:
            return jsonify({'error': 'Polígono no encontrado'}), 404
            
        # Obtener coordenadas del polígono actual
        coordenadas_corregidas = poligono_actual.coordenadas_corregidas
        if not coordenadas_corregidas:
            return jsonify({'error': 'El polígono no tiene coordenadas válidas'}), 400
        
        # Convertir coordenadas del polígono actual a geometría
        def coordenadas_a_geometria(coord_str):
            """Convierte string de coordenadas a geometría Shapely"""
            try:
                coords_list = coord_str.split(' | ')
                puntos = []
                for coord in coords_list:
                    if ',' in coord:
                        lat, lon = coord.strip().split(',')
                        puntos.append((float(lon.strip()), float(lat.strip())))  # Shapely usa (lon, lat)
                
                if len(puntos) >= 3:
                    return Polygon(puntos)
                else:
                    return None
            except Exception as e:
                app.logger.error(f"Error al convertir coordenadas a geometría: {e}")
                return None
        
        def coordenadas_a_geojson_coords(coord_str):
            """Convierte string de coordenadas a coordenadas de GeoJSON"""
            try:
                coords_list = coord_str.split(' | ')
                puntos = []
                for coord in coords_list:
                    if ',' in coord:
                        lat, lon = coord.strip().split(',')
                        puntos.append([float(lon.strip()), float(lat.strip())])  # GeoJSON usa [lon, lat]
                
                if len(puntos) >= 3:
                    # Cerrar el polígono si no está cerrado
                    if puntos[0] != puntos[-1]:
                        puntos.append(puntos[0])
                    return [puntos]  # GeoJSON Polygon requiere array de arrays
                else:
                    return None
            except Exception as e:
                app.logger.error(f"Error al convertir coordenadas a GeoJSON: {e}")
                return None
        
        geometria_actual = coordenadas_a_geometria(coordenadas_corregidas)
        if geometria_actual is None:
            return jsonify({'error': 'No se pudo procesar la geometría del polígono actual'}), 400
        
        # Obtener todos los demás polígonos de la base de datos
        otros_poligonos = Poligono.query.filter(Poligono.id != polygon_id).all()
        
        # Lista para almacenar polígonos que traslapan
        poligonos_traslapados = []
        features_geojson = []
        
        # Verificar traslapes con cada polígono
        for otro_poligono in otros_poligonos:
            if not otro_poligono.coordenadas_corregidas:
                continue
                
            otra_geometria = coordenadas_a_geometria(otro_poligono.coordenadas_corregidas)
            if otra_geometria is None:
                continue
            
            # Verificar si hay traslape
            try:
                if geometria_actual.intersects(otra_geometria) and not geometria_actual.touches(otra_geometria):
                    # Hay traslape (intersección pero no solo tocándose)
                    poligono_info = {
                        'id': otro_poligono.id,
                        'id_poligono': otro_poligono.id_poligono or f"DB_{otro_poligono.id}",
                        'area': otro_poligono.area_digitalizada or 0,
                        'estado': otro_poligono.estado or '',
                        'municipio': otro_poligono.municipio or ''
                    }
                    poligonos_traslapados.append(poligono_info)
                    
                    # Crear feature GeoJSON para este polígono
                    geojson_coords = coordenadas_a_geojson_coords(otro_poligono.coordenadas_corregidas)
                    if geojson_coords:
                        feature = {
                            "type": "Feature",
                            "properties": {
                                "id": otro_poligono.id,
                                "id_poligono": otro_poligono.id_poligono or f"DB_{otro_poligono.id}",
                                "area": otro_poligono.area_digitalizada or 0,
                                "estado": otro_poligono.estado or '',
                                "municipio": otro_poligono.municipio or '',
                                "tipo": "actual_traslapado"  # Identificador para aplicar estilo diferente
                            },
                            "geometry": {
                                "type": "Polygon",
                                "coordinates": geojson_coords
                            }
                        }
                        features_geojson.append(feature)
                        
            except Exception as e:
                app.logger.error(f"Error al verificar traslape con polígono {otro_poligono.id}: {e}")
                continue
        
        # Crear GeoJSON completo
        geojson_data = {
            "type": "FeatureCollection",
            "features": features_geojson
        }
        
        # Preparar respuesta
        respuesta = {
            'poligonos_traslapados': poligonos_traslapados,
            'total': len(poligonos_traslapados),
            'geojson': geojson_data
        }
        
        app.logger.debug(f"Polígonos actuales traslapados encontrados: {len(poligonos_traslapados)}")
        return jsonify(respuesta)
        
    except Exception as e:
        app.logger.error(f"Error al detectar traslapes entre polígonos actuales: {e}")
        return jsonify({'error': str(e)}), 500

def fix_encoding(val):
    """Fix latin-1 artifacts in strings (e.g. 'MichoacÃ¡n' -> 'Michoacán')."""
    if not isinstance(val, str):
        return val
    try:
        return val.encode('latin-1').decode('utf-8')
    except (UnicodeDecodeError, UnicodeEncodeError):
        return val


def enrich_with_location(gdf):
    """Agrega columnas ESTADO y MUNICIPIO al GeoDataFrame usando spatial join con municipios de México."""
    if shp_cache.municipios is None:
        gdf['ESTADO'] = None
        gdf['MUNICIPIO'] = None
        return gdf
    # Calcular centroide de cada polígono para el join
    gdf_copy = gdf.copy()
    gdf_copy['_centroid'] = gdf_copy.geometry.centroid
    # Crear GeoDataFrame temporal con centroides como geometría
    centroids_gdf = gpd.GeoDataFrame(gdf_copy, geometry='_centroid', crs=gdf.crs)
    # Spatial join con municipios
    joined = gpd.sjoin(centroids_gdf, shp_cache.municipios[['NOM_ENT', 'NOMGEO', 'geometry']], how='left', predicate='within')
    # Drop duplicates in case a centroid falls on a boundary between two municipalities
    joined = joined[~joined.index.duplicated(keep='first')]
    # Copiar resultados al gdf original
    gdf['ESTADO'] = joined['NOM_ENT'].values
    gdf['MUNICIPIO'] = joined['NOMGEO'].values
    # Limpiar encoding issues (latin-1 artifacts)
    for col in ['ESTADO', 'MUNICIPIO']:
        gdf[col] = gdf[col].apply(fix_encoding)
    return gdf


@app.route('/mapa-15k')
@login_required
def mapa_15k():
    return render_template('mapa_15k.html')


@app.route('/validacion-15k')
@login_required
def validacion_15k():
    return render_template('validacion_15k.html')


@app.route('/api/mapa-15k/validacion')
@login_required
def api_mapa_15k_validacion():
    try:
        gdf = gpd.read_file('data/VALIDACION_UNIFICADO.shp')
        if gdf.crs is not None and gdf.crs.to_epsg() != 4326:
            gdf = gdf.to_crs(epsg=4326)
        gdf['geometry'] = gdf['geometry'].simplify(tolerance=0.0001, preserve_topology=True)
        gdf = enrich_with_location(gdf)
        # Calcular estados_disponibles ANTES de filtrar
        estados_df = gdf.groupby('ESTADO').size().reset_index(name='count')
        estados_disponibles = [
            {'name': row['ESTADO'], 'count': int(row['count'])}
            for _, row in estados_df.iterrows()
            if row['ESTADO'] is not None
        ]
        estados_disponibles.sort(key=lambda x: x['name'] if x['name'] else '')
        # Leer query params opcionales
        estado = request.args.get('estado')
        municipio = request.args.get('municipio')
        id_credito = request.args.get('id_credito')
        id_poligono = request.args.get('id_poligono')
        # Calcular municipios_disponibles (del estado seleccionado, o todos)
        if estado:
            muns_gdf = gdf[gdf['ESTADO'] == estado]
        else:
            muns_gdf = gdf
        municipios_disponibles = sorted(muns_gdf['MUNICIPIO'].dropna().unique().tolist())
        # Aplicar filtros
        if estado:
            gdf = gdf[gdf['ESTADO'] == estado]
        if municipio:
            gdf = gdf[gdf['MUNICIPIO'] == municipio]
        if id_credito:
            gdf = gdf[gdf['ID_CREDITO'].astype(str).str.contains(id_credito, case=False, na=False)]
        if id_poligono:
            gdf = gdf[gdf['ID_POLIGON'].str.contains(id_poligono, case=False, na=False)]
        geojson_data = json.loads(gdf.to_json())
        return jsonify({
            'geojson': geojson_data,
            'total': len(gdf),
            'fields': ['ID_POLIGON', 'ID_CREDITO', 'NOMBRE_ZIP'],
            'estados_disponibles': estados_disponibles,
            'municipios_disponibles': municipios_disponibles
        })
    except Exception as e:
        app.logger.error(f"Error al cargar VALIDACION_UNIFICADO: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/mapa-15k/historico')
@login_required
def api_mapa_15k_historico():
    try:
        gdf = gpd.read_file('data/MEGA_CAPA_V1_OL.shp')
        if gdf.crs is not None and gdf.crs.to_epsg() != 4326:
            gdf = gdf.to_crs(epsg=4326)
        gdf['geometry'] = gdf['geometry'].simplify(tolerance=0.0001, preserve_topology=True)
        gdf = enrich_with_location(gdf)
        # Calcular estados_disponibles ANTES de filtrar
        estados_df = gdf.groupby('ESTADO').size().reset_index(name='count')
        estados_disponibles = [
            {'name': row['ESTADO'], 'count': int(row['count'])}
            for _, row in estados_df.iterrows()
            if row['ESTADO'] is not None
        ]
        estados_disponibles.sort(key=lambda x: x['name'] if x['name'] else '')
        # Leer query params opcionales
        estado = request.args.get('estado')
        municipio = request.args.get('municipio')
        id_credito = request.args.get('id_credito')
        id_poligono = request.args.get('id_poligono')
        # Calcular municipios_disponibles (del estado seleccionado, o todos)
        if estado:
            muns_gdf = gdf[gdf['ESTADO'] == estado]
        else:
            muns_gdf = gdf
        municipios_disponibles = sorted(muns_gdf['MUNICIPIO'].dropna().unique().tolist())
        # Aplicar filtros
        if estado:
            gdf = gdf[gdf['ESTADO'] == estado]
        if municipio:
            gdf = gdf[gdf['MUNICIPIO'] == municipio]
        if id_credito:
            gdf = gdf[gdf['ID_CREDITO'].astype(str).str.contains(id_credito, case=False, na=False)]
        if id_poligono:
            gdf = gdf[gdf['ID_POLIGON'].str.contains(id_poligono, case=False, na=False)]
        geojson_data = json.loads(gdf.to_json())
        return jsonify({
            'geojson': geojson_data,
            'total': len(gdf),
            'fields': ['ID_POLIGON', 'ID_CREDITO'],
            'estados_disponibles': estados_disponibles,
            'municipios_disponibles': municipios_disponibles
        })
    except Exception as e:
        app.logger.error(f"Error al cargar MEGA_CAPA_V1_OL: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/mapa-15k/estados')
@login_required
def api_mapa_15k_estados():
    """Retorna lista de estados y municipios disponibles en los shapefiles."""
    try:
        gdf = gpd.read_file('data/VALIDACION_UNIFICADO.shp')
        if gdf.crs is not None and gdf.crs.to_epsg() != 4326:
            gdf = gdf.to_crs(epsg=4326)
        gdf = enrich_with_location(gdf)
        # Estados
        estados = gdf.groupby('ESTADO').size().reset_index(name='count')
        estados_list = [
            {'name': row['ESTADO'], 'count': int(row['count'])}
            for _, row in estados.iterrows()
            if row['ESTADO'] is not None
        ]
        estados_list.sort(key=lambda x: x['name'] if x['name'] else '')
        # Municipios agrupados por estado
        municipios = {}
        for estado in gdf['ESTADO'].dropna().unique():
            muns = gdf[gdf['ESTADO'] == estado]['MUNICIPIO'].dropna().unique().tolist()
            municipios[estado] = sorted(muns)
        return jsonify({'estados': estados_list, 'municipios_por_estado': municipios})
    except Exception as e:
        app.logger.error(f'Error en /api/mapa-15k/estados: {e}')
        return jsonify({'error': str(e)}), 500


def clasificar_traslape(overlap_pct, area_ratio, same_credit):
    """Clasifica el traslape según las reglas de negocio.

    Returns: (clasificacion, color, descripcion)
    """
    if same_credit:
        if area_ratio >= 85 and overlap_pct >= 85:
            return 'duplicado', '#dc3545', 'Duplicado (mismo crédito)'
        elif overlap_pct >= 10:
            return 'traslape_interno', '#ffc107', 'Traslape interno - rechazar'
        else:
            return 'sin_conflicto', '#28a745', 'Sin conflicto'
    else:
        if area_ratio >= 85 and overlap_pct >= 85:
            return 'duplicado', '#dc3545', 'Duplicado (diferente crédito)'
        elif 30 <= overlap_pct <= 80:
            return 'traslape_relevante', '#fd7e14', 'Traslape relevante - revisar'
        else:
            return 'sin_conflicto', '#28a745', 'Sin conflicto'


def _build_nuevos_relacionados_cache():
    """Build/cache spatial structures for validacion polygons marked as nuevos."""
    global _nuevos_relacionados_cache, _indices_filtrados_cache

    if shp_cache.validacion is None or shp_cache.mega is None:
        raise RuntimeError('Shapefiles no cargados')

    if _nuevos_relacionados_cache is not None:
        return _nuevos_relacionados_cache

    if _indices_filtrados_cache is not None and 'nuevos' in _indices_filtrados_cache:
        nuevos_indices = [int(i) for i in _indices_filtrados_cache['nuevos']]
    else:
        mega_ids = set(shp_cache.mega['ID_POLIGON'].astype(str).str.strip())
        mask_nuevos = ~shp_cache.validacion['ID_POLIGON'].astype(str).str.strip().isin(mega_ids)
        nuevos_indices = [int(i) for i in shp_cache.validacion.index[mask_nuevos]]

    nuevos_gdf = shp_cache.validacion.iloc[nuevos_indices].copy()
    nuevos_gdf['__src_idx__'] = nuevos_indices

    try:
        nuevos_sindex = nuevos_gdf.sindex
    except Exception:
        nuevos_sindex = None

    _nuevos_relacionados_cache = {
        'nuevos_indices_set': set(nuevos_indices),
        'nuevos_gdf': nuevos_gdf,
        'nuevos_sindex': nuevos_sindex,
    }
    return _nuevos_relacionados_cache


def generar_propuesta_chapingo_nuevo(idx, analisis_mega=None):
    """Genera propuesta automatica Chapingo para un poligono nuevo basada SOLO en megacapa."""
    if shp_cache.validacion is None or shp_cache.mega is None:
        raise RuntimeError('Shapefiles no cargados')
    if idx < 0 or idx >= len(shp_cache.validacion):
        raise ValueError(f'Indice fuera de rango (0-{len(shp_cache.validacion)-1})')

    base_row = shp_cache.validacion.iloc[idx]
    id_poligono_base = str(base_row.get('ID_POLIGON', '') or '').strip()

    if analisis_mega is None:
        analisis_mega = calcular_traslapes(idx)

    # Calculate surface of current polygon
    poligono_props = (analisis_mega.get('poligono') or {}).get('properties') or {}
    superficie_base = poligono_props.get('area_ha')
    superficie_base = round(float(superficie_base), 4) if superficie_base is not None else None

    propuesta = {
        'estatus_chapingo_propuesto': None,
        'id_poligono_unico_propuesto': None,
        'superficie_chapingo_propuesta': superficie_base,
        'comentario_chapingo_propuesto': None,
    }

    # Check if idx belongs to nuevos
    cache = _build_nuevos_relacionados_cache()
    if idx not in cache['nuevos_indices_set']:
        propuesta['comentario_chapingo_propuesto'] = 'El indice no pertenece al subconjunto nuevos.'
        return propuesta

    matches_mega = analisis_mega.get('matches') or []

    # No matches in megacapa → NUEVO
    if not matches_mega:
        propuesta['estatus_chapingo_propuesto'] = 'NUEVO'
        propuesta['id_poligono_unico_propuesto'] = id_poligono_base
        propuesta['comentario_chapingo_propuesto'] = 'Sin traslape con Mega Capa. Poligono nuevo.'
        return propuesta

    # Find the worst (most severe) match classification
    # Priority: duplicado (VINCULAR) > traslape_interno (ELIMINAR) > traslape_relevante (REVISAR) > sin_conflicto (NUEVO)
    worst_match = None
    worst_priority = -1

    priority_map = {
        'duplicado': 3,
        'traslape_interno': 2,
        'traslape_relevante': 1,
        'sin_conflicto': 0,
    }

    for m in matches_mega:
        p = priority_map.get(m.get('clasificacion', ''), -1)
        if p > worst_priority:
            worst_priority = p
            worst_match = m

    if worst_match is None:
        propuesta['estatus_chapingo_propuesto'] = 'NUEVO'
        propuesta['id_poligono_unico_propuesto'] = id_poligono_base
        propuesta['comentario_chapingo_propuesto'] = 'Sin clasificacion determinante en Mega Capa.'
        return propuesta

    clasificacion = worst_match.get('clasificacion', '')
    mega_id_poligon = worst_match.get('id_poligon', '')
    mega_overlap = worst_match.get('overlap_pct', 0)
    mega_area_ratio = worst_match.get('area_ratio', 0)
    mega_same_credit = worst_match.get('same_credit', False)

    if clasificacion == 'duplicado':
        # Superficie >=85% + traslape >=85% → VINCULAR al ID_POLIGONO de mega
        propuesta['estatus_chapingo_propuesto'] = 'VINCULAR'
        propuesta['id_poligono_unico_propuesto'] = mega_id_poligon
        credito_tipo = 'mismo credito' if mega_same_credit else 'diferente credito'
        propuesta['comentario_chapingo_propuesto'] = (
            f'Duplicado en Mega ({credito_tipo}). '
            f'Traslape: {mega_overlap:.1f}%, similitud superficie: {mega_area_ratio:.1f}%. '
            f'Vinculado a {mega_id_poligon}.'
        )

    elif clasificacion == 'traslape_interno':
        # Mismo credito, traslape 10-85% → ELIMINAR
        propuesta['estatus_chapingo_propuesto'] = 'ELIMINAR'
        propuesta['id_poligono_unico_propuesto'] = mega_id_poligon
        propuesta['comentario_chapingo_propuesto'] = (
            f'Traslape interno con Mega (mismo credito). '
            f'Traslape: {mega_overlap:.1f}%. Se rechaza.'
        )

    elif clasificacion == 'traslape_relevante':
        # Diferente credito, traslape 30-80% → Revisar (no se asigna estatus automatico)
        propuesta['estatus_chapingo_propuesto'] = None
        propuesta['id_poligono_unico_propuesto'] = mega_id_poligon
        propuesta['comentario_chapingo_propuesto'] = (
            f'Traslape relevante con Mega (diferente credito). '
            f'Traslape: {mega_overlap:.1f}%. Revisar manualmente.'
        )

    else:
        # sin_conflicto → NUEVO
        propuesta['estatus_chapingo_propuesto'] = 'NUEVO'
        propuesta['id_poligono_unico_propuesto'] = id_poligono_base
        propuesta['comentario_chapingo_propuesto'] = (
            f'Traslape minimo con Mega ({mega_overlap:.1f}%). Sin conflicto.'
        )

    return propuesta


def construir_evidencia_flujo_chapingo():
    """Build reproducible evidence for key Chapingo flow scenarios."""
    if shp_cache.validacion is None or shp_cache.mega is None:
        raise RuntimeError('Shapefiles no cargados')

    cache = _build_nuevos_relacionados_cache()
    nuevos_indices = list(cache['nuevos_indices_set'])

    escenarios = {
        'nuevo_sin_match': None,
        'vincular_mismo_credito': None,
        'vincular_diferente_credito': None,
        'eliminar_traslape_interno': None,
    }

    for idx in nuevos_indices[:200]:
        try:
            data = calcular_traslapes(idx)
            matches = data.get('matches', [])

            if not matches and escenarios['nuevo_sin_match'] is None:
                escenarios['nuevo_sin_match'] = {'idx': idx, 'cumple': True}

            for m in matches:
                clasif = m.get('clasificacion', '')
                if clasif == 'duplicado' and m.get('same_credit') and escenarios['vincular_mismo_credito'] is None:
                    escenarios['vincular_mismo_credito'] = {'idx': idx, 'cumple': True, 'mega_id': m.get('id_poligon')}
                elif clasif == 'duplicado' and not m.get('same_credit') and escenarios['vincular_diferente_credito'] is None:
                    escenarios['vincular_diferente_credito'] = {'idx': idx, 'cumple': True, 'mega_id': m.get('id_poligon')}
                elif clasif == 'traslape_interno' and escenarios['eliminar_traslape_interno'] is None:
                    escenarios['eliminar_traslape_interno'] = {'idx': idx, 'cumple': True, 'mega_id': m.get('id_poligon')}

            if all(v is not None for v in escenarios.values()):
                break
        except Exception:
            continue

    return escenarios


@app.route('/api/analizador/total')
@login_required
def api_analizador_total():
    if shp_cache.validacion is None:
        return jsonify({'error': 'Shapefiles no cargados'}), 500
    return jsonify({'total': len(shp_cache.validacion)})


def calcular_traslapes(idx):
    """Helper: computes overlap analysis for shp_cache.validacion[idx] against shp_cache.mega.

    Returns a dict with keys: poligono, matches, match_features, resumen.
    Raises ValueError if idx is out of range.
    Raises RuntimeError if shapefiles are not loaded.
    """
    if shp_cache.validacion is None or shp_cache.mega is None:
        raise RuntimeError('Shapefiles no cargados')
    if idx < 0 or idx >= len(shp_cache.validacion):
        raise ValueError(f'Índice fuera de rango (0-{len(shp_cache.validacion)-1})')

    import pyproj
    import shapely
    from shapely.ops import transform

    vrow = shp_cache.validacion.iloc[idx]
    vgeom = vrow.geometry

    # Calcular área en hectáreas (proyectar a UTM zona basada en el centroide)
    centroid = vgeom.centroid
    utm_zone = int((centroid.x + 180) / 6) + 1
    transformer = pyproj.Transformer.from_crs('EPSG:4326', f'EPSG:326{utm_zone:02d}', always_xy=True)
    vgeom_utm = transform(transformer.transform, vgeom)
    area_ha = vgeom_utm.area / 10000

    # Buscar matches en MEGA usando spatial index
    candidates = list(shp_cache.mega.sindex.intersection(vgeom.bounds))
    matches = []
    match_features = []

    for ci in candidates:
        mrow = shp_cache.mega.iloc[ci]
        mgeom = mrow.geometry

        if not vgeom.intersects(mgeom):
            continue

        intersection = vgeom.intersection(mgeom)

        # Calcular áreas en UTM
        mgeom_utm = transform(transformer.transform, mgeom)
        intersection_utm = transform(transformer.transform, intersection)

        area_v = vgeom_utm.area
        area_m = mgeom_utm.area
        area_inter = intersection_utm.area

        overlap_pct = (area_inter / area_v * 100) if area_v > 0 else 0
        area_ratio = (min(area_v, area_m) / max(area_v, area_m) * 100) if max(area_v, area_m) > 0 else 0

        same_credit = str(vrow.get('ID_CREDITO', '')) == str(mrow.get('ID_CREDITO', ''))
        clasificacion, color, descripcion = clasificar_traslape(overlap_pct, area_ratio, same_credit)

        # Solo incluir matches con overlap > 0.1%
        if overlap_pct < 0.1:
            continue

        match_info = {
            'mega_index': int(ci),
            'id_poligon': str(mrow.get('ID_POLIGON', '')),
            'id_credito': str(mrow.get('ID_CREDITO', '')),
            'area_ha': round(area_m / 10000, 4),
            'overlap_pct': round(overlap_pct, 1),
            'area_ratio': round(area_ratio, 1),
            'same_credit': same_credit,
            'clasificacion': clasificacion,
            'color': color,
            'descripcion': descripcion
        }
        matches.append(match_info)

        match_feature = {
            'type': 'Feature',
            'properties': match_info,
            'geometry': json.loads(shapely.to_geojson(mgeom))
        }
        match_features.append(match_feature)

    # Ordenar matches por overlap descendente
    matches.sort(key=lambda x: x['overlap_pct'], reverse=True)
    match_features.sort(key=lambda x: x['properties']['overlap_pct'], reverse=True)

    # GeoJSON del polígono actual
    poligono_geojson = {
        'type': 'Feature',
        'properties': {
            'ID_POLIGON': str(vrow.get('ID_POLIGON', '')),
            'ID_CREDITO': str(vrow.get('ID_CREDITO', '')),
            'NOMBRE_ZIP': str(vrow.get('NOMBRE_ZIP', '')),
            'ESTATUS': obtener_estatus_validacion(vrow),
            'area_ha': round(area_ha, 4)
        },
        'geometry': json.loads(shapely.to_geojson(vgeom))
    }

    # Resumen de clasificaciones con desglose mismo/diferente crédito
    duplicados = [m for m in matches if m['clasificacion'] == 'duplicado']
    traslape_interno_list = [m for m in matches if m['clasificacion'] == 'traslape_interno']
    traslape_relevante_list = [m for m in matches if m['clasificacion'] == 'traslape_relevante']
    sin_conflicto_list = [m for m in matches if m['clasificacion'] == 'sin_conflicto']

    resumen = {
        'duplicados': len(duplicados),
        'duplicados_mismo_credito': sum(1 for m in duplicados if m.get('same_credit')),
        'duplicados_diferente_credito': sum(1 for m in duplicados if not m.get('same_credit')),
        'traslape_interno': len(traslape_interno_list),
        'traslape_interno_mismo_credito': sum(1 for m in traslape_interno_list if m.get('same_credit')),
        'traslape_interno_diferente_credito': sum(1 for m in traslape_interno_list if not m.get('same_credit')),
        'traslape_relevante': len(traslape_relevante_list),
        'traslape_relevante_mismo_credito': sum(1 for m in traslape_relevante_list if m.get('same_credit')),
        'traslape_relevante_diferente_credito': sum(1 for m in traslape_relevante_list if not m.get('same_credit')),
        'sin_conflicto': len(sin_conflicto_list),
        'sin_conflicto_mismo_credito': sum(1 for m in sin_conflicto_list if m.get('same_credit')),
        'sin_conflicto_diferente_credito': sum(1 for m in sin_conflicto_list if not m.get('same_credit')),
        'total_matches': len(matches)
    }

    return {
        'poligono': poligono_geojson,
        'matches': matches,
        'match_features': {
            'type': 'FeatureCollection',
            'features': match_features
        },
        'resumen': resumen
    }


def obtener_guardado_validacion_15k(idx):
    """Return persisted 15K validation row mapped to stable keys."""
    conn = get_db_connection()
    try:
        row = conn.execute(
            'SELECT * FROM validacion_15k WHERE idx = ?', (idx,)
        ).fetchone()
    finally:
        conn.close()

    if row is None:
        return {
            'estatus': 'pendiente',
            'estatus_chapingo': None,
            'id_poligono_unico': None,
            'superficie_chapingo': None,
            'comentario_chapingo': None,
            'id_poligon_historico': None,
            'mega_idx': None,
            'overlap_pct': None,
            'fecha_validacion': None,
            'superficie_calculada': None
        }

    return {
        'estatus': row['estatus'],
        'estatus_chapingo': row['estatus_chapingo'],
        'id_poligono_unico': row['id_poligono_unico'],
        'superficie_chapingo': row['superficie_chapingo'],
        'comentario_chapingo': row['comentario_chapingo'],
        'id_poligon_historico': row['id_poligon_historico'],
        'mega_idx': row['mega_idx'],
        'overlap_pct': row['overlap_pct'],
        'fecha_validacion': row['fecha_validacion'],
        'superficie_calculada': row['superficie_calculada']
    }


def indice_pertenece_a_nuevos(idx):
    """Return True when idx belongs to the computed subset 'nuevos'."""
    cache = _build_nuevos_relacionados_cache()
    return int(idx) in cache['nuevos_indices_set']


def _normalizar_texto_chapingo(value, field_name):
    """Normalize optional Chapingo text values and validate payload types."""
    if value is None:
        return None
    if isinstance(value, (dict, list, tuple, set)):
        raise ValueError(f'{field_name} debe ser texto')
    text = str(value).strip()
    return text or None


def _normalizar_superficie_chapingo(value):
    """Validate and normalize optional Chapingo area value."""
    if value in (None, ''):
        return None

    try:
        superficie = float(value)
    except (TypeError, ValueError):
        raise ValueError('superficie_chapingo debe ser numerica positiva')

    if math.isnan(superficie) or math.isinf(superficie) or superficie <= 0:
        raise ValueError('superficie_chapingo debe ser numerica positiva')

    return round(superficie, 4)


def _requiere_vinculacion_chapingo(idx, estatus_chapingo):
    """Determine if an index requires linked polygon id for Chapingo decision."""
    if estatus_chapingo == 'VINCULAR':
        return True

    try:
        propuesta = generar_propuesta_chapingo_nuevo(idx)
    except Exception:
        return False

    return propuesta.get('estatus_chapingo_propuesto') == 'VINCULAR'


@app.route('/api/analizador/poligono/<int:idx>')
@login_required
def api_analizador_poligono(idx):
    try:
        data = calcular_traslapes(idx)
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    return jsonify({
        'index': idx,
        'total': len(shp_cache.validacion),
        'poligono': data['poligono'],
        'matches': data['matches'],
        'match_features': data['match_features'],
        'resumen': data['resumen']
    })


@app.route('/api/analizador/propuesta-editable/<int:idx>')
@login_required
def api_analizador_propuesta_editable(idx):
    """Return complete editable payload for Analizador proposal panel."""
    try:
        analisis_mega = calcular_traslapes(idx)
        es_nuevo = indice_pertenece_a_nuevos(idx)
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    guardado = obtener_guardado_validacion_15k(idx)

    if es_nuevo:
        propuesta = generar_propuesta_chapingo_nuevo(idx, analisis_mega=analisis_mega)
        propuesta['aplica'] = True
    else:
        propuesta = {
            'aplica': False,
            'estatus_chapingo_propuesto': None,
            'id_poligono_unico_propuesto': None,
            'superficie_chapingo_propuesta': None,
            'comentario_chapingo_propuesto': 'El indice no pertenece al subconjunto nuevos; propuesta automatica no aplica.',
        }

    return jsonify({
        'index': idx,
        'total': len(shp_cache.validacion),
        'es_nuevo': es_nuevo,
        'poligono': analisis_mega['poligono'],
        'analisis_mega': {
            'matches': analisis_mega['matches'],
            'match_features': analisis_mega['match_features'],
            'resumen': analisis_mega['resumen'],
        },
        'propuesta': propuesta,
        'guardado': guardado,
    })


@app.route('/api/analizador/chapingo-evidencia')
@login_required
def api_analizador_chapingo_evidencia():
    """Expose reproducible evidence for critical Chapingo flow scenarios."""
    try:
        escenarios = construir_evidencia_flujo_chapingo()
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        return jsonify({'error': f'No fue posible generar evidencia Chapingo: {str(e)}'}), 500

    faltantes = [k for k, v in escenarios.items() if v is None]
    escenarios_fallidos = [
        k for k, v in escenarios.items()
        if isinstance(v, dict) and v.get('cumple') is False
    ]

    return jsonify({
        'ok': not faltantes and not escenarios_fallidos,
        'escenarios': escenarios,
        'faltantes': faltantes,
        'fallidos': escenarios_fallidos,
    })


@app.route('/api/analizador/propuesta-editable/<int:idx>/guardar', methods=['POST'])
@login_required
def api_analizador_guardar_propuesta_editable(idx):
    """Save editable Chapingo decision values for a single index."""
    if shp_cache.validacion is None:
        return jsonify({'error': 'Shapefiles no cargados'}), 500
    if idx < 0 or idx >= len(shp_cache.validacion):
        return jsonify({'error': f'idx fuera de rango (0-{len(shp_cache.validacion)-1})'}), 400

    data = request.get_json(force=True, silent=True) or {}

    estatus_raw = data.get('estatus_chapingo')
    if estatus_raw is None:
        return jsonify({'error': 'estatus_chapingo es requerido'}), 400

    estatus_chapingo = str(estatus_raw).strip().upper()
    if not estatus_chapingo:
        return jsonify({'error': 'estatus_chapingo no puede estar vacio'}), 400

    try:
        id_poligono_unico = _normalizar_texto_chapingo(
            data.get('id_poligono_unico'),
            'id_poligono_unico'
        )
        comentario_chapingo = _normalizar_texto_chapingo(
            data.get('comentario_chapingo'),
            'comentario_chapingo'
        )
        superficie_chapingo = _normalizar_superficie_chapingo(
            data.get('superficie_chapingo')
        )
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    if _requiere_vinculacion_chapingo(idx, estatus_chapingo) and not id_poligono_unico:
        return jsonify({
            'error': 'id_poligono_unico es requerido cuando la propuesta requiere vinculacion'
        }), 400

    # Auto-calculate surface area in hectares using UTM projection
    import pyproj
    from shapely.ops import transform
    vrow = shp_cache.validacion.iloc[idx]
    vgeom = vrow.geometry
    centroid = vgeom.centroid
    utm_zone = int((centroid.x + 180) / 6) + 1
    transformer = pyproj.Transformer.from_crs('EPSG:4326', f'EPSG:326{utm_zone:02d}', always_xy=True)
    vgeom_utm = transform(transformer.transform, vgeom)
    superficie_calculada = round(vgeom_utm.area / 10000, 4)

    conn = get_db_connection()
    try:
        conn.execute(
            '''INSERT INTO validacion_15k
               (idx, estatus_chapingo, id_poligono_unico, superficie_chapingo, comentario_chapingo, superficie_calculada)
               VALUES (?, ?, ?, ?, ?, ?)
               ON CONFLICT(idx) DO UPDATE SET
                   estatus_chapingo = excluded.estatus_chapingo,
                   id_poligono_unico = excluded.id_poligono_unico,
                   superficie_chapingo = excluded.superficie_chapingo,
                   comentario_chapingo = excluded.comentario_chapingo,
                   superficie_calculada = excluded.superficie_calculada''',
            (
                idx,
                estatus_chapingo,
                id_poligono_unico,
                superficie_chapingo,
                comentario_chapingo,
                superficie_calculada,
            )
        )
        conn.commit()
    finally:
        conn.close()

    guardado = obtener_guardado_validacion_15k(idx)
    return jsonify({
        'success': True,
        'idx': idx,
        'guardado': {
            'estatus_chapingo': guardado['estatus_chapingo'],
            'id_poligono_unico': guardado['id_poligono_unico'],
            'superficie_chapingo': guardado['superficie_chapingo'],
            'comentario_chapingo': guardado['comentario_chapingo'],
            'superficie_calculada': guardado['superficie_calculada'],
        }
    })


@app.route('/api/analizador/buscar')
@login_required
def api_analizador_buscar():
    if shp_cache.validacion is None:
        return jsonify({'error': 'Shapefiles no cargados'}), 500
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({'resultados': []})
    resultados = []
    for idx, row in shp_cache.validacion.iterrows():
        if (q.lower() in str(row.get('ID_CREDITO', '')).lower() or
                q.lower() in str(row.get('ID_POLIGON', '')).lower()):
            resultados.append({
                'index': int(idx),
                'id_poligon': str(row.get('ID_POLIGON', '')),
                'id_credito': str(row.get('ID_CREDITO', ''))
            })
        if len(resultados) >= 50:
            break
    return jsonify({'resultados': resultados, 'total': len(resultados)})


@app.route('/api/analizador/dashboard-estatus')
@login_required
def api_analizador_dashboard_estatus():
    global _dashboard_cache
    if shp_cache.validacion is None or shp_cache.mega is None:
        return jsonify({'error': 'Shapefiles no cargados'}), 500
    if _dashboard_cache is not None:
        return jsonify(_dashboard_cache)
    mega_ids = set(shp_cache.mega['ID_POLIGON'].astype(str).str.strip())
    total_15k = len(shp_cache.validacion)
    nuevos = int((~shp_cache.validacion['ID_POLIGON'].astype(str).str.strip().isin(mega_ids)).sum())
    existentes = total_15k - nuevos
    _dashboard_cache = {
        'total_15k': total_15k,
        'nuevos': nuevos,
        'existentes': existentes,
    }
    return jsonify(_dashboard_cache)


@app.route('/api/analizador/indices-filtrados')
@login_required
def api_analizador_indices_filtrados():
    global _indices_filtrados_cache
    if shp_cache.validacion is None or shp_cache.mega is None:
        return jsonify({'error': 'Shapefiles no cargados'}), 500
    filtro = request.args.get('filtro')
    if filtro not in ('nuevos', 'existentes'):
        return jsonify({'error': 'Parámetro filtro inválido. Use nuevos o existentes'}), 400
    if _indices_filtrados_cache is None:
        mega_ids = set(shp_cache.mega['ID_POLIGON'].astype(str).str.strip())
        mask_nuevos = ~shp_cache.validacion['ID_POLIGON'].astype(str).str.strip().isin(mega_ids)
        indices_nuevos = [int(i) for i in shp_cache.validacion.index[mask_nuevos]]
        indices_existentes = [int(i) for i in shp_cache.validacion.index[~mask_nuevos]]
        _indices_filtrados_cache = {
            'nuevos': indices_nuevos,
            'existentes': indices_existentes,
        }
    indices = _indices_filtrados_cache[filtro]
    return jsonify({'filtro': filtro, 'indices': indices, 'total': len(indices)})


def _run_clasificacion_nuevos():
    """Background thread: classifies all 'nuevo' polygons by their worst overlap category."""
    global _clasif_nuevos_state, _indices_filtrados_cache
    try:
        # Get or compute the list of 'nuevo' indices
        if _indices_filtrados_cache is None:
            mega_ids = set(shp_cache.mega['ID_POLIGON'].astype(str).str.strip())
            mask_nuevos = ~shp_cache.validacion['ID_POLIGON'].astype(str).str.strip().isin(mega_ids)
            indices_nuevos = [int(i) for i in shp_cache.validacion.index[mask_nuevos]]
            indices_existentes = [int(i) for i in shp_cache.validacion.index[~mask_nuevos]]
            _indices_filtrados_cache = {
                'nuevos': indices_nuevos,
                'existentes': indices_existentes,
            }
        indices = _indices_filtrados_cache['nuevos']
        total = len(indices)
        _clasif_nuevos_state['total'] = total

        counts = {
            'duplicado': 0,
            'duplicados_mismo_credito': 0,
            'duplicados_diferente_credito': 0,
            'traslape_interno': 0,
            'traslape_interno_mismo_credito': 0,
            'traslape_interno_diferente_credito': 0,
            'traslape_relevante': 0,
            'traslape_relevante_mismo_credito': 0,
            'traslape_relevante_diferente_credito': 0,
            'sin_conflicto': 0,
            'sin_conflicto_mismo_credito': 0,
            'sin_conflicto_diferente_credito': 0,
            'sin_matches': 0,
        }
        indices_por_clasif = {
            'duplicado': [],
            'duplicado_mismo_credito': [],
            'duplicado_diferente_credito': [],
            'traslape_interno': [],
            'traslape_interno_mismo_credito': [],
            'traslape_interno_diferente_credito': [],
            'traslape_relevante': [],
            'traslape_relevante_mismo_credito': [],
            'traslape_relevante_diferente_credito': [],
            'sin_conflicto': [],
            'sin_conflicto_mismo_credito': [],
            'sin_conflicto_diferente_credito': [],
            'sin_matches': [],
        }

        for i, idx in enumerate(indices):
            try:
                data = calcular_traslapes(idx)
                resumen = data['resumen']
                if resumen['duplicados'] > 0:
                    counts['duplicado'] += 1
                    indices_por_clasif['duplicado'].append(idx)
                    if resumen.get('duplicados_mismo_credito', 0) > 0:
                        counts['duplicados_mismo_credito'] += 1
                        indices_por_clasif['duplicado_mismo_credito'].append(idx)
                    else:
                        counts['duplicados_diferente_credito'] += 1
                        indices_por_clasif['duplicado_diferente_credito'].append(idx)
                elif resumen['traslape_interno'] > 0:
                    counts['traslape_interno'] += 1
                    indices_por_clasif['traslape_interno'].append(idx)
                    if resumen.get('traslape_interno_mismo_credito', 0) > 0:
                        counts['traslape_interno_mismo_credito'] += 1
                        indices_por_clasif['traslape_interno_mismo_credito'].append(idx)
                    else:
                        counts['traslape_interno_diferente_credito'] += 1
                        indices_por_clasif['traslape_interno_diferente_credito'].append(idx)
                elif resumen['traslape_relevante'] > 0:
                    counts['traslape_relevante'] += 1
                    indices_por_clasif['traslape_relevante'].append(idx)
                    if resumen.get('traslape_relevante_mismo_credito', 0) > 0:
                        counts['traslape_relevante_mismo_credito'] += 1
                        indices_por_clasif['traslape_relevante_mismo_credito'].append(idx)
                    else:
                        counts['traslape_relevante_diferente_credito'] += 1
                        indices_por_clasif['traslape_relevante_diferente_credito'].append(idx)
                elif resumen['sin_conflicto'] > 0:
                    counts['sin_conflicto'] += 1
                    indices_por_clasif['sin_conflicto'].append(idx)
                    if resumen.get('sin_conflicto_mismo_credito', 0) > 0:
                        counts['sin_conflicto_mismo_credito'] += 1
                        indices_por_clasif['sin_conflicto_mismo_credito'].append(idx)
                    else:
                        counts['sin_conflicto_diferente_credito'] += 1
                        indices_por_clasif['sin_conflicto_diferente_credito'].append(idx)
                else:
                    counts['sin_matches'] += 1
                    indices_por_clasif['sin_matches'].append(idx)
            except Exception:
                # Skip individual polygon errors without crashing the batch
                counts['sin_matches'] += 1
                indices_por_clasif['sin_matches'].append(idx)

            processed = i + 1
            _clasif_nuevos_state['processed'] = processed
            _clasif_nuevos_state['progress'] = int(processed / total * 100) if total > 0 else 100

        _clasif_nuevos_state['result'] = counts
        _clasif_nuevos_state['indices_por_clasif'] = indices_por_clasif
        _clasif_nuevos_state['status'] = 'done'
        _clasif_nuevos_state['progress'] = 100
    except Exception as e:
        _clasif_nuevos_state['status'] = 'error'
        _clasif_nuevos_state['error'] = str(e)


@app.route('/api/analizador/clasificacion-nuevos/iniciar', methods=['POST'])
@login_required
def api_clasificacion_nuevos_iniciar():
    global _clasif_nuevos_state
    if shp_cache.validacion is None or shp_cache.mega is None:
        return jsonify({'error': 'Shapefiles no cargados'}), 500

    if _clasif_nuevos_state['status'] == 'running':
        return jsonify({'error': 'Cálculo en progreso'}), 409

    if _clasif_nuevos_state['status'] == 'done':
        return jsonify({
            'status': 'done',
            'progress': 100,
            'processed': _clasif_nuevos_state['processed'],
            'total': _clasif_nuevos_state['total'],
            'result': _clasif_nuevos_state['result'],
        }), 200

    # Reset state and start background thread
    _clasif_nuevos_state['status'] = 'running'
    _clasif_nuevos_state['progress'] = 0
    _clasif_nuevos_state['processed'] = 0
    _clasif_nuevos_state['result'] = None
    _clasif_nuevos_state['indices_por_clasif'] = None
    _clasif_nuevos_state['error'] = None

    # Determine total upfront for the response (use cache if available)
    if _indices_filtrados_cache is not None:
        total = len(_indices_filtrados_cache['nuevos'])
    else:
        mega_ids = set(shp_cache.mega['ID_POLIGON'].astype(str).str.strip())
        mask_nuevos = ~shp_cache.validacion['ID_POLIGON'].astype(str).str.strip().isin(mega_ids)
        total = int(mask_nuevos.sum())
    _clasif_nuevos_state['total'] = total

    threading.Thread(target=_run_clasificacion_nuevos, daemon=True).start()
    return jsonify({'message': 'Cálculo iniciado', 'total': total}), 202


@app.route('/api/analizador/clasificacion-nuevos/estado')
@login_required
def api_clasificacion_nuevos_estado():
    state = _clasif_nuevos_state
    response = {
        'status': state['status'],
        'progress': state['progress'],
        'processed': state['processed'],
        'total': state['total'],
    }
    if state['status'] == 'done':
        response['result'] = state['result']
    if state['status'] == 'error':
        response['error'] = state['error']
    return jsonify(response)


@app.route('/api/analizador/clasificacion-nuevos/indices')
@login_required
def api_clasificacion_nuevos_indices():
    clasif = request.args.get('clasif')
    subfiltro = request.args.get('subfiltro')
    valid_clasifs = ['duplicado', 'traslape_interno', 'traslape_relevante', 'sin_conflicto', 'sin_matches']
    if clasif not in valid_clasifs:
        return jsonify({'error': 'Clasificación inválida. Use: ' + ', '.join(valid_clasifs)}), 400
    if _clasif_nuevos_state['status'] != 'done' or _clasif_nuevos_state['indices_por_clasif'] is None:
        return jsonify({'error': 'Clasificación no completada aún'}), 409

    indices_key = clasif
    # subfiltro applies to all categories that have mismo/diferente breakdown
    categories_with_subfiltro = ['duplicado', 'traslape_interno', 'traslape_relevante', 'sin_conflicto']
    if clasif in categories_with_subfiltro:
        valid_subfiltros = [None, '', 'mismo_credito', 'diferente_credito']
        if subfiltro not in valid_subfiltros:
            return jsonify({'error': f'Subfiltro inválido para {clasif}. Use: mismo_credito o diferente_credito'}), 400
        if subfiltro == 'mismo_credito':
            indices_key = clasif + '_mismo_credito'
        elif subfiltro == 'diferente_credito':
            indices_key = clasif + '_diferente_credito'
    elif subfiltro not in (None, ''):
        return jsonify({'error': 'El parámetro subfiltro no aplica para esta clasificación'}), 400

    indices = _clasif_nuevos_state['indices_por_clasif'].get(indices_key, [])
    response = {'clasif': clasif, 'indices': indices, 'total': len(indices)}
    if subfiltro not in (None, ''):
        response['subfiltro'] = subfiltro
    return jsonify(response)


# ─────────────────────────────────────────────────────────────────────────────
# Validación 15K — API endpoints
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/validacion-15k/guardar', methods=['POST'])
@login_required
def api_validacion_15k_guardar():
    """Save or update the validation status for a single 15K polygon (upsert)."""
    if shp_cache.validacion is None:
        return jsonify({'error': 'Shapefiles no cargados'}), 500

    data = request.get_json(force=True, silent=True) or {}

    # Validate idx
    idx = data.get('idx')
    if idx is None:
        return jsonify({'error': 'idx es requerido'}), 400
    try:
        idx = int(idx)
    except (TypeError, ValueError):
        return jsonify({'error': 'idx debe ser un entero'}), 400
    if idx < 0 or idx >= len(shp_cache.validacion):
        return jsonify({'error': f'idx fuera de rango (0-{len(shp_cache.validacion)-1})'}), 400

    # Validate estatus
    estatus = data.get('estatus')
    if estatus not in ('nuevo', 'encima'):
        return jsonify({'error': "estatus debe ser 'nuevo' o 'encima'"}), 400

    # Validate encima requirements
    id_poligon_historico = data.get('id_poligon_historico')
    mega_idx = data.get('mega_idx')
    overlap_pct = data.get('overlap_pct')

    if estatus == 'encima':
        if not id_poligon_historico:
            return jsonify({'error': "id_poligon_historico es requerido cuando estatus es 'encima'"}), 400
        if mega_idx is None:
            return jsonify({'error': "mega_idx es requerido cuando estatus es 'encima'"}), 400
    else:
        # estatus == 'nuevo': nullify linked fields
        id_poligon_historico = None
        mega_idx = None
        overlap_pct = None

    # Auto-populate from shp_cache.validacion
    vrow = shp_cache.validacion.iloc[idx]
    id_poligon_validacion = str(vrow.get('ID_POLIGON', '') or '')
    id_credito_validacion = str(vrow.get('ID_CREDITO', '') or '')
    nombre_zip = str(vrow.get('NOMBRE_ZIP', '') or '')

    from datetime import datetime, timezone
    fecha_validacion = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')

    conn = get_db_connection()
    try:
        cursor = conn.execute(
            '''INSERT OR REPLACE INTO validacion_15k
               (idx, id_poligon_validacion, id_credito_validacion, nombre_zip,
                estatus, id_poligon_historico, mega_idx, overlap_pct,
                fecha_validacion, validado_por)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'usuario')''',
            (idx, id_poligon_validacion, id_credito_validacion, nombre_zip,
             estatus, id_poligon_historico, mega_idx, overlap_pct,
             fecha_validacion)
        )
        conn.commit()
        val_id = cursor.lastrowid
    finally:
        conn.close()

    return jsonify({'success': True, 'val_id': val_id, 'estatus': estatus})


@app.route('/api/validacion-15k/progreso')
@login_required
def api_validacion_15k_progreso():
    """Return overall validation progress stats."""
    if shp_cache.validacion is None:
        return jsonify({'error': 'Shapefiles no cargados'}), 500

    total = len(shp_cache.validacion)
    conn = get_db_connection()
    try:
        rows = conn.execute(
            "SELECT estatus, COUNT(*) as cnt FROM validacion_15k GROUP BY estatus"
        ).fetchall()
    finally:
        conn.close()

    counts = {row['estatus']: row['cnt'] for row in rows}
    nuevos = counts.get('nuevo', 0)
    encima = counts.get('encima', 0)
    validados = nuevos + encima
    pendientes = total - validados
    porcentaje = round((validados / total * 100), 1) if total > 0 else 0.0

    return jsonify({
        'total': total,
        'pendientes': pendientes,
        'nuevos': nuevos,
        'encima': encima,
        'porcentaje': porcentaje
    })


@app.route('/api/validacion-15k/estado/<int:idx>')
@login_required
def api_validacion_15k_estado(idx):
    """Return the saved validation status for a specific polygon index."""
    if shp_cache.validacion is None:
        return jsonify({'error': 'Shapefiles no cargados'}), 500
    if idx < 0 or idx >= len(shp_cache.validacion):
        return jsonify({'error': f'idx fuera de rango (0-{len(shp_cache.validacion)-1})'}), 400

    guardado = obtener_guardado_validacion_15k(idx)
    return jsonify({'idx': idx, **guardado})


@app.route('/api/validacion-15k/poligono/<int:idx>')
@login_required
def api_validacion_15k_poligono(idx):
    """Return enhanced polygon detail with overlap analysis, saved status, and auto-suggestion."""
    try:
        data = calcular_traslapes(idx)
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    validacion = obtener_guardado_validacion_15k(idx)

    # Auto-suggestion based on overlap analysis
    matches = data['matches']
    best = matches[0] if matches else None  # already sorted by overlap_pct desc
    if best and best['overlap_pct'] >= 50:
        sugerencia = {
            'estatus': 'encima',
            'id_poligon_historico': best['id_poligon'],
            'mega_idx': best['mega_index'],
            'overlap_pct': best['overlap_pct'],
            'razon': f"Traslape >= 50% con polígono histórico {best['id_poligon']}"
        }
    else:
        sugerencia = {
            'estatus': 'nuevo',
            'id_poligon_historico': None,
            'mega_idx': None,
            'overlap_pct': None,
            'razon': 'Ningún traslape >= 50% con polígonos históricos'
        }

    try:
        propuesta_chapingo = generar_propuesta_chapingo_nuevo(idx, analisis_mega=data)
    except Exception as e:
        propuesta_chapingo = {
            'estatus_chapingo_propuesto': None,
            'id_poligono_unico_propuesto': None,
            'superficie_chapingo_propuesta': None,
            'comentario_chapingo_propuesto': f'No fue posible generar propuesta automatica: {str(e)}',
        }

    return jsonify({
        'index': idx,
        'total': len(shp_cache.validacion),
        'poligono': data['poligono'],
        'matches': data['matches'],
        'match_features': data['match_features'],
        'resumen': data['resumen'],
        'validacion': validacion,
        'sugerencia': sugerencia,
        'propuesta_chapingo': propuesta_chapingo
    })


@app.route('/api/validacion-15k/siguiente-pendiente')
@login_required
def api_validacion_15k_siguiente_pendiente():
    """Return the next unvalidated polygon index, optionally starting from ?desde=<idx>."""
    if shp_cache.validacion is None:
        return jsonify({'error': 'Shapefiles no cargados'}), 500

    try:
        desde = int(request.args.get('desde', 0))
    except (TypeError, ValueError):
        desde = 0
    desde = max(0, desde)

    total = len(shp_cache.validacion)

    # Fetch all validated indices (estatus != 'pendiente')
    conn = get_db_connection()
    try:
        rows = conn.execute(
            "SELECT idx FROM validacion_15k WHERE estatus != 'pendiente'"
        ).fetchall()
    finally:
        conn.close()

    validados = {row['idx'] for row in rows}
    total_pendientes = total - len(validados)

    # Search from `desde` to end, then wrap around 0 to `desde`
    for i in list(range(desde, total)) + list(range(0, desde)):
        if i not in validados:
            return jsonify({'idx': i, 'total_pendientes': total_pendientes})

    return jsonify({'idx': None, 'total_pendientes': 0, 'mensaje': 'Todos los polígonos han sido validados'})


@app.route('/api/validacion-15k/buscar')
@login_required
def api_validacion_15k_buscar():
    """Search shp_cache.validacion by ID_POLIGON or ID_CREDITO, including saved estatus."""
    if shp_cache.validacion is None:
        return jsonify({'error': 'Shapefiles no cargados'}), 500

    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({'resultados': [], 'total': 0})

    # Fetch all saved statuses in one query
    conn = get_db_connection()
    try:
        rows = conn.execute('SELECT idx, estatus FROM validacion_15k').fetchall()
    finally:
        conn.close()
    estatus_map = {row['idx']: row['estatus'] for row in rows}

    resultados = []
    q_lower = q.lower()
    for i, row in shp_cache.validacion.iterrows():
        if (q_lower in str(row.get('ID_CREDITO', '')).lower() or
                q_lower in str(row.get('ID_POLIGON', '')).lower()):
            resultados.append({
                'index': int(i),
                'id_poligon': str(row.get('ID_POLIGON', '')),
                'id_credito': str(row.get('ID_CREDITO', '')),
                'estatus': estatus_map.get(int(i), 'pendiente')
            })
        if len(resultados) >= 50:
            break

    return jsonify({'resultados': resultados, 'total': len(resultados)})


@app.route('/api/validacion-15k/exportar')
@login_required
def api_validacion_15k_exportar():
    """Generate and download an Excel file with all 15K validation results."""
    if shp_cache.validacion is None or shp_cache.mega is None:
        return jsonify({'error': 'Shapefiles no cargados'}), 500

    # Fetch all saved validation rows from DB
    conn = get_db_connection()
    try:
        db_rows = conn.execute('SELECT * FROM validacion_15k').fetchall()
    finally:
        conn.close()

    # Build a lookup dict: idx -> db row
    db_map = {row['idx']: row for row in db_rows}

    # Determine HIST_ columns from shp_cache.mega (all non-geometry columns)
    mega_data_cols = [c for c in shp_cache.mega.columns if c != 'geometry']

    # Build one record per polygon in shp_cache.validacion
    records = []
    for i in range(len(shp_cache.validacion)):
        vrow = shp_cache.validacion.iloc[i]
        db_row = db_map.get(i)

        if db_row is not None:
            estatus = db_row['estatus']
            id_poligon_validacion = db_row['id_poligon_validacion']
            id_credito_validacion = db_row['id_credito_validacion']
            nombre_zip = db_row['nombre_zip']
            id_poligon_historico = db_row['id_poligon_historico']
            overlap_pct = db_row['overlap_pct']
            fecha_validacion = db_row['fecha_validacion']
            mega_idx = db_row['mega_idx']
            superficie_calculada = db_row['superficie_calculada']
        else:
            # pendiente — populate from validacion_gdf
            estatus = 'pendiente'
            id_poligon_validacion = str(vrow.get('ID_POLIGON', '') or '')
            id_credito_validacion = str(vrow.get('ID_CREDITO', '') or '')
            nombre_zip = str(vrow.get('NOMBRE_ZIP', '') or '')
            id_poligon_historico = None
            overlap_pct = None
            fecha_validacion = None
            mega_idx = None
            superficie_calculada = None

        record = {
            'IDX': i,
            'ID_POLIGON_VALIDACION': id_poligon_validacion,
            'ID_CREDITO_VALIDACION': id_credito_validacion,
            'NOMBRE_ZIP': nombre_zip,
            'ESTATUS': estatus,
            'ID_POLIGON_HISTORICO': id_poligon_historico,
            'OVERLAP_PCT': overlap_pct,
            'FECHA_VALIDACION': fecha_validacion,
            'SUPERFICIE_CALCULADA': superficie_calculada,
        }

        # Add HIST_ columns from shp_cache.mega when estatus='encima'
        if estatus == 'encima' and mega_idx is not None:
            try:
                mrow = shp_cache.mega.iloc[int(mega_idx)]
                for col in mega_data_cols:
                    record[f'HIST_{col}'] = mrow.get(col, None)
            except (IndexError, TypeError):
                for col in mega_data_cols:
                    record[f'HIST_{col}'] = None
        else:
            for col in mega_data_cols:
                record[f'HIST_{col}'] = None

        records.append(record)

    # Sort by IDX ascending (already in order, but be explicit)
    records.sort(key=lambda r: r['IDX'])

    df = pd.DataFrame(records)

    # Generate Excel in memory
    excel_file = io.BytesIO()
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Validacion_15K')

        workbook = writer.book
        worksheet = writer.sheets['Validacion_15K']

        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    excel_file.seek(0)
    filename = f'validacion_15k_resultados_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/analizador')
@login_required
def analizador():
    return render_template('analizador.html')


shp_cache.preload_all()  # Preload in production for gunicorn preload_app

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
